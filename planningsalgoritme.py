# andere pauzes voor minderjarigen, maar die kloppen mogelijks niet lol (ook lange pauzes van pauzevlinders zijn beter)
# nieuw werkblad analyse
# zelfde versie als 3.5 maar pauzevlinders zijn ook volgens volgorde uit gekozen nummertje
#betere verdeling 3 uur blokken, maar te veel 6 uur bij zelfde attractie & 1+3 logica voor 9u30 ipv 3+1 & 2+2 logica voor 4 uur opt einde

#uitschakelen attracties op bepaalde uren lijkt te werken!
#samenvoegen attracties per uur werkttttt!!! Kleine bug is er uit gehaald
#hele dag bij attractie werkt
# probleem met twee


import streamlit as st
import random
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

# -----------------------------
# Datum
# -----------------------------
vandaag = datetime.date.today().strftime("%d-%m-%Y")

# -----------------------------
# Excelbestand uploaden
# -----------------------------
uploaded_file = st.file_uploader("Upload Excel bestand", type=["xlsx"])

if not uploaded_file:
    st.warning("Upload eerst het Excelbestand met de gegevens om verder te gaan.")
    st.stop()

wb = load_workbook(BytesIO(uploaded_file.read()), data_only=True)

ws = wb["Input"]

# -----------------------------

# Hulpfuncties
# -----------------------------
def max_consecutive_hours(urenlijst):
    if not urenlijst:
        return 0
    urenlijst = sorted(set(urenlijst))
    maxr = huidig = 1
    for i in range(1, len(urenlijst)):
        huidig = huidig + 1 if urenlijst[i] == urenlijst[i-1] + 1 else 1
        maxr = max(maxr, huidig)
    return maxr

def partition_run_lengths(L):
    """Flexibele blokken: prioritair 3 uur, dan 2,4,1 om shift te vullen."""
    blocks = [3,2,4,1]
    dp = [(10**9, [])]*(L+1)
    dp[0] = (0, [])
    for i in range(1, L+1):
        best = (10**9, [])
        for b in blocks:
            if i-b < 0:
                continue
            prev_ones, prev_blocks = dp[i-b]
            new_blocks = prev_blocks + [b]
            ones = prev_ones + (1 if b==1 else 0)
            if ones < best[0]:
                best = (ones, new_blocks)
        dp[i] = best
    return dp[L][1]

def contiguous_runs(sorted_hours):
    runs=[]
    if not sorted_hours:
        return runs
    run=[sorted_hours[0]]
    for h in sorted_hours[1:]:
        if h==run[-1]+1:
            run.append(h)
        else:
            runs.append(run)
            run=[h]
    runs.append(run)
    return runs

# Helpers die in meerdere delen gebruikt worden
def normalize_attr(name):
    """Normaliseer attractienaam zodat 'X 2' telt als 'X'; trim & lower-case voor vergelijking."""
    if not name:
        return ""
    s = str(name).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        s = parts[0]
    return s.strip().lower()

def parse_header_uur(header):
    """Map headertekst (bv. '14u', '14:00', '14:30') naar het hele uur (14)."""
    if not header:
        return None
    s = str(header).strip()
    try:
        if "u" in s:
            return int(s.split("u")[0])
        if ":" in s:
            uur, _min = s.split(":")
            return int(uur)
        return int(s)
    except:
        return None

# -----------------------------
# Studenten inlezen
# -----------------------------
studenten=[]
for rij in range(2,500):
    naam = ws.cell(rij,12).value
    if not naam:
        continue
    uren_beschikbaar=[10+(kol-3) for kol in range(3,12) if ws.cell(rij,kol).value in [1,True,"WAAR","X"]]
    attracties=[ws.cell(1,kol).value for kol in range(14,32) if ws.cell(rij,kol).value in [1,True,"WAAR","X"]]
    try:
        aantal_attracties=int(ws[f'AG{rij}'].value) if ws[f'AG{rij}'].value else len(attracties)
    except:
        aantal_attracties=len(attracties)
    studenten.append({
        "naam": naam,
        "uren_beschikbaar": sorted(uren_beschikbaar),
        "attracties":[a for a in attracties if a],
        "aantal_attracties":aantal_attracties,
        "is_pauzevlinder":False,
        "pv_number":None,
        "assigned_attracties":set(),
        "assigned_hours":[]
    })


# Nieuwe dictionary voor uren dat een attractie DICHT is
dichte_uren_per_attr = defaultdict(set)
# AJ t/m AR (kolom 36 t/m 44)
uur_kolommen = list(range(36, 45)) 

for rij in range(24, 30): # Rij 24 t/m 29
    attr_naam_raw = ws.cell(rij, 45).value # Kolom AS
    if attr_naam_raw:
        # Belangrijk: Gebruik normalize_attr voor een eerlijke vergelijking [3]
        attr_naam = normalize_attr(attr_naam_raw)
        for col_idx in uur_kolommen:
            val = ws.cell(rij, col_idx).value
            if val in [1, True, "WAAR", "X"]:
                uur = 10 + (col_idx - 36)
                dichte_uren_per_attr[attr_naam].add(uur)

# -----------------------------
# Samenvoeg-attracties (per uur)
# -----------------------------


# In DEEL 1 bij "Samenvoeg-attracties (per uur)"
uur_samenvoegingen = defaultdict(list)
uur_kolommen = list(range(36, 45)) 

for rij in range(14, 22):  # Rij 14 t/m 21 voor samenvoegingen
    # Lees de groep (AS, AT, AU)
    groep = []
    for col in range(45, 48): 
        val = ws.cell(rij, col).value
        if val: groep.append(str(val).strip())
    
    if len(groep) > 1:
        # Check per uur of de samenvoeging actief is (AJ t/m AR)
        for col_idx in uur_kolommen:
            if ws.cell(rij, col_idx).value in [1, True, "WAAR", "X"]:
                uur = 10 + (col_idx - 36)
                uur_samenvoegingen[uur].append(groep)


# -----------------------------
# Alle mogelijke samengevoegde attracties (namen)
# -----------------------------

samengevoegde_attracties = set()

for groepen in uur_samenvoegingen.values():
    for groep in groepen:
        samengevoegde_attracties.add(" + ".join(groep))



# -----------------------------
# Voeg samengestelde attracties toe aan individuele studenten
# -----------------------------
for s in studenten:
    huidige = set(s["attracties"])
    for sameng in samengevoegde_attracties:
        onderdelen = [a.strip() for a in sameng.split("+")]
        if all(o in huidige for o in onderdelen):
            s["attracties"].append(sameng)  # voeg de samengestelde attractie toe




# -----------------------------
# Openingsuren
# -----------------------------
open_uren=[int(str(ws.cell(1,kol).value).replace("u","").strip()) for kol in range(36,45) if ws.cell(2,kol).value in [1,True,"WAAR","X"]]
if not open_uren:
    open_uren=list(range(10,19))
open_uren=sorted(set(open_uren))



# -----------------------------
# Sorteervolgorde studenten
# Eerst op aantal attracties,
# daarna op vaste tie-break regel uit BU2
# -----------------------------
bu2_waarde = ws["BU2"].value
try:
    tie_break_mode = int(bu2_waarde)
except:
    tie_break_mode = 1

if tie_break_mode not in [1, 2, 3, 4, 5]:
    tie_break_mode = 1


def naam_tie_break_key(naam_raw):
    naam = str(naam_raw).strip().lower()

    if tie_break_mode == 1:
        # gewone alfabetische volgorde
        return naam

    elif tie_break_mode == 2:
        # omgekeerde alfabetische volgorde
        return "".join(chr(255 - ord(c)) for c in naam)

    elif tie_break_mode == 3:
        # eerst op aantal letters, daarna alfabetisch
        return (len(naam), naam)

    elif tie_break_mode == 4:
        # alfabetisch op basis van laatste letters
        return naam[::-1]

    elif tie_break_mode == 5:
        # omgekeerde van mode 4
        return "".join(chr(255 - ord(c)) for c in naam[::-1])

    return naam




# -----------------------------

# Pauzevlinders
# -----------------------------
pauzevlinder_namen=[ws[f'BN{rij}'].value for rij in range(4,11) if ws[f'BN{rij}'].value]

def compute_pauze_hours(open_uren):
    if 10 in open_uren and 18 in open_uren:
        return [h for h in open_uren if 12 <= h <= 16]
    elif 10 in open_uren and 17 in open_uren:
        return [h for h in open_uren if 12 <= h <= 16]
    elif 12 in open_uren and 18 in open_uren:
        return [h for h in open_uren if 14 <= h <= 17]
    elif 14 in open_uren and 18 in open_uren:
        return [h for h in open_uren if 16 <= h <= 17]
    else:
        return list(open_uren)

required_pauze_hours=compute_pauze_hours(open_uren)

for idx,pvnaam in enumerate(pauzevlinder_namen,start=1):
    for s in studenten:
        if s["naam"]==pvnaam:
            s["is_pauzevlinder"]=True
            s["pv_number"]=idx
            s["uren_beschikbaar"]=[u for u in s["uren_beschikbaar"] if u not in required_pauze_hours]
            break

# Maak 'selected' lijst van pauzevlinders (dicts met naam en attracties)
selected = [s for s in studenten if s.get("is_pauzevlinder")]
selected = sorted(selected, key=lambda s: naam_tie_break_key(s["naam"]))

# -----------------------------
# Attracties & aantallen (raw)
# -----------------------------
aantallen_raw = {}
attracties_te_plannen = []
for kol in range(47, 65):  # AU-BL
    naam = ws.cell(1, kol).value
    if naam:
        try:
            aantal = int(ws.cell(2, kol).value)
        except:
            aantal = 0
        aantallen_raw[naam] = max(0, min(2, aantal))
        if aantallen_raw[naam] >= 1:
            attracties_te_plannen.append(naam)

# Priority order for second spots (column BA, rows 5-11)
second_priority_order = [
    ws["BA" + str(rij)].value for rij in range(5, 12)
    if ws["BA" + str(rij)].value
]


# -----------------------------
# Attractielijst uitbreiden met samengevoegde attracties (globaal)
# -----------------------------

for nieuwe in samengevoegde_attracties:
    if nieuwe not in attracties_te_plannen:
        attracties_te_plannen.append(nieuwe)
    aantallen_raw[nieuwe] = 1


# -----------------------------
# Actieve attracties per uur (ivm samenvoegingen)
# -----------------------------

actieve_attracties_per_uur = {}
# Gebruik de raw aantallen als basis
aantallen = {uur: {a: aantallen_raw.get(a, 1) for a in attracties_te_plannen} for uur in open_uren}

for uur in open_uren:
    actief = set()
    # 1. Voeg eerst alle individuele attracties toe die NIET dicht zijn
    for a in attracties_te_plannen:
        if " + " in a: continue # Sla samengevoegde namen hier nog even over
        
        if uur in dichte_uren_per_attr.get(normalize_attr(a), set()):
            aantallen[uur][a] = 0
        else:
            actief.add(a)

    # 2. Verwerk de samenvoegingen voor dit specifieke uur
    huidige_groepen = uur_samenvoegingen.get(uur, [])
    for groep in huidige_groepen:
        samengevoegde_naam = " + ".join(groep)
        
        # Voeg de samengevoegde attractie toe aan de planning
        actief.add(samengevoegde_naam)
        aantallen[uur][samengevoegde_naam] = 1
        
        # VERWIJDER de onderdelen uit de actieve lijst (voorkomt dubbele telling)
        for onderdeel in groep:
            if onderdeel in actief:
                actief.remove(onderdeel)
            aantallen[uur][onderdeel] = 0

    actieve_attracties_per_uur[uur] = actief



### -----------------------------
### Compute aantallen per hour + red spots (GEÏNTEGREERD)
### -----------------------------
red_spots = {uur: set() for uur in open_uren}          
second_spot_blocked = {uur: set() for uur in open_uren}  

for uur in open_uren:
    # 1. Hoeveel studenten zijn er dit uur echt beschikbaar? [1]
    student_count = sum(
        1 for s in studenten
        if uur in s["uren_beschikbaar"] and not (
            s["is_pauzevlinder"] and uur in required_pauze_hours
        )
    )
    
    # 2. Hoeveel attracties moeten dit uur minimaal 1 persoon hebben? [1]
    # We kijken naar de actieve lijst van dat uur (rekening houdend met uitschakelingen/samenvoegingen)
    base_spots = sum(1 for a in actieve_attracties_per_uur[uur] if aantallen[uur].get(a, 0) >= 1)
    
    # 3. Bereken het overschot
    extra_spots = student_count - base_spots

    # 4. Verdeel de tweede plekken op basis van de prioriteitslijst uit Excel (BA5:BA11) [2]
    for attr in second_priority_order:
        # Check of de attractie dit uur actief is én of hij normaal 2 personen nodig heeft [2, 3]
        if attr in actieve_attracties_per_uur[uur] and aantallen_raw.get(attr) == 2:
            if extra_spots > 0:
                # Er is nog een student over voor een tweede plek
                aantallen[uur][attr] = 2
                extra_spots -= 1
            else:
                # Geen studenten meer over? Blokkeer de tweede plek voor dit uur
                second_spot_blocked[uur].add(attr)
                aantallen[uur][attr] = 1  # Forceer het aantal voor dit uur naar 1


# -----------------------------
# Red spots for samengestelde attracties
# -----------------------------

for uur in open_uren:
    # Groepen die dit uur samengevoegd zijn
    groepen = uur_samenvoegingen.get(uur, [])

    # Samengestelde attracties die DIT uur actief zijn
    samengestelde = set(" + ".join(g) for g in groepen)

    # Losse attracties die in een samenvoeging zitten
    losse_in_samenvoeging = set(a for g in groepen for a in g)

    # 1️⃣ Samenvoeging actief → losse attracties verbieden
    for attr in losse_in_samenvoeging:
        red_spots[uur].add(attr)

    # 2️⃣ Samenvoeging NIET actief → samenvoeging verbieden
    for samengestelde_attr in samengevoegde_attracties:
        if samengestelde_attr not in samengestelde:
            red_spots[uur].add(samengestelde_attr)


# -----------------------------
# Studenten die effectief inzetbaar zijn
# -----------------------------
studenten_workend = [
    s for s in studenten if any(u in open_uren for u in s["uren_beschikbaar"])
]


# -----------------------------
# Blacklist van attracties per student (BB16:BG79)
# -----------------------------
student_blacklist = defaultdict(set)

for rij in range(16, 80):  # BB16 t/m BG79
    naam = ws[f'BB{rij}'].value
    if not naam:
        continue
    naam = str(naam).strip()
    # attracties in BC t/m BG
    for col in range(54, 60):  # BC=54, BD=55, ..., BG=59
        attr = ws.cell(rij, col).value
        if attr:
            student_blacklist[naam].add(str(attr).strip().lower())


# Sorteer attracties op "kritieke score" (hoeveel studenten ze kunnen doen)
def kritieke_score(attr, studenten_list):
    return sum(1 for s in studenten_list if attr in s["attracties"])

attracties_te_plannen.sort(key=lambda a: kritieke_score(a, studenten_workend))



# -----------------------------
# Assign per student
# -----------------------------
assigned_map = defaultdict(list)  # (uur, attr) -> list of student-names
per_hour_assigned_counts = {uur: {a: 0 for a in attracties_te_plannen} for uur in open_uren}
extra_assignments = defaultdict(list)

MAX_CONSEC = 4
MAX_PER_STUDENT_ATTR = 6


# -----------------------------
# Vaste dag-attracties (BG–BI)
# -----------------------------

vaste_plaatsingen = []  # lijst van dicts: {naam, attractie}

for rij in range(5, 9):  # BG5 t.e.m. BI26
    if ws[f"BG{rij}"].value in [1, True, "WAAR", "X"]:
        naam = ws[f"BH{rij}"].value
        attractie = ws[f"BI{rij}"].value
        if naam and attractie:
            vaste_plaatsingen.append({
                "naam": str(naam).strip(),
                "attractie": str(attractie).strip()
            })


# -----------------------------
# Vaste plaatsingen toepassen
# -----------------------------

for vp in vaste_plaatsingen:
    student = next((s for s in studenten if s["naam"] == vp["naam"]), None)
    if not student:
        continue

    attr = vp["attractie"]

    # effectieve werkuren van deze student
    uren = [
        u for u in student["uren_beschikbaar"]
        if u in open_uren
        and not (student["is_pauzevlinder"] and u in required_pauze_hours)
    ]

    for uur in uren:
        # attractie moet dit uur actief zijn
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            continue

        # rode attracties overslaan
        if attr in red_spots.get(uur, set()):
            continue

        # capaciteit check
        max_spots = aantallen[uur].get(attr, 1)
        if attr in second_spot_blocked.get(uur, set()):
            max_spots = 1

        if per_hour_assigned_counts[uur][attr] >= max_spots:
            continue

        # plaats student
        assigned_map[(uur, attr)].append(student["naam"])
        per_hour_assigned_counts[uur][attr] += 1
        student["assigned_hours"].append(uur)
        student["assigned_attracties"].add(attr)

    # student mag niet meer door de normale planner
    student["uren_beschikbaar"] = []





def student_tie_break_key(student):
    return naam_tie_break_key(student["naam"])

studenten_sorted = sorted(
    studenten_workend,
    key=lambda s: (s["aantal_attracties"], student_tie_break_key(s))
)


# -----------------------------
# Voorbereiden: expand naar posities per uur
# -----------------------------
positions_per_hour = {uur: [] for uur in open_uren}
for uur in open_uren:
    for attr in actieve_attracties_per_uur[uur]:
        max_pos = aantallen[uur].get(attr, 1)
        for pos in range(1, max_pos+1):
            # sla rode posities over
            if attr in second_spot_blocked[uur] and pos == 2:
                continue
            positions_per_hour[uur].append((attr, pos))
# -----------------------------
# occupied_positions vullen op basis van bestaande assigned_map
# -----------------------------
occupied_positions = {uur: {} for uur in open_uren}

for (uur, attr), namen in assigned_map.items():
    for idx, naam in enumerate(namen, start=1):
        occupied_positions[uur][(attr, idx)] = naam


# -----------------------------
# Hulpfunctie: kan blok geplaatst worden?
# -----------------------------
def can_place_block(student, block_hours, attr):
    for h in block_hours:
        # check of attractie beschikbaar is in dit uur
        if (attr, 1) not in positions_per_hour[h] and (attr, 2) not in positions_per_hour[h]:
            return False
        # alle posities bezet?
        taken1 = (attr,1) in occupied_positions[h]
        taken2 = (attr,2) in occupied_positions[h]
        if taken1 and taken2:
            return False
    return True

# -----------------------------
# Plaats blok
# -----------------------------
def place_block(student, block_hours, attr):
    for h in block_hours:
        # kies positie: eerst pos1, anders pos2
        if (attr,1) in positions_per_hour[h] and (attr,1) not in occupied_positions[h]:
            pos = 1
        else:
            pos = 2
        occupied_positions[h][(attr,pos)] = student["naam"]
        assigned_map[(h, attr)].append(student["naam"])
        student["assigned_hours"].append(h)
    student["assigned_attracties"].add(attr)


# =============================
# FLEXIBELE BLOKKEN & PLAATSLOGICA
# =============================

def student_kan_attr(student, attr):
    if " + " not in attr:
        # check blacklist
        if attr.lower() in student_blacklist.get(student["naam"], set()):
            return False
        return attr in student["attracties"]
    onderdelen = [a.strip() for a in attr.split("+")]
    # check elk onderdeel tegen blacklist
    for o in onderdelen:
        if o.lower() in student_blacklist.get(student["naam"], set()):
            return False
    return all(o in student["attracties"] for o in onderdelen)


def _max_spots_for(attr, uur):
    """Houd rekening met red_spots: 2e plek dicht als het rood is."""
    max_spots = aantallen[uur].get(attr, 1)
    if attr in second_spot_blocked.get(uur, set()):
        max_spots = 1
    return max_spots

def _has_capacity(attr, uur):
    if attr in red_spots.get(uur, set()):
        return False
    return per_hour_assigned_counts[uur][attr] < _max_spots_for(attr, uur)


def _try_place_block_on_attr(student, block_hours, attr):
    """Check capaciteit in alle uren en plaats dan in één keer.
    Regels:
    - max 6 uur totaal per attractie per dag
    - max 4 aaneengesloten uren op dezelfde attractie
    """
    # Capaciteit check
    for h in block_hours:
        if not _has_capacity(attr, h):
            return False

    # Verzamel alle uren waarop deze student al bij deze attractie staat
    uren_bij_attr = set()
    for h in student["assigned_hours"]:
        namen = assigned_map.get((h, attr), [])
        if student["naam"] in namen:
            uren_bij_attr.add(h)

    # Check max 6 unieke uren per attractie per dag
    nieuwe_uren = set(block_hours)
    totaal_uren = uren_bij_attr | nieuwe_uren
    if len(totaal_uren) > 6:
        return False

    # Check max 4 aaneengesloten uren op dezelfde attractie
    alle_uren_attr = sorted(totaal_uren)
    if max_consecutive_hours(alle_uren_attr) > 4:
        return False

    # Plaatsen
    for h in block_hours:
        assigned_map[(h, attr)].append(student["naam"])
        per_hour_assigned_counts[h][attr] += 1
        student["assigned_hours"].append(h)

    student["assigned_attracties"].add(attr)
    return True



def _try_place_block_any_attr(student, block_hours):
    """Probeer dit blok te plaatsen op eender welke attractie die student kan.
    Fairness-regel:
    - Studenten met veel mogelijke attracties moeten minder snel naar 5e/6e uur
      op dezelfde attractie gaan.
    - Studenten met weinig mogelijke attracties blijven soepeler behandeld.
    """

    def uren_bij_attr(student, attr):
        uren = set()
        for h in student["assigned_hours"]:
            namen = assigned_map.get((h, attr), [])
            if student["naam"] in namen:
                uren.add(h)
        return uren

    def candidate_score(attr):
        # Hoeveel studenten kunnen deze attractie? Lager = kritieker
        schaarste = sum(1 for s in studenten_workend if attr in s["attracties"])

        bestaande_uren = uren_bij_attr(student, attr)
        totaal_na_plaatsing = len(bestaande_uren | set(block_hours))
        reeds_gebruikt = attr in student["assigned_attracties"]

        # Hoe breed is deze student inzetbaar?
        # We nemen aantal_attracties als hoofdsignaal, met fallback op echte lijstlengte
        breedte_profiel = student.get("aantal_attracties", len(student.get("attracties", [])))

        # Fairness-straf:
        # - Studenten met veel attracties krijgen zware straf als ze naar uur 5/6
        #   op dezelfde attractie gaan.
        # - Studenten met weinig attracties krijgen weinig of geen straf.
        fairness_straf = 0

        if totaal_na_plaatsing > 4:
            if breedte_profiel >= 6:
                fairness_straf = 100
            elif breedte_profiel >= 5:
                fairness_straf = 60
            elif breedte_profiel >= 4:
                fairness_straf = 25
            else:
                fairness_straf = 0

        # Lichte voorkeur om eerst nieuwe attracties te gebruiken,
        # maar minder belangrijk dan fairness boven 4 uur.
        hergebruik_straf = 1 if reeds_gebruikt else 0

        # Eventueel nog mini-voorkeur voor attracties waar student nog 0 uur stond
        # en pas daarna voor attracties met al wat uren.
        huidige_uren_op_attr = len(bestaande_uren)

        return (
            fairness_straf,
            hergebruik_straf,
            huidige_uren_op_attr,
            schaarste,
            attr
        )

    candidate_attrs = [
        a for a in attracties_te_plannen
        if student_kan_attr(student, a)
    ]

    candidate_attrs.sort(key=candidate_score)

    for attr in candidate_attrs:
        if _try_place_block_on_attr(student, block_hours, attr):
            return True

    return False
    

def _place_block_with_fallback(student, hours_seq, preferred_sizes=None):
    """
    Probeer een reeks opeenvolgende uren te plaatsen.
    - Standaard: eerst 3, dan 2, dan 1.
    - Via preferred_sizes kan je lokaal een andere voorkeur afdwingen.
    - Als niets lukt aan het begin van de reeks, schuif 1 uur op (dat uur gaat voorlopig naar extra),
      en probeer verder; tweede pass zal het later alsnog proberen op te vullen.
    Retourneert: lijst 'unplaced' uren die (voorlopig) niet geplaatst raakten.
    """
    if not hours_seq:
        return []

    if preferred_sizes is None:
        preferred_sizes = [3, 2, 1]

    # Probeer blok aan de voorkant volgens voorkeur
    for size in preferred_sizes:
        if len(hours_seq) >= size:
            first_block = hours_seq[:size]
            if _try_place_block_any_attr(student, first_block):
                return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes)

    # Helemaal niks paste aan de voorkant: markeer eerste uur tijdelijk als 'unplaced' en schuif door
    return [hours_seq[0]] + _place_block_with_fallback(student, hours_seq[1:], preferred_sizes)


# -----------------------------
# Vinkjes uitlezen voor bloklogica
# AR = kolom 44, AS = kolom 45, rij 2
# -----------------------------
ar2_vinkje = ws.cell(2, 44).value
as2_vinkje = ws.cell(2, 45).value

laatste_blok_is_anderhalf_uur = ar2_vinkje in [1, True, "WAAR", "X"]
eerste_blok_is_anderhalf_uur = as2_vinkje in [1, True, "WAAR", "X"]

    
# -----------------------------
# Nieuwe assign_student
# -----------------------------


def assign_student(s):
    """
    Plaats één student in de planning volgens alle regels:
    - Alleen uren waar de student beschikbaar is én open_uren zijn.
    - Geen overlap met pauzevlinder-uren.
    - Alleen attracties die de student kan.
    - Standaard voorkeur: 3 uur, dan 2, dan 1.
    - Speciaal geval begin van de dag:
      * student met exact 4 effectieve werkuren
      * én AS2 aangevinkt
      * én run start op het eerste open uur
      => probeer expliciet 1 + 3
    - Speciaal geval einde van de dag:
      * student met exact 4 effectieve werkuren
      * én AR2 aangevinkt
      * én run eindigt op het laatste open uur
      => probeer expliciet 2 + 2
    - Blokken die niet passen, gaan voorlopig naar extra_assignments.
    """
    # Filter op effectieve inzetbare uren
    uren = sorted(u for u in s["uren_beschikbaar"] if u in open_uren)
    if s["is_pauzevlinder"]:
        uren = [u for u in uren if u not in required_pauze_hours]

    if not uren:
        return

    runs = contiguous_runs(uren)
    eerste_open_uur = min(open_uren) if open_uren else None
    laatste_open_uur = max(open_uren) if open_uren else None

    for run in runs:
        # -----------------------------
        # Speciaal geval einde van de dag:
        # bij AR2 aangevinkt willen we voor een run van exact 4 uren
        # die eindigt op het laatste open uur liever 2 + 2
        # -----------------------------
        if (
            laatste_blok_is_anderhalf_uur
            and len(run) == 4
            and laatste_open_uur is not None
            and run[-1] == laatste_open_uur
        ):
            eerste_blok = run[:2]
            tweede_blok = run[2:]

            if _try_place_block_any_attr(s, eerste_blok):
                if _try_place_block_any_attr(s, tweede_blok):
                    unplaced = []
                else:
                    # Eerste 2 uur zijn al geplaatst, rest valt terug op normale logica
                    unplaced = _place_block_with_fallback(s, tweede_blok)
            else:
                # Als 2+2 niet lukt, val volledig terug op normale logica
                unplaced = _place_block_with_fallback(s, run)

        # -----------------------------
        # Speciaal geval begin van de dag:
        # bij AS2 aangevinkt telt het eerste blok als 1,5 uur (9u30-11u),
        # dus voor een run van exact 4 uren die start op het eerste open uur
        # proberen we eerst expliciet 1 + 3
        # -----------------------------
        elif (
            eerste_blok_is_anderhalf_uur
            and len(run) == 4
            and eerste_open_uur is not None
            and run[0] == eerste_open_uur
        ):
            eerste_blok = [run[0]]
            rest_blok = run[1:]

            if _try_place_block_any_attr(s, eerste_blok):
                if _try_place_block_any_attr(s, rest_blok):
                    unplaced = []
                else:
                    # Eerste uur is al geplaatst, rest valt terug op normale logica
                    unplaced = _place_block_with_fallback(s, rest_blok)
            else:
                # Als 1+3 niet lukt, val volledig terug op normale logica
                unplaced = _place_block_with_fallback(s, run)

        else:
            # Normale logica
            unplaced = _place_block_with_fallback(s, run)

        for h in unplaced:
            extra_assignments[h].append(s["naam"])



for s in studenten_sorted:
    assign_student(s)

# -----------------------------
# Post-processing: lege plekken opvullen door doorschuiven
# -----------------------------

def doorschuif_leegplek(uur, attr, pos_idx, student_naam, stap, max_stappen=5):
    if stap > max_stappen:
        return False
    namen = assigned_map.get((uur, attr), [])
    naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
    if naam:
        return False

    kandidaten = []
    for b_attr in attracties_te_plannen:
        b_namen = assigned_map.get((uur, b_attr), [])
        for b_pos, b_naam in enumerate(b_namen):
            if not b_naam or b_naam == student_naam:
                continue
            cand_student = next((s for s in studenten_workend if s["naam"] == b_naam), None)
            if not cand_student:
                continue
            # Mag deze student de lege attractie doen?
            if attr not in cand_student["attracties"]:
                continue
            # Mag de extra de vrijgekomen plek doen?
            extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
            if not extra_student:
                continue
            if b_attr not in extra_student["attracties"]:
                continue
            # Score: zo min mogelijk 1-uursblokken creëren
            uren_cand = sorted([u for u in cand_student["assigned_hours"] if u != uur] + [uur])
            uren_extra = sorted(extra_student["assigned_hours"] + [uur])
            def count_1u_blokken(uren):
                if not uren:
                    return 0
                runs = contiguous_runs(uren)
                return sum(1 for r in runs if len(r) == 1)
            score = count_1u_blokken(uren_cand) + count_1u_blokken(uren_extra)
            kandidaten.append((score, b_attr, b_pos, b_naam, cand_student))
    kandidaten.sort()

    for score, b_attr, b_pos, b_naam, cand_student in kandidaten:
        extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
        if not extra_student:
            continue
        # Voer de swap uit
        assigned_map[(uur, b_attr)][b_pos] = student_naam
        extra_student["assigned_hours"].append(uur)
        extra_student["assigned_attracties"].add(b_attr)
        per_hour_assigned_counts[uur][b_attr] += 0  # netto gelijk
        assigned_map[(uur, attr)].insert(pos_idx-1, b_naam)
        assigned_map[(uur, attr)] = assigned_map[(uur, attr)][:aantallen[uur].get(attr, 1)]
        cand_student["assigned_hours"].remove(uur)
        cand_student["assigned_attracties"].discard(b_attr)
        cand_student["assigned_hours"].append(uur)
        cand_student["assigned_attracties"].add(attr)
        per_hour_assigned_counts[uur][attr] += 0  # netto gelijk
        # Check of alles klopt (geen dubbele, geen restricties overtreden)
        # (optioneel: extra checks toevoegen)
        return True
    return False

max_iterations = 5
for _ in range(max_iterations):
    changes_made = False
    for uur in open_uren:
        for attr in actieve_attracties_per_uur[uur]:
            max_pos = aantallen[uur].get(attr, 1)
            if attr in red_spots.get(uur, set()):
                max_pos = 1
            for pos_idx in range(1, max_pos+1):
                namen = assigned_map.get((uur, attr), [])
                naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
                if naam:
                    continue
                # Probeer voor alle extra's op dit uur
                extras_op_uur = list(extra_assignments[uur])  # kopie ivm mutatie
                for extra_naam in extras_op_uur:
                    extra_student = next((s for s in studenten_workend if s["naam"] == extra_naam), None)
                    if not extra_student:
                        continue
                    if attr in extra_student["attracties"]:
                        # Kan direct geplaatst worden, dus hoort niet bij dit scenario
                        continue
                    # Probeer doorschuiven
                    if doorschuif_leegplek(uur, attr, pos_idx, extra_naam, 1, max_iterations):
                        extra_assignments[uur].remove(extra_naam)
                        changes_made = True
                        break  # stop met deze plek, ga naar volgende lege plek
    if not changes_made:
        break



# -----------------------------
# Post-processing: wissel laatste blok van 2 of 3 uren
# als iemand 5 of 6 uur op 1 attractie staat
# -----------------------------

vaste_studenten = {vp["naam"] for vp in vaste_plaatsingen}

def get_student_by_name(naam):
    return next((s for s in studenten_workend if s["naam"] == naam), None)

def get_student_attr_on_hour(student_naam, uur):
    for attr in actieve_attracties_per_uur.get(uur, set()):
        if student_naam in assigned_map.get((uur, attr), []):
            return attr
    return None

def get_hours_on_attr(student, attr):
    uren = []
    for uur in sorted(set(student["assigned_hours"])):
        if student["naam"] in assigned_map.get((uur, attr), []):
            uren.append(uur)
    return sorted(uren)

def get_runs_on_attr(student, attr):
    uren = get_hours_on_attr(student, attr)
    return contiguous_runs(uren)

def count_attr_switches(student):
    uur_attr = []
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            uur_attr.append((uur, attr))

    if not uur_attr:
        return 0

    switches = 0
    prev_attr = uur_attr[0][1]
    for _, attr in uur_attr[1:]:
        if attr != prev_attr:
            switches += 1
        prev_attr = attr
    return switches

def remove_assignment(student, uur, attr):
    namen = assigned_map.get((uur, attr), [])
    if student["naam"] in namen:
        namen.remove(student["naam"])
    if uur in student["assigned_hours"]:
        student["assigned_hours"].remove(uur)

def add_assignment(student, uur, attr):
    assigned_map[(uur, attr)].append(student["naam"])
    student["assigned_hours"].append(uur)
    student["assigned_attracties"].add(attr)

def rebuild_student_attrs(student):
    attrs = set()
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            attrs.add(attr)
    student["assigned_attracties"] = attrs

def is_valid_attr_for_student_on_hours(student, attr, uren):
    # vaste dagplaatsingen niet aanpassen
    if student["naam"] in vaste_studenten:
        return False

    # student moet attractie kunnen doen
    if not student_kan_attr(student, attr):
        return False

    # attractie moet op al die uren actief en geldig zijn
    for uur in uren:
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            return False
        if attr in red_spots.get(uur, set()):
            return False

    return True

def respects_student_attr_rules(student, attr):
    uren = get_hours_on_attr(student, attr)
    if len(uren) > 6:
        return False
    if max_consecutive_hours(uren) > 4:
        return False
    return True

def can_swap_exact_block(student_a, attr_a, block_hours, student_b, attr_b):
    # zelfde student of zelfde attractie heeft geen zin
    if student_a["naam"] == student_b["naam"]:
        return False
    if attr_a == attr_b:
        return False

    # beide richtingen moeten kunnen
    if not is_valid_attr_for_student_on_hours(student_a, attr_b, block_hours):
        return False
    if not is_valid_attr_for_student_on_hours(student_b, attr_a, block_hours):
        return False

    # student_b moet op exact deze uren ook éénzelfde blok hebben op attr_b
    for uur in block_hours:
        if student_b["naam"] not in assigned_map.get((uur, attr_b), []):
            return False
        # en niet tegelijk nog ergens anders zitten
        current_attr = get_student_attr_on_hour(student_b["naam"], uur)
        if current_attr != attr_b:
            return False

    # student_a moet natuurlijk ook exact daar staan
    for uur in block_hours:
        if student_a["naam"] not in assigned_map.get((uur, attr_a), []):
            return False
        current_attr = get_student_attr_on_hour(student_a["naam"], uur)
        if current_attr != attr_a:
            return False

    return True

def count_problem_attrs(student):
    """
    Tel voor hoeveel attracties deze student meer dan 4 uur ingepland staat.
    """
    count = 0
    for attr in list(student["assigned_attracties"]):
        if len(get_hours_on_attr(student, attr)) > 4:
            count += 1
    return count

def total_overflow_hours(student):
    """
    Tel hoeveel uren boven de limiet van 4 uur deze student in totaal heeft.
    Voorbeeld:
    - 5 uur op een attractie => +1
    - 6 uur op een attractie => +2
    """
    overflow = 0
    for attr in list(student["assigned_attracties"]):
        uren = len(get_hours_on_attr(student, attr))
        if uren > 4:
            overflow += (uren - 4)
    return overflow

def can_use_block_as_swap_target(student, attr, block_hours):
    """
    Check of student op exact deze uren op exact dezelfde attractie staat.
    """
    for uur in block_hours:
        if student["naam"] not in assigned_map.get((uur, attr), []):
            return False
        huidige_attr = get_student_attr_on_hour(student["naam"], uur)
        if huidige_attr != attr:
            return False
    return True

def try_swap_specific_block(student, attr, block_hours):
    """
    Probeer één specifiek blok (eerste OF laatste) van student/attr te wisselen.
    Alleen als:
    - het blok 2 of 3 uur lang is
    - de andere student op exact die uren ook één blok op één attractie heeft
    - alle regels geldig blijven
    - max 1 extra wissel ontstaat
    - het totaal aantal >4u-problemen niet stijgt
    - en liefst daalt
    """
    if len(block_hours) not in [2, 3]:
        return False

    orig_switches_a = count_attr_switches(student)
    orig_problem_count_a = count_problem_attrs(student)
    orig_overflow_a = total_overflow_hours(student)

    eerste_uur = block_hours[0]
    kandidaten = []

    for andere_student in studenten_workend:
        if andere_student["naam"] == student["naam"]:
            continue
        if andere_student["naam"] in vaste_studenten:
            continue

        attr_b = get_student_attr_on_hour(andere_student["naam"], eerste_uur)
        if not attr_b or attr_b == attr:
            continue

        # Andere student moet exact op dit hele blok op dezelfde attractie staan
        if not can_use_block_as_swap_target(andere_student, attr_b, block_hours):
            continue

        # Beide studenten moeten elkaars attractie op die uren mogen doen
        if not is_valid_attr_for_student_on_hours(student, attr_b, block_hours):
            continue
        if not is_valid_attr_for_student_on_hours(andere_student, attr, block_hours):
            continue

        kandidaten.append((andere_student["naam"], attr_b, andere_student))

    for _, attr_b, andere_student in kandidaten:
        orig_switches_b = count_attr_switches(andere_student)
        orig_problem_count_b = count_problem_attrs(andere_student)
        orig_overflow_b = total_overflow_hours(andere_student)

        # --- tijdelijke swap uitvoeren ---
        for uur in block_hours:
            remove_assignment(student, uur, attr)
            remove_assignment(andere_student, uur, attr_b)

        for uur in block_hours:
            add_assignment(student, uur, attr_b)
            add_assignment(andere_student, uur, attr)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

        valid = True

        # Regels voor beide studenten / beide attracties
        for s, a in [
            (student, attr),
            (student, attr_b),
            (andere_student, attr),
            (andere_student, attr_b),
        ]:
            if not respects_student_attr_rules(s, a):
                valid = False

        # Max 1 extra wissel in totaal
        new_switches_a = count_attr_switches(student)
        new_switches_b = count_attr_switches(andere_student)
        extra_wissels = (new_switches_a - orig_switches_a) + (new_switches_b - orig_switches_b)

        if extra_wissels > 1:
            valid = False

        # Problemen na swap
        new_problem_count_a = count_problem_attrs(student)
        new_problem_count_b = count_problem_attrs(andere_student)
        new_overflow_a = total_overflow_hours(student)
        new_overflow_b = total_overflow_hours(andere_student)

        orig_total_problem_count = orig_problem_count_a + orig_problem_count_b
        new_total_problem_count = new_problem_count_a + new_problem_count_b

        orig_total_overflow = orig_overflow_a + orig_overflow_b
        new_total_overflow = new_overflow_a + new_overflow_b

        # Geen nieuw probleem creëren
        if new_total_problem_count > orig_total_problem_count:
            valid = False

        # Geen grotere overschrijding creëren
        if new_total_problem_count == orig_total_problem_count and new_total_overflow > orig_total_overflow:
            valid = False

        # Moet minstens iets verbeteren
        verbetering = (
            (new_total_problem_count < orig_total_problem_count)
            or (
                new_total_problem_count == orig_total_problem_count
                and new_total_overflow < orig_total_overflow
            )
        )

        if not verbetering:
            valid = False

        if valid:
            return True

        # --- rollback ---
        for uur in block_hours:
            remove_assignment(student, uur, attr_b)
            remove_assignment(andere_student, uur, attr)

        for uur in block_hours:
            add_assignment(student, uur, attr)
            add_assignment(andere_student, uur, attr_b)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

    return False

def try_swap_last_or_first_block(student, attr):
    """
    Probeer eerst het laatste blok op deze attractie te wisselen.
    Lukt dat niet, probeer dan het eerste blok.
    Alleen relevant als student >4 uur op deze attractie staat.
    """
    uren_op_attr = get_hours_on_attr(student, attr)
    if len(uren_op_attr) <= 4:
        return False

    runs = get_runs_on_attr(student, attr)
    if not runs:
        return False

    laatste_run = runs[-1]
    eerste_run = runs[0]

    # Eerst laatste blok proberen
    if len(laatste_run) in [2, 3]:
        if try_swap_specific_block(student, attr, laatste_run):
            return True

    # Daarna eerste blok proberen
    if len(eerste_run) in [2, 3]:
        # niet dubbel proberen als er maar 1 run is en die identiek is
        if eerste_run != laatste_run:
            if try_swap_specific_block(student, attr, eerste_run):
                return True

    return False


# Iteratief toepassen tot er niets meer verandert
max_block_swap_passes = 5
for _ in range(max_block_swap_passes):
    wijziging = False

    for student in studenten_workend:
        probleem_attracties = [
            a for a in list(student["assigned_attracties"])
            if len(get_hours_on_attr(student, a)) > 4
        ]

        # Eerst de zwaarste problemen proberen
        probleem_attracties.sort(
            key=lambda a: (
                -len(get_hours_on_attr(student, a)),
                -max(get_hours_on_attr(student, a))
            )
        )

        for attr in probleem_attracties:
            if try_swap_last_or_first_block(student, attr):
                wijziging = True
                break

    if not wijziging:
        break


# -----------------------------
# Volgorde attracties uit Input!BL16:BL33
# -----------------------------
input_volgorde = []
for rij in range(16, 34):  # 16 t.e.m. 33
    waarde = ws[f"BL{rij}"].value
    if waarde:
        input_volgorde.append(str(waarde).strip())

# -----------------------------
# Alle attracties die minstens één keer actief zijn (voor output)
# -----------------------------
alle_actieve_attracties = set()
for uur in open_uren:
    alle_actieve_attracties |= actieve_attracties_per_uur.get(uur, set())

# Eerst de gewone attracties in de volgorde van BL16:BL33
geordende_attracties = [a for a in input_volgorde if a in alle_actieve_attracties]

# Samengevoegde attracties slim invoegen:
# bv. "A + B" direct na de laatste van A/B in de inputvolgorde
samengestelde_attracties = [a for a in alle_actieve_attracties if " + " in str(a)]
overige_attracties = [
    a for a in alle_actieve_attracties
    if a not in geordende_attracties and a not in samengestelde_attracties
]

for sameng in samengestelde_attracties:
    onderdelen = [x.strip() for x in str(sameng).split("+")]

    # Zoek de positie van het laatste onderdeel in de huidige lijst
    laatst_gevonden_index = -1
    for onderdeel in onderdelen:
        if onderdeel in geordende_attracties:
            idx = geordende_attracties.index(onderdeel)
            laatst_gevonden_index = max(laatst_gevonden_index, idx)

    if laatst_gevonden_index >= 0:
        geordende_attracties.insert(laatst_gevonden_index + 1, sameng)
    else:
        # Als geen enkel onderdeel in de inputvolgorde staat,
        # zet hem voorlopig bij de rest
        overige_attracties.append(sameng)

# Voeg tenslotte nog attracties toe die niet in BL16:BL33 stonden
alle_actieve_attracties = geordende_attracties + overige_attracties


# -----------------------------
# Output-fix: houd studenten zo veel mogelijk
# op dezelfde plek (1 of 2) per attractie over opeenvolgende uren
# -----------------------------
def stabiliseer_assigned_map_voor_output():
    """
    Deze functie verandert niets aan WIE waar staat,
    maar alleen in welke volgorde namen in assigned_map[(uur, attr)] staan.

    Doel:
    - Studenten zo veel mogelijk op dezelfde plek (1 of 2) houden over opeenvolgende uren.
    - Extra slim omgaan met uren waarop plek 2 later verdwijnt:
      als iemand doorloopt naar een volgend uur met slechts 1 plek,
      dan zetten we die student liefst al op plek 1 in het uur ervoor.
    """

    def get_namen_op_uur(attr, uur):
        namen = list(assigned_map.get((uur, attr), []))
        unieke_namen = []
        for naam in namen:
            if naam and naam not in unieke_namen:
                unieke_namen.append(naam)
        return unieke_namen

    def get_max_pos(attr, uur):
        max_pos = aantallen[uur].get(attr, 1)
        if attr in second_spot_blocked.get(uur, set()):
            max_pos = 1
        return max_pos

    def naam_staat_op_attr_in_volgend_uur(attr, huidig_uur, naam):
        volgende_uren = [u for u in sorted(open_uren) if u > huidig_uur]
        if not volgende_uren:
            return False
        volgend_uur = volgende_uren[0]
        return naam in get_namen_op_uur(attr, volgend_uur)

    def naam_moet_liefst_naar_plek1(attr, huidig_uur, naam):
        """
        True als deze naam in het volgende uur nog op dezelfde attractie staat
        én het volgende uur maar 1 plek heeft.
        Dan is het logisch om deze student nu al op plek 1 te zetten.
        """
        volgende_uren = [u for u in sorted(open_uren) if u > huidig_uur]
        if not volgende_uren:
            return False

        volgend_uur = volgende_uren[0]
        if get_max_pos(attr, volgend_uur) != 1:
            return False

        return naam in get_namen_op_uur(attr, volgend_uur)

    for attr in alle_actieve_attracties:
        vorige_slots = {1: None, 2: None}

        for uur in sorted(open_uren):
            namen = get_namen_op_uur(attr, uur)
            max_pos = get_max_pos(attr, uur)

            if not namen:
                assigned_map[(uur, attr)] = []
                vorige_slots = {1: None, 2: None}
                continue

            if max_pos <= 1:
                assigned_map[(uur, attr)] = [namen[0]]
                vorige_slots = {1: namen[0], 2: None}
                continue

            # Vanaf hier: 2 plekken beschikbaar
            slots = {1: None, 2: None}
            resterend = namen[:]

            # 1) Eerst vooruitkijken:
            # als een student in het volgende uur doorloopt terwijl daar nog maar 1 plek is,
            # dan krijgt die student nu voorrang op plek 1.
            voorkeursnaam_plek1 = None
            kandidaten_plek1 = [n for n in resterend if naam_moet_liefst_naar_plek1(attr, uur, n)]
            if len(kandidaten_plek1) == 1:
                voorkeursnaam_plek1 = kandidaten_plek1[0]
            elif len(kandidaten_plek1) > 1:
                # Als er meerdere kandidaten zijn:
                # geef voorkeur aan wie vorige uur al op plek 1 stond,
                # anders gewoon de eerste in de huidige lijst.
                if vorige_slots.get(1) in kandidaten_plek1:
                    voorkeursnaam_plek1 = vorige_slots.get(1)
                else:
                    voorkeursnaam_plek1 = kandidaten_plek1[0]

            if voorkeursnaam_plek1 in resterend:
                slots[1] = voorkeursnaam_plek1
                resterend.remove(voorkeursnaam_plek1)

            # 2) Daarna achterwaartse stabiliteit:
            # probeer dezelfde student op dezelfde plek te houden
            for pos in [1, 2]:
                if slots[pos] is not None:
                    continue
                vorige_naam = vorige_slots.get(pos)
                if vorige_naam in resterend:
                    slots[pos] = vorige_naam
                    resterend.remove(vorige_naam)

            # 3) Als plek 1 nog leeg is, geef lichte voorkeur aan iemand
            # die ook in het volgende uur op deze attractie blijft staan
            if slots[1] is None:
                doorlopers = [n for n in resterend if naam_staat_op_attr_in_volgend_uur(attr, uur, n)]
                if len(doorlopers) == 1:
                    slots[1] = doorlopers[0]
                    resterend.remove(doorlopers[0])
                elif len(doorlopers) > 1:
                    # behoud indien mogelijk de oude plek-1 volgorde
                    if vorige_slots.get(1) in doorlopers:
                        slots[1] = vorige_slots.get(1)
                        resterend.remove(vorige_slots.get(1))
                    else:
                        slots[1] = doorlopers[0]
                        resterend.remove(doorlopers[0])

            # 4) Vul de rest gewoon op
            for pos in [1, 2]:
                if slots[pos] is None and resterend:
                    slots[pos] = resterend.pop(0)

            nieuwe_volgorde = []
            if slots[1]:
                nieuwe_volgorde.append(slots[1])
            if slots[2]:
                nieuwe_volgorde.append(slots[2])

            assigned_map[(uur, attr)] = nieuwe_volgorde
            vorige_slots = {1: slots[1], 2: slots[2]}

stabiliseer_assigned_map_voor_output()


# -----------------------------

# Excel output
# -----------------------------
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Planning"

gray_fill = PatternFill(start_color="808080", fill_type="solid")

# Witte fill voor headers en attracties
white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
pv_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
extra_fill = PatternFill(start_color="FCE4D6", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Felle, maar lichte pastelkleuren (gelijkmatige felheid, veel variatie)
studenten_namen = sorted({s["naam"] for s in studenten})
# Pauzevlinders krijgen ook een kleur uit het schema
alle_namen = studenten_namen + [pv for pv in pauzevlinder_namen if pv not in studenten_namen]
# Unieke kleuren genereren: als er te weinig kleuren zijn, maak er meer met lichte variatie
base_colors = [
    "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", "E0BBE4", "957DAD", "D291BC", "FEC8D8", "FFDFD3",
    "B5EAD7", "C7CEEA", "FFDAC1", "E2F0CB", "F6DFEB", "F9E2AE", "B6E2D3", "B6D0E2", "F6E2B3", "F7C5CC",
    "F7E6C5", "C5F7D6", "C5E6F7", "F7F6C5", "F7C5F7", "C5C5F7", "C5F7F7", "F7C5C5", "C5F7C5", "F7E2C5",
    "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "F7F7C5", "C5F7F7", "F7C5F7", "C5C5F7", "F7C5C5",
    "C5F7C5", "F7E2C5", "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "E2C5F7", "C5F7E2", "E2F7C5"
]
import colorsys
def pastel_variant(hex_color, variant):
    # hex_color: 'RRGGBB', variant: int
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    # kleine variatie in lichtheid en saturatie
    l = min(1, l + 0.03 * (variant % 3))
    s = max(0.3, s - 0.04 * (variant % 5))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return f"{int(r2*255):02X}{int(g2*255):02X}{int(b2*255):02X}"

unique_colors = []
needed = len(alle_namen)
variant = 0
while len(unique_colors) < needed:
    for base in base_colors:
        if len(unique_colors) >= needed:
            break
        # voeg lichte variatie toe als nodig
        color = pastel_variant(base, variant) if variant > 0 else base
        if color not in unique_colors:
            unique_colors.append(color)
    variant += 1

student_kleuren = dict(zip(alle_namen, unique_colors))

ws_out.cell(1, 1, vandaag).font = Font(bold=True)
ws_out.cell(1, 1).fill = white_fill

for col_idx, uur in enumerate(sorted(open_uren), start=2):
    ws_out.cell(1, col_idx, f"{uur}:00").font = Font(bold=True)
    ws_out.cell(1, col_idx).fill = white_fill
    ws_out.cell(1, col_idx).alignment = center_align
    ws_out.cell(1, col_idx).border = thin_border

# --- NIEUWE LOGICA VOOR AS2 VINKJE ---
# AS is de 45e kolom in de Input-sheet (ws)
as2_vinkje = ws.cell(2, 45).value 
if as2_vinkje in [1, True, "WAAR", "X"]:
    # Cel B1 is kolom 2, rij 1
    ws_out.cell(1, 2).value = "9u30-11u"
    # Cel J1 is kolom 10, rij 1
    ws_out.cell(1, 10).value = "18u-19u30"
# -------------------------------------
    

rij_out = 2
for attr in alle_actieve_attracties:
    # 1. Bepaal hoeveel rijen deze attractie nodig heeft (1 of 2 plekken)
    max_pos = max(
        max(aantallen[uur].get(attr, 1) for uur in open_uren),
        max(per_hour_assigned_counts[uur].get(attr, 0) for uur in open_uren)
    )

    for pos_idx in range(1, max_pos + 1):
        # --- LAYOUT: Naam gevolgd door spatie en nummer (zonder haakjes) ---
        display_name = f"{attr} {pos_idx}" if max_pos > 1 else attr
        ws_out.cell(rij_out, 1, display_name).font = Font(bold=True)
        ws_out.cell(rij_out, 1).fill = white_fill
        ws_out.cell(rij_out, 1).border = thin_border

        for col_idx, uur in enumerate(sorted(open_uren), start=2):
            cell = ws_out.cell(rij_out, col_idx)

            # Haal de studentnaam op voor dit uur en deze positie
            namen = assigned_map.get((uur, attr), [])
            naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""

            # --- LOGICA VOOR GRIJS KLEUREN ---
            current_attr_norm = normalize_attr(attr)
            is_samengesteld = " + " in attr
            groepen_dit_uur = uur_samenvoegingen.get(uur, [])
            
            moet_grijs = False

            # A. Check of de attractie dit uur gesloten is
            if uur in dichte_uren_per_attr.get(current_attr_norm, set()):
                moet_grijs = True

            # B. Check voor samengestelde attracties (bv. 'A + B')
            elif is_samengesteld:
                # De samengevoegde rij is grijs als deze specifieke groep dit uur NIET actief is
                onderdelen_set = {normalize_attr(x.strip()) for x in attr.split("+")}
                actief_als_groep = any({normalize_attr(g) for g in groep} == onderdelen_set for groep in groepen_dit_uur)
                if not actief_als_groep:
                    moet_grijs = True

            # C. Check voor individuele attracties (bv. 'A')
            else:
                # De individuele rij wordt grijs als de attractie opgaat in een samenvoeging
                is_onderdeel_van_samenvoeging = any(current_attr_norm in [normalize_attr(g) for g in groep] for groep in groepen_dit_uur)
                if is_onderdeel_van_samenvoeging:
                    moet_grijs = True

            # D. Check of de tweede plek geblokkeerd is (red spots)
            if pos_idx == 2 and attr in second_spot_blocked.get(uur, set()):
                moet_grijs = True

            # --- Cel invullen en opmaken ---
            cell.value = naam
            cell.alignment = center_align
            cell.border = thin_border

            if moet_grijs:
                cell.fill = gray_fill  # Grijs uit je bronnen
            elif naam and naam in student_kleuren:
                cell.fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
            else:
                cell.fill = white_fill

        rij_out += 1
        
# Pauzevlinders
rij_out += 1
pauzevlinder_namen_sorted = [pv["naam"] for pv in selected]

for pv_idx, pvnaam in enumerate(pauzevlinder_namen_sorted, start=1):
    ws_out.cell(rij_out, 1, f"Pauzevlinder {pv_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        naam = pvnaam if uur in required_pauze_hours else ""
        ws_out.cell(rij_out, col_idx, naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if naam and naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
    rij_out += 1

# Extra's per rij
rij_out += 1
extras_flat = []
for uur in sorted(open_uren):
    for naam in extra_assignments[uur]:
        if naam not in extras_flat:
            extras_flat.append(naam)
for extra_idx, naam in enumerate(extras_flat, start=1):
    ws_out.cell(rij_out, 1, f"Extra {extra_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        # Toon naam alleen als deze extra op dit uur is ingepland
        cell_naam = naam if naam in extra_assignments[uur] else ""
        ws_out.cell(rij_out, col_idx, cell_naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if cell_naam and cell_naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[cell_naam], fill_type="solid")
    rij_out += 1

# Kolombreedte
for col in range(1, len(open_uren) + 2):
    ws_out.column_dimensions[get_column_letter(col)].width = 18

# ---- student_totalen beschikbaar maken voor volgende delen ----
from collections import defaultdict
student_totalen = defaultdict(int)
for row in ws_out.iter_rows(min_row=2, values_only=True):
    for naam in row[1:]:
        if naam and str(naam).strip() != "":
            student_totalen[naam] += 1


# -----------------------------
# Analyse-sheet maken indien nodig
# Alleen als er nog extra's zijn terwijl er elders echte lege plekken zijn
# -----------------------------

def heeft_echte_lege_plek():
    """
    True als er minstens 1 echte lege plek bestaat op de planning:
    - attractie is actief op dat uur
    - niet gesloten / niet red spot
    - geen geblokkeerde 2e plek
    - plaats is binnen de capaciteit
    - er staat nog niemand op die plek
    """
    for uur in open_uren:
        for attr in actieve_attracties_per_uur.get(uur, set()):
            if attr in red_spots.get(uur, set()):
                continue

            max_pos = aantallen[uur].get(attr, 1)
            if attr in second_spot_blocked.get(uur, set()):
                max_pos = 1

            namen = assigned_map.get((uur, attr), [])
            for pos_idx in range(1, max_pos + 1):
                naam = namen[pos_idx - 1] if pos_idx - 1 < len(namen) else ""
                if not naam:
                    return True
    return False


def heeft_extra_studenten():
    return any(len(namen) > 0 for namen in extra_assignments.values())


def student_is_aanwezig_op_uur_zonder_pauzevlinder(student, uur):
    """
    Student telt mee in analyse voor dit uur als:
    - student effectief aanwezig is op dit uur
      (ingepland of extra)
    - en NIET als pauzevlinder bezig is op dit uur
    """
    naam = student["naam"]

    # Pauzevlinder tijdens pauzevlinderuur telt niet mee
    if student.get("is_pauzevlinder") and uur in required_pauze_hours:
        return False

    if uur in set(student.get("assigned_hours", [])):
        return True

    if naam in extra_assignments.get(uur, []):
        return True

    return False


def student_kan_attr_in_analyse(student, attr):
    """
    Voor analyse:
    - respecteer blacklist
    - samengevoegde attractie mag enkel als student alle onderdelen kan
    """
    naam = student["naam"]

    if " + " not in attr:
        return attr.lower() not in student_blacklist.get(naam, set()) and attr in student.get("attracties", [])

    onderdelen = [a.strip() for a in attr.split("+")]
    for onderdeel in onderdelen:
        if onderdeel.lower() in student_blacklist.get(naam, set()):
            return False

    return all(onderdeel in student.get("attracties", []) for onderdeel in onderdelen)


def actieve_analyse_attracties_op_uur(uur):
    """
    Geeft attracties terug in de volgorde van Input!BL16:BL33,
    maar aangepast aan het specifieke uur:
    - losse attracties als ze actief zijn
    - samengevoegde attracties enkel als ze dat uur actief samengevoegd zijn
    """
    actieve_set = actieve_attracties_per_uur.get(uur, set())

    input_volgorde_lokaal = []
    for rij_bl in range(16, 34):  # BL16 t.e.m. BL33
        attr = ws[f"BL{rij_bl}"].value
        if attr:
            input_volgorde_lokaal.append(str(attr).strip())

    resultaat = []
    gebruikte = set()

    # Eerst gewone attracties in inputvolgorde
    for attr in input_volgorde_lokaal:
        if attr in actieve_set and attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

        # Kijk of een samengestelde attractie met dit onderdeel actief is op dit uur
        for actief_attr in actieve_set:
            if " + " not in str(actief_attr):
                continue
            onderdelen = [x.strip() for x in str(actief_attr).split("+")]
            if attr in onderdelen and actief_attr not in gebruikte:
                # pas invoegen na laatste onderdeel uit de inputvolgorde
                if all(o in input_volgorde_lokaal for o in onderdelen):
                    laatst_idx = max(input_volgorde_lokaal.index(o) for o in onderdelen)
                    huidig_idx = input_volgorde_lokaal.index(attr)
                    if huidig_idx == laatst_idx:
                        resultaat.append(actief_attr)
                        gebruikte.add(actief_attr)

    # Daarna nog eventuele actieve attracties die niet in BL-lijst zaten
    for attr in actieve_set:
        if attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

    return resultaat


ws_analyse = wb_out.create_sheet(title="Analyse")

analyse_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
witte_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# -----------------------------
# Titel bovenaan de Analyse-pagina
# -----------------------------
titel = "Hier zie je per uur welke studenten aanwezig zijn en welke attracties ze kunnen:"
ws_analyse.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)

titel_cel = ws_analyse.cell(1, 1, titel)
titel_cel.font = Font(bold=True, size=12)
titel_cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
titel_cel.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
titel_cel.border = thin_border

start_rij = 3

for uur in sorted(open_uren):
    analyse_studenten_uur = [
        s for s in studenten
        if student_is_aanwezig_op_uur_zonder_pauzevlinder(s, uur)
    ]
    analyse_studenten_uur = sorted(analyse_studenten_uur, key=lambda s: naam_tie_break_key(s["naam"]))

    analyse_attracties_uur = actieve_analyse_attracties_op_uur(uur)

    # Als er voor dit uur niets te tonen is, sla over
    if not analyse_studenten_uur or not analyse_attracties_uur:
        continue

    # Uur in plaats van datum
    ws_analyse.cell(start_rij, 1, f"{uur}:00").font = Font(bold=True)
    ws_analyse.cell(start_rij, 1).fill = analyse_header_fill
    ws_analyse.cell(start_rij, 1).alignment = center_align
    ws_analyse.cell(start_rij, 1).border = thin_border

    ws_analyse.cell(start_rij, 2, "Student").font = Font(bold=True)
    ws_analyse.cell(start_rij, 2).fill = analyse_header_fill
    ws_analyse.cell(start_rij, 2).alignment = center_align
    ws_analyse.cell(start_rij, 2).border = thin_border

    # 1 kolom per attractie
    start_col_attr = 3
    for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
        cel = ws_analyse.cell(start_rij, idx, attr)
        cel.font = Font(bold=True)
        cel.fill = analyse_header_fill
        cel.alignment = center_align
        cel.border = thin_border

    # Data voor dit uur
    rij = start_rij + 1
    for s in analyse_studenten_uur:
        naam = s["naam"]

        ws_analyse.cell(rij, 1, rij - start_rij).alignment = center_align
        ws_analyse.cell(rij, 1).border = thin_border
        ws_analyse.cell(rij, 1).fill = witte_fill

        naam_cel = ws_analyse.cell(rij, 2, naam)
        naam_cel.alignment = center_align
        naam_cel.border = thin_border
        student_fill = witte_fill

        if naam in student_kleuren:
            student_fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
            naam_cel.fill = student_fill
        else:
            naam_cel.fill = witte_fill

        for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
            cel = ws_analyse.cell(rij, idx)
            cel.alignment = center_align
            cel.border = thin_border
            cel.font = Font(color="000000")

            if student_kan_attr_in_analyse(s, attr):
                cel.value = attr
                cel.fill = student_fill
            else:
                cel.value = ""
                cel.fill = witte_fill

        rij += 1

    # Kolombreedtes
    ws_analyse.column_dimensions["A"].width = 8
    ws_analyse.column_dimensions["B"].width = 24
    for idx in range(start_col_attr, start_col_attr + len(analyse_attracties_uur)):
        ws_analyse.column_dimensions[get_column_letter(idx)].width = 13.5

    # Enkele lege rijen tussen uurblokken
    start_rij = rij + 3


#DEEL 2
#oooooooooooooooooooo
#oooooooooooooooooooo

# -----------------------------
# DEEL 2: Pauzevlinder overzicht
# -----------------------------
ws_pauze = wb_out.create_sheet(title="Pauzevlinders")

light_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# -----------------------------
# Rij 1: Uren
# -----------------------------
# Gebruik compute_pauze_hours/open_uren als basis voor de pauzeplanning-urenrij
uren_rij1 = []
from datetime import datetime, timedelta
if required_pauze_hours:
    start_uur = min(required_pauze_hours)
    eind_uur = max(required_pauze_hours)
    tijd = datetime(2020,1,1,start_uur,0)
    # Laatste pauze mag een kwartier vóór het einde starten
    laatste_pauze = datetime(2020,1,1,eind_uur,30)
    while tijd <= laatste_pauze:
        uren_rij1.append(f"{tijd.hour}u" if tijd.minute==0 else f"{tijd.hour}u{tijd.minute:02d}")
        tijd += timedelta(minutes=15)
else:
    # fallback: gebruik open_uren
    for uur in sorted(open_uren):
        uren_rij1.append(f"{uur}u")

# Schrijf uren in rij 1, start in kolom B
for col_idx, uur in enumerate(uren_rij1, start=2):
    c = ws_pauze.cell(1, col_idx, uur)
    c.fill = light_fill
    c.alignment = center_align
    c.border = thin_border

### Zet de datum van vandaag in cel A1 van de pauzeplanning
a1 = ws_pauze.cell(1, 1, vandaag)
a1.font = Font(bold=True)
a1.fill = light_fill
a1.alignment = center_align
a1.border = thin_border

# -----------------------------
# Pauzevlinders en namen
# -----------------------------
rij_out = 2
for pv_idx, pv in enumerate(selected, start=1):
    # Titel: Pauzevlinder X
    title_cell = ws_pauze.cell(rij_out, 1, f"Pauzevlinder {pv_idx}")
    title_cell.font = Font(bold=True)
    title_cell.fill = light_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border

    # Naam eronder (zelfde stijl en kleur)
    naam_cel = ws_pauze.cell(rij_out + 1, 1, pv["naam"])
    naam_cel.fill = light_fill
    naam_cel.alignment = center_align
    naam_cel.border = thin_border

    rij_out += 3  # lege rij ertussen

# -----------------------------
# Kolombreedte voor overzicht
# -----------------------------

# Automatisch de breedte van kolom A instellen op basis van de langste tekst
max_len_colA = 0
for row in range(1, ws_pauze.max_row + 1):
    val = ws_pauze.cell(row, 1).value
    if val:
        max_len_colA = max(max_len_colA, len(str(val)))
# Voeg wat extra ruimte toe
ws_pauze.column_dimensions['A'].width = max(12, max_len_colA + 2)

for col in range(2, len(uren_rij1) + 2):
    ws_pauze.column_dimensions[get_column_letter(col)].width = 10

# Gebruik exact dezelfde open_uren en headers als in deel 1 voor de pauzeplanning
uren_rij1 = []
for uur in sorted(open_uren):
    # Zoek de originele header uit ws_out (de hoofdplanning)
    for col in range(2, ws_out.max_column + 1):
        header = ws_out.cell(1, col).value
        if header and str(header).startswith(str(uur)):
            uren_rij1.append(header)
            break

# Opslaan met dezelfde unieke naam

# Maak in-memory bestand
output = BytesIO()





#oooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


#DEEL 3
#oooooooooooooooooooo
#oooooooooooooooooooo

import random
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side, PatternFill
import datetime

# -----------------------------
# DEEL 3: Extra naam voor pauzevlinders die langer dan 6 uur werken
# -----------------------------

# Sheet referenties
ws_planning = wb_out["Planning"]
ws_pauze = wb_out["Pauzevlinders"]

# Pauzekolommen (B–G in Pauzevlinders sheet)
# Dynamisch: alle kolommen waar in rij 1 een uur staat (bv. '13u45', '14u', ...)
pauze_cols = []
for col in range(2, ws_pauze.max_column + 1):
    header = ws_pauze.cell(1, col).value
    if header and ("u" in str(header)):
        pauze_cols.append(col)

# Bouw lijst met pauzevlinder-rijen
pv_rows = []
for pv in selected:
    row_found = None
    for r in range(2, ws_pauze.max_row + 1):
        if str(ws_pauze.cell(r, 1).value).strip() == str(pv["naam"]).strip():
            row_found = r
            break
    if row_found is not None:
        pv_rows.append((pv, row_found))

# Bereken totaal uren per student in Werkblad "Planning"
student_totalen = defaultdict(int)
for row in ws_planning.iter_rows(min_row=2, values_only=True):
    for naam in row[1:]:
        if naam and str(naam).strip() != "":
            student_totalen[naam] += 1

# Loop door pauzevlinders in Werkblad "Pauzevlinders"


# ---- OPTIMALISATIE: Meerdere planningen genereren en de beste kiezen ----
import copy
best_score = None
best_state = None
num_runs = 5
for _run in range(num_runs):
    # Maak een deep copy van de relevante werkbladen en variabelen
    ws_pauze_tmp = wb_out.copy_worksheet(ws_pauze)
    ws_pauze_tmp.title = f"Pauzevlinders_tmp_{_run}"
    # Reset alle naamcellen
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            ws_pauze_tmp.cell(pv_row, col).value = None
    # Herhaal de bestaande logica voor pauzeplanning, maar werk op ws_pauze_tmp
    # ...existing code for pauzeplanning, but use ws_pauze_tmp instead of ws_pauze...
    # (Voor deze patch: laat de bestaande logica staan, dit is een structuurvoorzet. Zie opmerking hieronder)
    # ---- Evalueer deze planning ----
    # 1. Iedereen een pauze?
    korte_pauze_ontvangers = set()
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            cel = ws_pauze_tmp.cell(pv_row, col)
            if cel.value and str(cel.value).strip() != "":
                # Check of het een korte pauze is (enkel blok, niet dubbel)
                idx = pauze_cols.index(col)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze_tmp.cell(pv_row, next_col)
                    if cel_next.value == cel.value:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze_tmp.cell(pv_row, prev_col)
                    if prev_cel.value == cel.value:
                        is_lange = True
                if not is_lange:
                    korte_pauze_ontvangers.add(str(cel.value).strip())
    alle_studenten = [s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) >= 4]
    iedereen_pauze = all(naam in korte_pauze_ontvangers for naam in alle_studenten)
    # 2. Eerlijkheid: verschil max-min korte pauzes per pauzevlinder
    from collections import Counter
    pv_korte_pauze_count = Counter()
    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            cel = ws_pauze_tmp.cell(pv_row, col)
            if cel.value and str(cel.value).strip() != "":
                idx = pauze_cols.index(col)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze_tmp.cell(pv_row, next_col)
                    if cel_next.value == cel.value:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze_tmp.cell(pv_row, prev_col)
                    if prev_cel.value == cel.value:
                        is_lange = True
                if not is_lange:
                    pv_korte_pauze_count[pv["naam"]] += 1
    if pv_korte_pauze_count:
        eerlijkheid = max(pv_korte_pauze_count.values()) - min(pv_korte_pauze_count.values())
    else:
        eerlijkheid = 999
    # Score: eerst iedereen_pauze, dan eerlijkheid
    score = (iedereen_pauze, -eerlijkheid)
    if (best_score is None) or (score > best_score):
        best_score = score
        best_state = copy.deepcopy(ws_pauze_tmp)

# Na alle runs: kopieer best_state naar ws_pauze
if best_state is not None:

    for pv, pv_row in pv_rows:
        for col in pauze_cols:
            ws_pauze.cell(pv_row, col).value = best_state.cell(pv_row, col).value

# ---- Verwijder tijdelijke werkbladen ----
tmp_sheets = [ws for ws in wb_out.worksheets if ws.title.startswith("Pauzevlinders_tmp")]
for ws in tmp_sheets:
    wb_out.remove(ws)

# ---- Lege naamcellen inkleuren ----
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
naam_leeg_fill_pp2 = naam_leeg_fill
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")

for pv, pv_row in pv_rows:
    for col in pauze_cols:
        if ws_pauze.cell(pv_row, col).value in [None, ""]:
            ws_pauze.cell(pv_row, col).fill = naam_leeg_fill






#ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo

# DEEL 4: Lange werkers (>6 uur) pauze inplannen – gegarandeerd
# -----------------------------

from openpyxl.styles import Alignment, Border, Side, PatternFill
import random  # <— toegevoegd voor willekeurige verdeling

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center")
# Zachtblauw, anders dan je titelkleuren; alleen voor naamcellen
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Alleen kolommen B..G
# Dynamisch: alle kolommen waar in rij 1 een uur staat (bv. '13u45', '14u', ...)
pauze_cols = []
for col in range(2, ws_pauze.max_column + 1):
    header = ws_pauze.cell(1, col).value
    if header and ("u" in str(header)):
        pauze_cols.append(col)


def is_student_extra(naam):
    """Check of student in Planning bij 'extra' staat (kolom kan 'Extra' zijn of specifieke marker)."""
    for row in range(2, ws_planning.max_row + 1):
        if ws_planning.cell(row, 1).value:  # rij met attractienaam
            for col in range(2, ws_planning.max_column + 1):
                if str(ws_planning.cell(row, col).value).strip().lower() == str(naam).strip().lower():
                    header = str(ws_planning.cell(1, col).value).strip().lower()
                    if "extra" in header:  # of een andere logica afhankelijk van hoe 'extra' wordt aangeduid
                        return True
    return False


def is_pauzevlinder(naam):
    """Is deze naam een pauzevlinder?"""
    return any(pv["naam"] == naam for pv in selected)



# ---- Helpers ----
def parse_header_uur(header):
    """Map headertekst (bv. '14u', '14:00', '14:30') naar het hele uur (14)."""
    if not header:
        return None
    s = str(header).strip()
    try:
        if "u" in s:
            # '14u' of '14u30' -> 14
            return int(s.split("u")[0])
        if ":" in s:
            # '14:00' of '14:30' -> 14 (halfuur koppelen aan het hele uur)
            uur, _min = s.split(":")
            return int(uur)
        return int(s)  # fallback
    except:
        return None

# ---- Pauze-restrictie: geen korte pauze in eerste 12 kwartieren van de pauzeplanning (tenzij <=6u open) ----
def get_verboden_korte_pauze_kolommen():
    """Geeft de kolomnummers van de eerste 12 kwartieren in ws_pauze (B t/m M)."""
    return list(range(2, 12))  # kolommen 2 t/m 11 (B t/m M)

def is_korte_pauze_toegestaan_col(col, student_naam=None):
    """
    Controleert of een korte pauze in deze kolom mag.
    Uitzondering: als een student vóór 15:00 stopt, mag de pauze ALTIJD.
    """
    if len(open_uren) <= 6:
        return True
    
    # Check of de student een vroege stopper is (stopt vóór 15u)
    if student_naam:
        werk_uren = get_student_work_hours(student_naam)
        # De check 'if werk_uren' voorkomt dat de code crasht bij een lege lijst
        if werk_uren and max(werk_uren) < 15:
            return True
            
    return col not in get_verboden_korte_pauze_kolommen()

def normalize_attr(name):
    """Normaliseer attractienaam zodat 'X 2' telt als 'X'; trim & lower-case voor vergelijking."""
    if not name:
        return ""
    s = str(name).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        s = parts[0]
    return s.strip().lower()

# Build: pauzevlinder-rijen en capability-sets
pv_rows = []  # lijst: (pv_dict, naam_rij_index)
pv_cap_set = {}  # pv-naam -> set genormaliseerde attracties
for pv in selected:
    # zoek de rij waar de pv-naam in kolom A staat
    row_found = None
    for r in range(2, ws_pauze.max_row + 1):
        if str(ws_pauze.cell(r, 1).value).strip() == str(pv["naam"]).strip():
            row_found = r
            break
    if row_found is not None:
        pv_rows.append((pv, row_found))
        pv_cap_set[pv["naam"]] = {normalize_attr(a) for a in pv.get("attracties", [])}


# -----------------------------
# DEEL 1.5: Samengevoegde attracties correct registreren
# -----------------------------
# Plaats dit nadat je de PV-capabilities hebt opgebouwd, bv. na:
# pv_cap_set[pv["naam"]] = {normalize_attr(a) for a in pv.get("attracties", [])}

for pv in selected:
    nieuwe_caps = set()
    for attr in pv.get("attracties", []):
        attr_norm = normalize_attr(attr)
        # Check: bevat '+' → samengevoegde attractie
        if "+" in attr_norm:
            # split en normaliseer elk onderdeel
            onderdelen = [normalize_attr(x) for x in attr_norm.split("+")]
            # als PV elk onderdeel kan, voeg samengevoegde attractie toe
            if all(x in pv_cap_set[pv["naam"]] for x in onderdelen):
                nieuwe_caps.add(attr_norm)  # hele samengestelde attractie toevoegen
        else:
            nieuwe_caps.add(attr_norm)
    # overschrijf set met de uitgebreide mogelijkheden
    pv_cap_set[pv["naam"]] = nieuwe_caps


# Lange werkers: namen-set voor snelle checks


lange_werkers = [
    s for s in studenten
    if student_totalen.get(s["naam"], 0) > 6
    and s["naam"] not in [pv["naam"] for pv in selected]
]
lange_werkers_names = {s["naam"] for s in lange_werkers}

def get_student_work_hours(naam):
    """Welke uren werkt deze student echt (zoals te zien in werkblad 'Planning')?"""
    uren = set()
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        uur = parse_header_uur(header)
        if uur is None:
            continue
        # check of student in deze kolom ergens staat
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                uren.add(uur)
                break
    return sorted(uren)

def vind_attractie_op_uur(naam, uur):
    """Geef attractienaam (exact zoals in Planning-kolom A) waar student staat op dit uur; None als niet gevonden."""
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        col_uur = parse_header_uur(header)
        if col_uur != uur:
            continue
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                return ws_planning.cell(row, 1).value
    return None

def pv_kan_attr(pv, attr):
    """Check of pv attr kan (met normalisatie, zodat 'X 2' telt als 'X')."""
    base = normalize_attr(attr)
    if base == "extra":
        return True
    return base in pv_cap_set.get(pv["naam"], set())

# Willekeurige volgorde van pauzeplekken (pv-rij x kolom) om lege cellen random te spreiden
slot_order = [(pv, pv_row, col) for (pv, pv_row) in pv_rows for col in pauze_cols]
random.shuffle(slot_order)  # <— kern om lege plekken later willekeurig te verspreiden

def plaats_student(student, harde_mode=False):
    """
    Plaats student bij een geschikte pauzevlinder in B..G op een uur waar student effectief werkt.
    - Overschrijven alleen in harde_mode én alleen als de huidige inhoud géén lange werker is.
    - Volgorde van slots is willekeurig (slot_order) zodat lege plekken random verdeeld blijven.
    """
    naam = student["naam"]
    werk_uren = get_student_work_hours(naam)  # echte uren waarop student in 'Planning' staat
    # Pauze mag niet in eerste of laatste werkuur vallen
    werk_uren_set = set(werk_uren)
    if len(werk_uren) > 2:
        verboden_uren = {werk_uren[0], werk_uren[-1]}
    else:
        verboden_uren = set(werk_uren)  # als maar 1 of 2 uur: geen pauze mogelijk

    # Sorteer alle pauzekolommen op volgorde
    pauze_cols_sorted = sorted(pauze_cols)
    # Zoek alle (uur, col) paren, filter verboden uren
    uur_col_pairs = []
    for col in pauze_cols_sorted:
        col_header = ws_pauze.cell(1, col).value
        col_uur = parse_header_uur(col_header)
        if col_uur is not None and col_uur in werk_uren_set and col_uur not in verboden_uren:
            uur_col_pairs.append((col_uur, col))

    # Houd bij of deze student al een lange/korte pauze heeft gekregen
    if not hasattr(plaats_student, "pauze_registry"):
        plaats_student.pauze_registry = {}
    reg = plaats_student.pauze_registry.setdefault(naam, {"lange": False, "korte": False})

    # Eerst: zoek alle mogelijke dubbele blokjes voor de lange pauze
    lange_pauze_opties = []
    for i in range(len(uur_col_pairs)-1):
        uur1, col1 = uur_col_pairs[i]
        uur2, col2 = uur_col_pairs[i+1]
        if col2 == col1 + 1:
            lange_pauze_opties.append((i, uur1, col1, uur2, col2))

    # Probeer alle opties voor de lange pauze (max 1x per student)
    if not reg["lange"] and not heeft_al_lange_pauze(naam):
        for optie in lange_pauze_opties:
            i, uur1, col1, uur2, col2 = optie
            attr1 = vind_attractie_op_uur(naam, uur1)
            attr2 = vind_attractie_op_uur(naam, uur2)
            if not attr1 or not attr2:
                continue
            for (pv, pv_row, _) in slot_order:
                if not pv_kan_attr(pv, attr1) and not is_student_extra(naam):
                    continue
                cel1 = ws_pauze.cell(pv_row, col1)
                cel2 = ws_pauze.cell(pv_row, col2)
                boven_cel1 = ws_pauze.cell(pv_row-1, col1)
                boven_cel2 = ws_pauze.cell(pv_row-1, col2)
                if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
                    # Vul beide blokjes voor lange pauze
                    boven_cel1.value = attr1
                    boven_cel1.alignment = center_align
                    boven_cel1.border = thin_border
                    boven_cel2.value = attr2
                    boven_cel2.alignment = center_align
                    boven_cel2.border = thin_border
                    cel1.value = naam
                    cel1.alignment = center_align
                    cel1.border = thin_border
                    cel2.value = naam
                    cel2.alignment = center_align
                    cel2.border = thin_border
                    reg["lange"] = True
                    # Nu: zoek een korte pauze, eerst exact 10 blokjes afstand, dan 11, 12, ... tot einde, daarna 9, 8, ... tot 1
                    if not reg["korte"]:
                        found = False
                        # Eerst exact 10 blokjes afstand
                        for min_blokjes in list(range(10, len(uur_col_pairs)-i)) + list(range(9, 0, -1)):
                            for j in range(i+min_blokjes, len(uur_col_pairs)):
                                uur_kort, col_kort = uur_col_pairs[j]
                                if not is_korte_pauze_toegestaan_col(col_kort):
                                    continue
                                attr_kort = vind_attractie_op_uur(naam, uur_kort)
                                if not attr_kort:
                                    continue
                                # Belangrijk: alleen op deze PV-rij plaatsen als de pauzevlinder deze attractie kan, behalve bij EXTRA
                                if (not pv_kan_attr(pv, attr_kort)) and (not is_student_extra(naam)):
                                    continue
                                # Alleen zoeken in dezelfde rij als de lange pauze (dus bij dezelfde pauzevlinder)
                                cel_kort = ws_pauze.cell(pv_row, col_kort)
                                boven_cel_kort = ws_pauze.cell(pv_row-1, col_kort)
                                if cel_kort.value in [None, ""]:
                                    boven_cel_kort.value = attr_kort
                                    boven_cel_kort.alignment = center_align
                                    boven_cel_kort.border = thin_border
                                    cel_kort.value = naam
                                    cel_kort.alignment = center_align
                                    cel_kort.border = thin_border
                                    reg["korte"] = True
                                    found = True
                                    return True
                                elif harde_mode:
                                    occupant = str(cel_kort.value).strip() if cel_kort.value else ""
                                    if occupant not in lange_werkers_names:
                                        boven_cel_kort.value = attr_kort
                                        boven_cel_kort.alignment = center_align
                                        boven_cel_kort.border = thin_border
                                        cel_kort.value = naam
                                        cel_kort.alignment = center_align
                                        cel_kort.border = thin_border
                                        reg["korte"] = True
                                        found = True
                                        return True
                            if found:
                                break
                    # Geen korte pauze gevonden, maar lange pauze is wel gezet
                    return True
                elif harde_mode:
                    occupant1 = str(cel1.value).strip() if cel1.value else ""
                    occupant2 = str(cel2.value).strip() if cel2.value else ""
                    if (occupant1 not in lange_werkers_names) and (occupant2 not in lange_werkers_names) and not heeft_al_lange_pauze(naam):
                        boven_cel1.value = attr1
                        boven_cel1.alignment = center_align
                        boven_cel1.border = thin_border
                        boven_cel2.value = attr2
                        boven_cel2.alignment = center_align
                        boven_cel2.border = thin_border
                        cel1.value = naam
                        cel1.alignment = center_align
                        cel1.border = thin_border
                        cel2.value = naam
                        cel2.alignment = center_align
                        cel2.border = thin_border
                        reg["lange"] = True
                        # Nu: zoek een korte pauze minstens 6 blokjes verderop
                        if not reg["korte"]:
                            for j in range(i+6, len(uur_col_pairs)):
                                uur_kort, col_kort = uur_col_pairs[j]
                                attr_kort = vind_attractie_op_uur(naam, uur_kort)
                                if not attr_kort:
                                    continue
                                for (pv2, pv_row2, _) in slot_order:
                                    if not pv_kan_attr(pv2, attr_kort) and not is_student_extra(naam):
                                        continue
                                    cel_kort = ws_pauze.cell(pv_row2, col_kort)
                                    boven_cel_kort = ws_pauze.cell(pv_row2-1, col_kort)
                                    if cel_kort.value in [None, ""]:
                                        boven_cel_kort.value = attr_kort
                                        boven_cel_kort.alignment = center_align
                                        boven_cel_kort.border = thin_border
                                        cel_kort.value = naam
                                        cel_kort.alignment = center_align
                                        cel_kort.border = thin_border
                                        reg["korte"] = True
                                        return True
                                    elif harde_mode:
                                        occupant = str(cel_kort.value).strip() if cel_kort.value else ""
                                        if occupant not in lange_werkers_names:
                                            boven_cel_kort.value = attr_kort
                                            boven_cel_kort.alignment = center_align
                                            boven_cel_kort.border = thin_border
                                            cel_kort.value = naam
                                            cel_kort.alignment = center_align
                                            cel_kort.border = thin_border
                                            reg["korte"] = True
                                            return True
                        return True
    # Als geen geldige combinatie gevonden, probeer fallback (oude logica)
    # Korte pauze alleen als nog niet toegekend
    for uur in random.sample(werk_uren, len(werk_uren)):
        if uur in verboden_uren:
            continue
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            continue
        for (pv, pv_row, col) in slot_order:
            col_header = ws_pauze.cell(1, col).value
            col_uur = parse_header_uur(col_header)
            if col_uur != uur:
                continue
            if not is_korte_pauze_toegestaan_col(col, naam):
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            cel = ws_pauze.cell(pv_row, col)
            boven_cel = ws_pauze.cell(pv_row - 1, col)
            current_val = cel.value
            if current_val in [None, ""]:
                if not reg["korte"]:
                    boven_cel.value = attr
                    boven_cel.alignment = center_align
                    boven_cel.border = thin_border
                    cel.value = naam
                    cel.alignment = center_align
                    cel.border = thin_border
                    reg["korte"] = True
                    return True
            else:
                if harde_mode:
                    occupant = str(current_val).strip()
                    if occupant not in lange_werkers_names:
                        if not reg["korte"]:
                            boven_cel.value = attr
                            boven_cel.alignment = center_align
                            boven_cel.border = thin_border
                            cel.value = naam
                            cel.alignment = center_align
                            cel.border = thin_border
                            reg["korte"] = True
                            return True
    return False

# ---- Fase 1: zachte toewijzing (niet overschrijven) ----
def heeft_al_lange_pauze(naam):
    # Check of student al een dubbele blok (lange pauze) heeft in het pauzeoverzicht
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        return True
    return False

lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # lange pauze
lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")  # kwartierpauze
naam_leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
lange_pauze_ontvangers = set()
# --- Verspreid lange pauzes van lange werkers net als bij pauzevlinders ---
niet_geplaatst = []
for s in random.sample(lange_werkers, len(lange_werkers)):
    naam = s["naam"]
    if naam in lange_pauze_ontvangers or heeft_al_lange_pauze(naam):
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)
        continue
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) <= 6:
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)
        continue
    # Alleen de eerste 11 kwartieren (indices 0 t/m 10) zijn toegestaan voor lange pauzes
    halve_uren = []  # lijst van (col1, col2, uur1, uur2, pv, pv_row)
    werk_uren_set = set(werk_uren)
    verboden_uren = {werk_uren[0], werk_uren[-1]} if len(werk_uren) > 2 else set(werk_uren)
    max_start_idx = min(8, len(pauze_cols)-2)  # idx 0 t/m 10 zijn halve uren binnen eerste 11 kwartieren
    for pv, pv_row in pv_rows:
        for idx in range(max_start_idx+1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx+1]
            col1_header = ws_pauze.cell(1, col1).value
            col2_header = ws_pauze.cell(1, col2).value
            # Alleen starten op heel of half uur
            try:
                min1 = int(str(col1_header).split('u')[1]) if 'u' in str(col1_header) and len(str(col1_header).split('u')) > 1 else 0
            except:
                min1 = 0
            if min1 not in (0, 30):
                continue
            uur1 = parse_header_uur(col1_header)
            uur2 = parse_header_uur(col2_header)
            if uur1 is None or uur2 is None:
                continue
            if uur1 not in werk_uren_set or uur2 not in werk_uren_set:
                continue
            if uur1 in verboden_uren or uur2 in verboden_uren:
                continue
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            # Attractie moet kloppen
            attr1 = vind_attractie_op_uur(naam, uur1)
            attr2 = vind_attractie_op_uur(naam, uur2)
            if not attr1 or not attr2:
                continue
            if not pv_kan_attr(pv, attr1) and not is_student_extra(naam):
                continue
            if cel1.value in [None, ""] and cel2.value in [None, ""]:
                halve_uren.append((col1, col2, uur1, uur2, pv, pv_row))
    random.shuffle(halve_uren)
    # Fairness: keep a live counter of long breaks per pauzevlinder
    from collections import Counter
    if not hasattr(plaats_student, "pv_lange_pauze_count"):
        plaats_student.pv_lange_pauze_count = Counter()
    pv_lange_pauze_count = plaats_student.pv_lange_pauze_count
    for pv, _ in pv_rows:
        if pv["naam"] not in pv_lange_pauze_count:
            pv_lange_pauze_count[pv["naam"]] = 0
    geplaatst = False
    # Sorteer bij elke poging op actuele telling
    halve_uren_sorted = sorted(halve_uren, key=lambda x: pv_lange_pauze_count[x[4]["naam"]])
    for col1, col2, uur1, uur2, pv, pv_row in halve_uren_sorted:
        cel1 = ws_pauze.cell(pv_row, col1)
        cel2 = ws_pauze.cell(pv_row, col2)
        boven_cel1 = ws_pauze.cell(pv_row-1, col1)
        boven_cel2 = ws_pauze.cell(pv_row-1, col2)
        attr1 = vind_attractie_op_uur(naam, uur1)
        attr2 = vind_attractie_op_uur(naam, uur2)
        if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
            boven_cel1.value = attr1
            boven_cel1.alignment = center_align
            boven_cel1.border = thin_border
            boven_cel2.value = attr2
            boven_cel2.alignment = center_align
            boven_cel2.border = thin_border
            cel1.value = naam
            cel1.alignment = center_align
            cel1.border = thin_border
            cel1.fill = lichtgroen_fill
            cel2.value = naam
            cel2.alignment = center_align
            cel2.border = thin_border
            cel2.fill = lichtgroen_fill
            lange_pauze_ontvangers.add(naam)
            geplaatst = True
            # Fairness: niet meetellen als deze lange pauze (een van de twee blokken) een 'Extra' overname is
            if (normalize_attr(attr1) != 'extra') and (normalize_attr(attr2) != 'extra'):
                pv_lange_pauze_count[pv["naam"]] += 1
            break
    if not geplaatst:
        if not plaats_student(s, harde_mode=False):
            niet_geplaatst.append(s)

# ---- Fase 2: iteratief, met gecontroleerd overschrijven van niet-lange-werkers ----
# we herhalen meerdere passes om iedereen >6u te kunnen plaatsen
max_passes = 12
for _ in range(max_passes):
    if not niet_geplaatst:
        break
    rest = []
    # Ook hier willekeurige volgorde voor extra spreiding
    for s in random.sample(niet_geplaatst, len(niet_geplaatst)):
        if not plaats_student(s, harde_mode=True):
            rest.append(s)
    # Als niets veranderde in een hele pass, stoppen we
    if len(rest) == len(niet_geplaatst):
        break
    niet_geplaatst = rest

# ---- Lege naamcellen inkleuren (alleen de NAAM-rij; bovenliggende attractie-rij NIET kleuren) ----

# ---- Pauze-kleuren: lichtgroen voor lange pauze (dubbele blok), lichtpaars voor kwartierpauze (enkel blok) ----

lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # lange pauze
lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")  # kwartierpauze
roze_fill = PatternFill(start_color="FFD6E7", end_color="FFD6E7", fill_type="solid")

# Pauze kleuren invullen (lange en korte pauzes)
for pv, pv_row in pv_rows:
    for idx, col in enumerate(pauze_cols):
        cel = ws_pauze.cell(pv_row, col)
        if cel.value in [None, ""]:
            cel.fill = naam_leeg_fill
        else:
            # Check of dit een lange pauze is (dubbele blok: zelfde naam in 2 naast elkaar liggende cellen)
            is_langepauze = False
            # Kijk vooruit: als deze en de volgende cel dezelfde naam hebben, kleur beide groen
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value and cel.value not in [None, ""]:
                    is_langepauze = True
                    cel.fill = lichtgroen_fill
                    cel_next.fill = lichtgroen_fill
                    continue  # sla de volgende cel over, die is al gekleurd
            # Kijk achteruit: als vorige cel al groen is door lange pauze, deze niet opnieuw kleuren
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value and cel.value not in [None, ""]:
                    # Deze cel is al als tweede helft van lange pauze gekleurd
                    continue
            # Anders: kwartierpauze
            cel.fill = lichtpaars_fill

# ---- Korte pauze voor pauzevlinders zelf toevoegen (eerst, met afstandscriterium) ----
def _pv_has_short_pause(naam, pv_row):
    for idx, col in enumerate(pauze_cols):
        if ws_pauze.cell(pv_row, col).value == naam:
            left_same = idx > 0 and ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam
            right_same = idx+1 < len(pauze_cols) and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam
            if not left_same and not right_same:
                return True
    return False

def _pv_place_best_short(pv, pv_row, target_gap=10):
    """Plaats korte pauze op eigen rij met voorkeur: exact +10 blokken na lange pauze-einde,
    anders +11, +12, ...; als dat niet past, probeer +9, +8, ...; valt terug op globale laatste lange-pauze-einde als geen eigen lange pauze."""
    naam = pv["naam"]
    # Als er al een korte pauze staat, niets doen
    if _pv_has_short_pause(naam, pv_row):
        return False

    # Hulpfuncties
    def is_toegestaan_pv_col(col):
        if len(open_uren) <= 6:
            return True
        return col not in get_verboden_korte_pauze_kolommen()

    # Zoek eigen lange pauze-einde op deze rij
    lange_blok_einde = None
    i = 0
    while i < len(pauze_cols)-1:
        c1 = pauze_cols[i]
        c2 = pauze_cols[i+1]
        if ws_pauze.cell(pv_row, c1).value == naam and ws_pauze.cell(pv_row, c2).value == naam:
            lange_blok_einde = i+1
            # ga door; kies de laatste indien meerdere
            i += 2
        else:
            i += 1

    # Geen eigen lange pauze: kies een geldige plek op eigen rij (liefst vroegste index >= target_gap)
    if lange_blok_einde is None:
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        candidates = []  # (prefer, idx, col, uur)
        for i, col in enumerate(pauze_cols):
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if not is_toegestaan_pv_col(col):
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            prefer = 1 if i >= target_gap else 0
            candidates.append((prefer, i, col, uur))
        if not candidates:
            return False
        # Kies met voorkeur voor index >= target_gap; binnen die groep de laatste (grootste index) om niet te vroeg te vallen
        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        _pref, i, col, uur = candidates[0]
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            return False
        # Voor PV-korte pauzes: laat het vakje erboven leeg
        ws_pauze.cell(pv_row-1, col).value = None
        ws_pauze.cell(pv_row-1, col).alignment = center_align
        ws_pauze.cell(pv_row-1, col).border = thin_border
        cel = ws_pauze.cell(pv_row, col)
        cel.value = naam
        cel.fill = lichtpaars_fill
        cel.alignment = center_align
        cel.border = thin_border
        return True
    else:
        anchor_idx = lange_blok_einde

    if anchor_idx is None or anchor_idx < 0:
        # Geen anchor beschikbaar: kies de eerste toegestane lege cel op eigen rij (zeldzaam)
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        for i, col in enumerate(pauze_cols):
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if is_toegestaan_pv_col(col) and ws_pauze.cell(pv_row, col).value in [None, ""]:
                # schrijf bovenliggende attr
                attr = vind_attractie_op_uur(naam, uur)
                if not attr:
                    continue
                # Voor PV-korte pauzes: laat het vakje erboven leeg
                ws_pauze.cell(pv_row-1, col).value = None
                ws_pauze.cell(pv_row-1, col).alignment = center_align
                ws_pauze.cell(pv_row-1, col).border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.fill = lichtpaars_fill
                cel.alignment = center_align
                cel.border = thin_border
                return True
        return False

    # Eerst exact +10 blokken, dan +11, +12, ...
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) > 2:
        verboden_uren = {werk_uren[0], werk_uren[-1]}
    else:
        verboden_uren = set(werk_uren)
    for d in range(target_gap, len(pauze_cols)-anchor_idx):
        idx = anchor_idx + d
        if idx >= len(pauze_cols):
            break
        col = pauze_cols[idx]
        if not is_toegestaan_pv_col(col):
            continue
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            continue
        if ws_pauze.cell(pv_row, col).value in [None, ""]:
            # schrijf bovenliggende attr
            attr = vind_attractie_op_uur(naam, uur)
            if not attr:
                continue
            # Voor PV-korte pauzes: laat het vakje erboven leeg
            ws_pauze.cell(pv_row-1, col).value = None
            ws_pauze.cell(pv_row-1, col).alignment = center_align
            ws_pauze.cell(pv_row-1, col).border = thin_border
            cel = ws_pauze.cell(pv_row, col)
            cel.value = naam
            cel.fill = lichtpaars_fill
            cel.alignment = center_align
            cel.border = thin_border
            return True

    # Dan lagere alternatieven: +9, +8, ... +1
    for d in range(target_gap-1, 0, -1):
        idx = anchor_idx + d
        if 0 <= idx < len(pauze_cols):
            col = pauze_cols[idx]
            if not is_toegestaan_pv_col(col):
                continue
            header = ws_pauze.cell(1, col).value
            uur = parse_header_uur(header)
            if uur not in werk_uren or uur in verboden_uren:
                continue
            if ws_pauze.cell(pv_row, col).value in [None, ""]:
                attr = vind_attractie_op_uur(naam, uur)
                if not attr:
                    continue
                # Voor PV-korte pauzes: laat het vakje erboven leeg
                ws_pauze.cell(pv_row-1, col).value = None
                ws_pauze.cell(pv_row-1, col).alignment = center_align
                ws_pauze.cell(pv_row-1, col).border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.fill = lichtpaars_fill
                cel.alignment = center_align
                cel.border = thin_border
                return True

    return False

for pv, pv_row in pv_rows:
    _pv_place_best_short(pv, pv_row, target_gap=8)


# ---- Korte pauze voor ALLE studenten (ook <=6u, behalve pauzevlinders en lange werkers) ----
# --- Houd bij wie al een korte pauze heeft gekregen ---
korte_pauze_ontvangers = set()
# Zoek alle namen die al een korte pauze hebben in het pauzeoverzicht
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())



# ---- Korte pauze voor ALLE studenten (ook <=6u, behalve pauzevlinders en lange werkers) ----
# --- Houd bij wie al een korte pauze heeft gekregen ---
korte_pauze_ontvangers = set()
# Zoek alle namen die al een korte pauze hebben in het pauzeoverzicht
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())


# Nieuwe logica: eerlijke verdeling van korte pauzes over pauzevlinders
from collections import Counter

# Tel per pauzevlinder het aantal korte pauzes dat al is toegekend (EXTRA niet meetellen)
pv_korte_pauze_count = Counter()
for pv, pv_row in pv_rows:
    for col in pauze_cols:
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # Check of het een korte pauze is (enkel blok, niet dubbel)
            idx = pauze_cols.index(col)
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                # Kijk naar bovenliggende attractie; tel niet als dit 'Extra' is of leeg (zoals bij PV zelf)
                attr_above = ws_pauze.cell(pv_row-1, col).value
                if attr_above and normalize_attr(attr_above) != 'extra':
                    pv_korte_pauze_count[pv["naam"]] += 1

niet_geplaatste_korte_pauze = []

# --- NIEUW: Eerst korte pauzes toewijzen aan iedereen met een LANGE pauze,
# in volgorde van wie het LAATST z'n lange pauze had ---

def _has_long_pause(naam):
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols[:-1]):
            if ws_pauze.cell(pv_row, col).value == naam and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam:
                return True
    return False

def _last_long_pause_end_index(naam):
    """Geef de hoogste kolomindex (in pauze_cols) die een tweede helft is van een dubbele blok voor deze naam; -1 indien geen lange pauze."""
    last_idx = -1
    for _pv, pv_row in pv_rows:
        for idx in range(len(pauze_cols)-1):
            c1 = pauze_cols[idx]
            c2 = pauze_cols[idx+1]
            if ws_pauze.cell(pv_row, c1).value == naam and ws_pauze.cell(pv_row, c2).value == naam:
                last_idx = max(last_idx, idx+1)
    return last_idx

def _has_short_pause(naam):
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            if ws_pauze.cell(pv_row, col).value == naam:
                # korte pauze: geen buur met dezelfde naam
                left_same = (idx > 0 and ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam)
                right_same = (idx+1 < len(pauze_cols) and ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam)
                if not left_same and not right_same:
                    return True
    return False

def _place_short_pause_for(naam):
    if _has_short_pause(naam):
        return True
    werk_uren = get_student_work_hours(naam)
    if not werk_uren:
        return False
    verboden_uren = {werk_uren[0], werk_uren[-1]} if len(werk_uren) > 2 else set(werk_uren)
    # Zoek anker = einde van eigen lange pauze (kolomindex in pauze_cols)
    def _last_long_end_index_for(naam):
        best = -1
        for _pv, _row in pv_rows:
            for idx in range(len(pauze_cols)-1):
                if ws_pauze.cell(_row, pauze_cols[idx]).value == naam and ws_pauze.cell(_row, pauze_cols[idx+1]).value == naam:
                    best = max(best, idx+1)
        return best
    anchor = _last_long_end_index_for(naam)

    # Helper om te checken en plaatsen op gewenste col
    def _try_place_at_col(col):
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            return False
        if not is_korte_pauze_toegestaan_col(col, naam):
            return False
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            return False
        # verzamel geldige (pv_row) volgens regels
        rows = []
        for pv, pv_row in pv_rows:
            if is_pauzevlinder(naam) and pv["naam"] != naam:
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            rows.append((pv, pv_row))
        if not rows:
            return False
        # fairness op pv-rijen
        rows.sort(key=lambda r: pv_korte_pauze_count[r[0]["naam"]])
        pv, pv_row = rows[0]
        # Voor PV-korte pauzes: laat het vakje erboven leeg
        if is_pauzevlinder(naam):
            ws_pauze.cell(pv_row-1, col).value = None
        else:
            ws_pauze.cell(pv_row-1, col).value = attr
        ws_pauze.cell(pv_row-1, col).alignment = center_align
        ws_pauze.cell(pv_row-1, col).border = thin_border
        cel = ws_pauze.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = lichtpaars_fill
        # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
        if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
            pv_korte_pauze_count[pv["naam"]] += 1
        return True

    # Als anchor bestaat, probeer exact +10, dan groter, anders lagere alternatieven
    if anchor >= 0:
        # +10 en verder
        for d in range(10, len(pauze_cols)-anchor):
            if _try_place_at_col(pauze_cols[anchor + d]):
                return True
        # lager
        for d in range(9, 0, -1):
            idx = anchor + d
            if 0 <= idx < len(pauze_cols) and _try_place_at_col(pauze_cols[idx]):
                return True

    # Als geen anchor of niets gevonden: val terug op fairness, maar zonder links-bias
    # Kies uit alle geldige (pv_row, col) paren, sorteer op pv fairness en dan op kolomindex die het verst ligt van begin (rechts-bias)
    pairs = []
    for col in pauze_cols:
        if not is_korte_pauze_toegestaan_col(col, naam):
            continue
        header = ws_pauze.cell(1, col).value
        uur = parse_header_uur(header)
        if uur not in werk_uren or uur in verboden_uren:
            continue
        attr = vind_attractie_op_uur(naam, uur)
        if not attr:
            continue
        for pv, pv_row in pv_rows:
            if is_pauzevlinder(naam) and pv["naam"] != naam:
                continue
            if ws_pauze.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                continue
            pairs.append((pv, pv_row, col))
    if not pairs:
        return False
    pairs.sort(key=lambda x: (pv_korte_pauze_count[x[0]["naam"]], -pauze_cols.index(x[2])))
    pv, pv_row, col = pairs[0]
    attr = vind_attractie_op_uur(naam, parse_header_uur(ws_pauze.cell(1, col).value))
    ws_pauze.cell(pv_row-1, col).value = attr
    ws_pauze.cell(pv_row-1, col).alignment = center_align
    ws_pauze.cell(pv_row-1, col).border = thin_border
    cel = ws_pauze.cell(pv_row, col)
    cel.value = naam
    cel.alignment = center_align
    cel.border = thin_border
    cel.fill = lichtpaars_fill
    # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
    if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
        pv_korte_pauze_count[pv["naam"]] += 1
    return True

# verzamel alle namen met een lange pauze en sorteer op laatste einde (desc)
names_with_long = []
alle_studenten_namen = {s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) >= 4}
for naam in alle_studenten_namen:
    if _has_long_pause(naam):
        end_idx = _last_long_pause_end_index(naam)
        names_with_long.append((end_idx, naam))
names_with_long.sort(reverse=True)  # laatste eerst

for _end, naam in names_with_long:
    _place_short_pause_for(naam)

# Zorg dat latere rondes deze personen overslaan: recompute korte_pauze_ontvangers nu
korte_pauze_ontvangers = set()
for pv, pv_row in pv_rows:
    for idx, col in enumerate(pauze_cols):
        cel = ws_pauze.cell(pv_row, col)
        if cel.value and str(cel.value).strip() != "":
            # korte pauze = enkel blok
            is_lange = False
            if idx+1 < len(pauze_cols):
                next_col = pauze_cols[idx+1]
                cel_next = ws_pauze.cell(pv_row, next_col)
                if cel_next.value == cel.value:
                    is_lange = True
            if idx > 0:
                prev_col = pauze_cols[idx-1]
                prev_cel = ws_pauze.cell(pv_row, prev_col)
                if prev_cel.value == cel.value:
                    is_lange = True
            if not is_lange:
                korte_pauze_ontvangers.add(str(cel.value).strip())

# Bepaal wie geen lange pauze heeft gekregen
studenten_zonder_lange_pauze = []
for s in studenten:
    naam = s["naam"]
    heeft_lange = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        heeft_lange = True
                        break
        if heeft_lange:
            break
    if not heeft_lange:
        studenten_zonder_lange_pauze.append(s)

# Eerst: korte pauze toewijzen aan studenten zonder lange pauze
def korte_pauze_toewijzen(studenten_lijst):
    for s in studenten_lijst:
        if s["naam"] in korte_pauze_ontvangers or _has_short_pause(s["naam"]):
            continue
        naam = s["naam"]
        werk_uren = get_student_work_hours(naam)
        if len(werk_uren) > 2:
            verboden_uren = {werk_uren[0], werk_uren[-1]}
        else:
            verboden_uren = set(werk_uren)
        pauze_cols_sorted = sorted(pauze_cols)
        geplaatst = False
        for uur in random.sample(werk_uren, len(werk_uren)):
            if uur in verboden_uren:
                continue
            attr = vind_attractie_op_uur(naam, uur)
            if not attr:
                continue
            geldige_slots = []
            for (pv, pv_row) in pv_rows:
                # Pauzevlinders: enkel op eigen rij
                if is_pauzevlinder(naam) and pv["naam"] != naam:
                    continue
                for col in pauze_cols:
                    col_header = ws_pauze.cell(1, col).value
                    col_uur = parse_header_uur(col_header)
                    if col_uur != uur:
                        continue
                    if not is_korte_pauze_toegestaan_col(col, naam):
                        continue
                    if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
                        continue
                    cel = ws_pauze.cell(pv_row, col)
                    if cel.value in [None, ""]:
                        geldige_slots.append((pv, pv_row, col))
            geldige_slots.sort(key=lambda slot: pv_korte_pauze_count[slot[0]["naam"]])
            for (pv, pv_row, col) in geldige_slots:
                boven_cel = ws_pauze.cell(pv_row - 1, col)
                # PV korte pauze: laat boven leeg
                boven_cel.value = None if is_pauzevlinder(naam) else attr
                boven_cel.alignment = center_align
                boven_cel.border = thin_border
                cel = ws_pauze.cell(pv_row, col)
                cel.value = naam
                cel.alignment = center_align
                cel.border = thin_border
                cel.fill = lichtpaars_fill
                korte_pauze_ontvangers.add(naam)
                # Niet meetellen als dit een EXTRA overname is of als de pauze voor een PV zelf is
                if (not is_pauzevlinder(naam)) and (normalize_attr(attr) != 'extra'):
                    pv_korte_pauze_count[pv["naam"]] += 1
                geplaatst = True
                break
            if geplaatst:
                break
        if not geplaatst:
            niet_geplaatste_korte_pauze.append(naam)

korte_pauze_toewijzen(studenten_zonder_lange_pauze)
# Daarna: de rest
korte_pauze_toewijzen([s for s in studenten if s not in studenten_zonder_lange_pauze])
korte_pauze_toewijzen([s for s in studenten if s not in studenten_zonder_lange_pauze])

# --- Iteratief wisselen: studenten zonder korte pauze proberen te ruilen met anderen (geen pauzevlinders) ---

def vind_korte_pauze_cell(naam):
    """Vind (pv_row, col) van de korte pauze van deze student, of None."""
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of het een korte pauze is (enkel blok, niet dubbel)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    return (pv_row, col)
    return None

def kan_student_korte_pauze_op_plek(naam, pv_row, col):
    """Check of student naam op deze plek een korte pauze mag hebben."""
    # Mag niet op pauzevlinder-rij
    if is_pauzevlinder(naam):
        return False
    # Moet werken op dit uur
    col_header = ws_pauze.cell(1, col).value
    col_uur = parse_header_uur(col_header)
    werk_uren = get_student_work_hours(naam)
    if col_uur not in werk_uren:
        return False
    # Niet in eerste/laatste werkuur
    if len(werk_uren) > 2:
        if col_uur == werk_uren[0] or col_uur == werk_uren[-1]:
            return False
    # Attractie moet kloppen
    attr = vind_attractie_op_uur(naam, col_uur)
    if not attr:
        return False
    # Pauzevlinder moet deze attractie kunnen
    pv = None
    for pv_obj, row in pv_rows:
        if row == pv_row:
            pv = pv_obj
            break
    if not pv:
        return False
    if not pv_kan_attr(pv, attr) and not is_student_extra(naam):
        return False
    # Kolom moet korte pauze toestaan
    if not is_korte_pauze_toegestaan_col(col, naam):
        return False
    return True

# Verzamel actuele lijst van studenten zonder korte pauze
werkende_studenten = [s for s in studenten if student_totalen.get(s["naam"], 0) >= 4 and not is_pauzevlinder(s["naam"])]
studenten_zonder_korte_pauze = []
for s in werkende_studenten:
    naam = s["naam"]
    heeft_korte = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of GEEN dubbele blok (dus geen lange pauze)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    heeft_korte = True
                    break
        if heeft_korte:
            break
    if not heeft_korte:
        studenten_zonder_korte_pauze.append(naam)

max_wissel_passes = 10
for _ in range(max_wissel_passes):
    if not studenten_zonder_korte_pauze:
        break
    verbeterd = False
    for naam_zonder in studenten_zonder_korte_pauze:
        # Probeer te ruilen met een student die wél een korte pauze heeft (geen pauzevlinder)
        for s in werkende_studenten:
            naam_met = s["naam"]
            if naam_met == naam_zonder:
                continue
            if naam_met in studenten_zonder_korte_pauze:
                continue
            # Vind de korte pauze van deze student
            plek = vind_korte_pauze_cell(naam_met)
            if not plek:
                continue
            pv_row, col = plek
            # Mag naam_zonder op deze plek een korte pauze hebben?
            if not kan_student_korte_pauze_op_plek(naam_zonder, pv_row, col):
                continue
            # Bepaal attractie voor naam_zonder op deze plek
            col_header = ws_pauze.cell(1, col).value
            col_uur = parse_header_uur(col_header)
            attr_zonder = vind_attractie_op_uur(naam_zonder, col_uur)
            if not attr_zonder:
                continue
            # Mag naam_met elders een korte pauze krijgen?
            # Zoek alternatieve plek voor naam_met
            gevonden = False
            for pv2, pv_row2 in pv_rows:
                if is_pauzevlinder(naam_met):
                    continue
                for col2 in pauze_cols:
                    if (pv_row2, col2) == (pv_row, col):
                        continue
                    cel2 = ws_pauze.cell(pv_row2, col2)
                    if cel2.value not in [None, ""]:
                        continue
                    if not kan_student_korte_pauze_op_plek(naam_met, pv_row2, col2):
                        continue
                    # Bepaal attractie voor naam_met op nieuwe plek
                    col2_header = ws_pauze.cell(1, col2).value
                    col2_uur = parse_header_uur(col2_header)
                    attr_met = vind_attractie_op_uur(naam_met, col2_uur)
                    if not attr_met:
                        continue
                    # Wissel uitvoeren
                    # 1. naam_met uit oude plek halen
                    ws_pauze.cell(pv_row, col).value = None
                    ws_pauze.cell(pv_row, col).fill = naam_leeg_fill
                    ws_pauze.cell(pv_row-1, col).value = None
                    # 2. naam_zonder op deze plek zetten
                    ws_pauze.cell(pv_row, col).value = naam_zonder
                    ws_pauze.cell(pv_row, col).fill = lichtpaars_fill
                    ws_pauze.cell(pv_row, col).alignment = center_align
                    ws_pauze.cell(pv_row, col).border = thin_border
                    ws_pauze.cell(pv_row-1, col).value = attr_zonder
                    ws_pauze.cell(pv_row-1, col).alignment = center_align
                    ws_pauze.cell(pv_row-1, col).border = thin_border
                    # 3. naam_met op nieuwe plek zetten
                    ws_pauze.cell(pv_row2, col2).value = naam_met
                    ws_pauze.cell(pv_row2, col2).fill = lichtpaars_fill
                    ws_pauze.cell(pv_row2, col2).alignment = center_align
                    ws_pauze.cell(pv_row2, col2).border = thin_border
                    ws_pauze.cell(pv_row2-1, col2).value = attr_met
                    ws_pauze.cell(pv_row2-1, col2).alignment = center_align
                    ws_pauze.cell(pv_row2-1, col2).border = thin_border
                    verbeterd = True
                    gevonden = True
                    break
                if gevonden:
                    break
            if verbeterd:
                break
        if verbeterd:
            break
    # Update lijst van studenten zonder korte pauze
    studenten_zonder_korte_pauze = []
    for s in werkende_studenten:
        naam = s["naam"]
        heeft_korte = False
        for pv, pv_row in pv_rows:
            for idx, col in enumerate(pauze_cols):
                cel = ws_pauze.cell(pv_row, col)
                if cel.value == naam:
                    # Check of GEEN dubbele blok (dus geen lange pauze)
                    is_lange = False
                    if idx+1 < len(pauze_cols):
                        next_col = pauze_cols[idx+1]
                        cel_next = ws_pauze.cell(pv_row, next_col)
                        if cel_next.value == naam:
                            is_lange = True
                    if idx > 0:
                        prev_col = pauze_cols[idx-1]
                        prev_cel = ws_pauze.cell(pv_row, prev_col)
                        if prev_cel.value == naam:
                            is_lange = True
                    if not is_lange:
                        heeft_korte = True
                        break
            if heeft_korte:
                break
        if not heeft_korte:
            studenten_zonder_korte_pauze.append(naam)
    if not verbeterd:
        break  # geen verbetering meer mogelijk

# Iteratieve optimalisatie: verschuif korte pauzes van "rijke" naar "arme" pauzevlinders
max_opt_passes = 10
for _ in range(max_opt_passes):
    # Zoek max en min aantal korte pauzes
    if not pv_korte_pauze_count:
        break
    max_pv = max(pv_korte_pauze_count, key=lambda k: pv_korte_pauze_count[k])
    min_pv = min(pv_korte_pauze_count, key=lambda k: pv_korte_pauze_count[k])
    if pv_korte_pauze_count[max_pv] - pv_korte_pauze_count[min_pv] <= 1:
        break  # verdeling is al redelijk
    # Zoek een korte pauze van max_pv die overgezet kan worden naar min_pv
    found = False
    for col in pauze_cols:
        pv_row_max = next((row for pv, row in pv_rows if pv["naam"] == max_pv), None)
        pv_row_min = next((row for pv, row in pv_rows if pv["naam"] == min_pv), None)
        if pv_row_max is None or pv_row_min is None:
            continue
        cel_max = ws_pauze.cell(pv_row_max, col)
        naam = cel_max.value
        if not naam or str(naam).strip() == "":
            continue
        # Check of het een korte pauze is (enkel blok, niet dubbel)
        idx = pauze_cols.index(col)
        is_lange = False
        if idx+1 < len(pauze_cols):
            next_col = pauze_cols[idx+1]
            cel_next = ws_pauze.cell(pv_row_max, next_col)
            if cel_next.value == cel_max.value:
                is_lange = True
        if idx > 0:
            prev_col = pauze_cols[idx-1]
            prev_cel = ws_pauze.cell(pv_row_max, prev_col)
            if prev_cel.value == cel_max.value:
                is_lange = True
        if is_lange:
            continue
        # Mag min_pv deze attractie overnemen?
        attr = ws_pauze.cell(pv_row_max-1, col).value
        if not pv_kan_attr(next(pv for pv, _ in pv_rows if pv["naam"] == min_pv), attr):
            continue
        # Is de cel bij min_pv vrij?
        cel_min = ws_pauze.cell(pv_row_min, col)
        if cel_min.value not in [None, ""]:
            continue
        # Wissel de korte pauze van max_pv naar min_pv
        cel_min.value = naam
        cel_min.alignment = center_align
        cel_min.border = thin_border
        cel_min.fill = lichtpaars_fill
        ws_pauze.cell(pv_row_min-1, col).value = attr
        ws_pauze.cell(pv_row_min-1, col).alignment = center_align
        ws_pauze.cell(pv_row_min-1, col).border = thin_border
        cel_max.value = None
        ws_pauze.cell(pv_row_max-1, col).value = None
        # Pas telling aan enkel als dit geen EXTRA-overname is
        if attr and normalize_attr(attr) != 'extra':
            pv_korte_pauze_count[max_pv] -= 1
            pv_korte_pauze_count[min_pv] += 1
        found = True
        break
    if not found:
        break  # geen verschuiving meer mogelijk



# --- Iteratieve optimalisatie: verdeel lange pauzes zo eerlijk mogelijk over pauzevlinders ---

max_opt_passes_lange = 10
from collections import Counter
for _ in range(max_opt_passes_lange):
    pass  # (oude optimalisatie-code is verwijderd, want niet meer nodig)

# --- Pauzevlinders met >6u: altijd lange pauze in eigen rij ---
import random
# --- Pauzevlinders met >6u: altijd lange pauze in eigen rij, gespreid over eerste drie pauzeuren ---
for pv, pv_row in pv_rows:
    naam = pv["naam"]
    werk_uren = get_student_work_hours(naam)
    if len(werk_uren) > 6:
        # Alleen de eerste 11 kwartieren (indices 0 t/m 10) zijn toegestaan voor lange pauzes
        if heeft_al_lange_pauze(naam):
            continue
        halve_uren = []  # lijst van (idx, col1, col2)
        max_start_idx = min(8, len(pauze_cols)-2)  # idx 0 t/m 10 zijn halve uren binnen eerste 11 kwartieren
        for idx in range(max_start_idx+1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx+1]
            col1_header = ws_pauze.cell(1, col1).value
            # Alleen starten op heel of half uur
            try:
                min1 = int(str(col1_header).split('u')[1]) if 'u' in str(col1_header) and len(str(col1_header).split('u')) > 1 else 0
            except:
                min1 = 0
            if min1 not in (0, 30):
                continue
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            if cel1.value in [None, ""] and cel2.value in [None, ""]:
                halve_uren.append((idx, col1, col2))
        # Shuffle de halve uren
        random.shuffle(halve_uren)
        # Probeer in geshuffelde volgorde een lange pauze te plaatsen
        geplaatst = False
        for idx, col1, col2 in halve_uren:
            cel1 = ws_pauze.cell(pv_row, col1)
            cel2 = ws_pauze.cell(pv_row, col2)
            if cel1.value in [None, ""] and cel2.value in [None, ""] and not heeft_al_lange_pauze(naam):
                cel1.value = naam
                cel2.value = naam
                cel1.alignment = center_align
                cel2.alignment = center_align
                cel1.border = thin_border
                cel2.border = thin_border
                cel1.fill = lichtgroen_fill
                cel2.fill = lichtgroen_fill
                geplaatst = True
                break
        # Indien geen plek gevonden, doe niets (komt zelden voor)



output = BytesIO()

# -----------------------------
# ENFORCE: Korte pauzes minimaal 10 blokken uit elkaar (maximaliseer anders)
# -----------------------------

# We beschouwen alle pauzecellen (korte én beide helften van lange pauzes).
# We verplaatsen alleen korte pauzes (enkelblok) naar lege geschikte slots.

def _get_student_pause_cols(naam):
    cols = []
    for _pv, pv_row in pv_rows:
        for col in pauze_cols:
            if ws_pauze.cell(pv_row, col).value == naam:
                cols.append(col)
    return sorted(cols)

def _get_student_short_pause_positions(naam):
    pos = []  # lijst van (pv_row, col)
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value != naam:
                continue
            # controleer of GEEN deel van een dubbele blok (lange pauze)
            is_dubbel = False
            if idx+1 < len(pauze_cols):
                if ws_pauze.cell(pv_row, pauze_cols[idx+1]).value == naam:
                    is_dubbel = True
            if idx > 0:
                if ws_pauze.cell(pv_row, pauze_cols[idx-1]).value == naam:
                    is_dubbel = True
            if not is_dubbel:
                pos.append((pv_row, col))
    return pos

def _min_gap(cols):
    if len(cols) < 2:
        return 10**9
    cols = sorted(cols)
    mg = min(cols[i+1]-cols[i] for i in range(len(cols)-1))
    return mg

def _can_place_short_pause(naam, pv_row, col):
    # cel moet leeg zijn
    if ws_pauze.cell(pv_row, col).value not in [None, ""]:
        return False
    # kolom moet korte pauze toelaten
    if not is_korte_pauze_toegestaan_col(col, naam):
        return False
    # student moet werken op dit uur en niet in eerste/laatste werkuur
    header = ws_pauze.cell(1, col).value
    uur = parse_header_uur(header)
    if uur is None:
        return False
    werk_uren = get_student_work_hours(naam)
    if uur not in werk_uren:
        return False
    if len(werk_uren) > 2:
        if uur == werk_uren[0] or uur == werk_uren[-1]:
            return False
    else:
        # bij 1-2 uren: geen korte pauze plannen
        return False
    # pauzevlinder-capability: pauzevlinder op pv_row moet attractie kunnen
    attr = vind_attractie_op_uur(naam, uur)
    if not attr:
        return False
    pv_obj = None
    for pv, row in pv_rows:
        if row == pv_row:
            pv_obj = pv
            break
    if pv_obj is None:
        return False
    if not pv_kan_attr(pv_obj, attr) and not is_student_extra(naam):
        return False
    return True

def _move_short_pause(naam, from_row, from_col, to_row, to_col):
    # leegmaken bron
    ws_pauze.cell(from_row, from_col).value = None
    ws_pauze.cell(from_row-1, from_col).value = None
    # invullen doel
    header = ws_pauze.cell(1, to_col).value
    uur = parse_header_uur(header)
    attr = vind_attractie_op_uur(naam, uur)
    # Voor PV korte pauze: laat boven leeg
    ws_pauze.cell(to_row-1, to_col).value = None if is_pauzevlinder(naam) else attr
    ws_pauze.cell(to_row-1, to_col).alignment = center_align
    ws_pauze.cell(to_row-1, to_col).border = thin_border
    ws_pauze.cell(to_row, to_col).value = naam
    ws_pauze.cell(to_row, to_col).alignment = center_align
    ws_pauze.cell(to_row, to_col).border = thin_border

def _recolor_pauze_sheet():
    # Kleur korte pauze paars, lange (dubbel) groen, leeg lichtblauw
    for _pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            val = cel.value
            if val in [None, ""]:
                cel.fill = naam_leeg_fill
                continue
            # lange pauze als naastliggende cel dezelfde naam heeft
            is_dubbel = False
            if idx+1 < len(pauze_cols):
                c2 = ws_pauze.cell(pv_row, pauze_cols[idx+1])
                if c2.value == val:
                    cel.fill = lichtgroen_fill
                    c2.fill = lichtgroen_fill
                    is_dubbel = True
                    continue
            if idx > 0:
                c1 = ws_pauze.cell(pv_row, pauze_cols[idx-1])
                if c1.value == val:
                    # onderdeel van reeds gekleurde lange pauze
                    continue
            # anders korte pauze
            cel.fill = lichtpaars_fill

def _enforce_min_gap_for_short_pauses(desired_gap=10, max_passes=5):
    changed = False
    for _ in range(max_passes):
        improved = False
        # itereren over alle studenten met minstens 1 korte pauze
        alle_namen = {s["naam"] for s in studenten if student_totalen.get(s["naam"], 0) > 0}
        for naam in alle_namen:
            short_pos = _get_student_short_pause_positions(naam)
            if not short_pos:
                continue
            all_cols = _get_student_pause_cols(naam)
            # bekijk elke korte pauze afzonderlijk
            for (from_row, from_col) in short_pos:
                # huidige minimale gap voor deze korte pauze
                other_cols = [c for c in all_cols if c != from_col]
                cur_gap = min((abs(from_col - c) for c in other_cols), default=10**9)
                if cur_gap >= desired_gap:
                    continue  # al goed
                # zoek beste lege slot
                best = None  # (best_gap, to_row, to_col)
                for _pv, pv_row in pv_rows:
                    # Pauzevlinders: korte pauze enkel op eigen rij verplaatsen/plaatsen
                    if is_pauzevlinder(naam) and _pv["naam"] != naam:
                        continue
                    for col in pauze_cols:
                        if not _can_place_short_pause(naam, pv_row, col):
                            continue
                        # gap als we hierheen verplaatsen
                        new_gap = min((abs(col - c) for c in other_cols), default=10**9)
                        if (best is None) or (new_gap > best[0]) or (new_gap == best[0] and col < best[2]):
                            best = (new_gap, pv_row, col)
                            # early exit als we desired halen
                            if new_gap >= desired_gap:
                                break
                    if best and best[0] >= desired_gap:
                        break
                if best and best[0] > cur_gap:
                    _move_short_pause(naam, from_row, from_col, best[1], best[2])
                    improved = True
                    changed = True
                    # update caches voor volgende iteraties
                    all_cols = _get_student_pause_cols(naam)
        if not improved:
            break
    # herkleuren na eventuele wijzigingen
    if changed:
        _recolor_pauze_sheet()
    return changed

# Voer de enforce stap uit met doelafstand 10 blokken
_enforce_min_gap_for_short_pauses(desired_gap=10, max_passes=6)

# Optionele samenvatting in Streamlit
# Debug samenvatting (globale minimale pauze-afstand) verwijderd om UI schoon te houden.
# Indien opnieuw nodig: functie _global_min_gap_summary() herstellen.

# --- FEEDBACK SHEET ---
ws_feedback = wb_out.create_sheet("Feedback")
row_fb = 1

# 1. Lange werkers (>6u) zonder lange pauze
lange_werkers_zonder_lange_pauze = set()

def _heeft_lange_pauze_naam(naam: str) -> bool:
    """Zoek in ws_pauze of deze persoon een dubbele blok (lange pauze) heeft."""
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of volgende cel ook deze naam heeft (dubbele blok)
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        return True
    return False

# a) reguliere lange werkers
for s in lange_werkers:
    naam = s["naam"]
    if not _heeft_lange_pauze_naam(naam):
        lange_werkers_zonder_lange_pauze.add(naam)

# b) pauzevlinders die >6u werken meenemen
for pv, _pv_row in pv_rows:
    naam = pv["naam"]
    werk_uren = get_student_work_hours(naam) or []
    if len(werk_uren) > 6:
        if not _heeft_lange_pauze_naam(naam):
            lange_werkers_zonder_lange_pauze.add(naam)

ws_feedback.cell(row_fb, 1, "Lange werkers (>6u) zonder lange pauze:")
row_fb += 1
if lange_werkers_zonder_lange_pauze:
    for naam in sorted(lange_werkers_zonder_lange_pauze):
        ws_feedback.cell(row_fb, 1, naam)
        row_fb += 1
else:
    vinkje_cel = ws_feedback.cell(row_fb, 1, "✓")
    ws_feedback.cell(row_fb, 2, "Iedereen heeft een lange pauze gekregen.")
    from openpyxl.styles import PatternFill, Font
    vinkje_cel.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # opvallend groen
    vinkje_cel.font = Font(bold=True, color="006100")  # donkergroen
    row_fb += 1

# 2. Werkende studenten zonder korte pauze
werkende_studenten = [s for s in studenten if student_totalen.get(s["naam"], 0) >= 4]
studenten_zonder_korte_pauze = []
for s in werkende_studenten:
    naam = s["naam"]
    # Zoek in ws_pauze of deze student een korte pauze heeft (enkel blok, niet dubbel)
    heeft_korte = False
    for pv, pv_row in pv_rows:
        for idx, col in enumerate(pauze_cols):
            cel = ws_pauze.cell(pv_row, col)
            if cel.value == naam:
                # Check of GEEN dubbele blok (dus geen lange pauze)
                is_lange = False
                if idx+1 < len(pauze_cols):
                    next_col = pauze_cols[idx+1]
                    cel_next = ws_pauze.cell(pv_row, next_col)
                    if cel_next.value == naam:
                        is_lange = True
                if idx > 0:
                    prev_col = pauze_cols[idx-1]
                    prev_cel = ws_pauze.cell(pv_row, prev_col)
                    if prev_cel.value == naam:
                        is_lange = True
                if not is_lange:
                    heeft_korte = True
                    break
        if heeft_korte:
            break
    if not heeft_korte:
        studenten_zonder_korte_pauze.append(naam)

ws_feedback.cell(row_fb, 1, "Werkende studenten zonder korte pauze:")
row_fb += 1
if studenten_zonder_korte_pauze:
    for naam in studenten_zonder_korte_pauze:
        ws_feedback.cell(row_fb, 1, naam)
        row_fb += 1
else:
    vinkje_cel = ws_feedback.cell(row_fb, 1, "✓")
    ws_feedback.cell(row_fb, 2, "Iedereen heeft een korte pauze gekregen.")
    vinkje_cel.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    vinkje_cel.font = Font(bold=True, color="006100")
    row_fb += 1


##### EXTRA INFO TOEVOEGEN AAN PAUZEPLANNING (A12 e.v.)
##### -------------------------------------------------------------
### We gebruiken de 'Input' sheet van het geüploade bestand
### en de 'Pauzevlinders' sheet van het resultaat
ws_input_data = wb["Input"]
ws_pauze_sheet = wb_out["Pauzevlinders"]

### Definieer de witte achtergrond
witte_fill = PatternFill(start_color="FFFFFF", fill_type="solid")

# --- NIEUWE LOGICA VOOR BN15 VINKJE ---
# BN is de 66e kolom in Excel. We controleren cel BN15.
bn15_vinkje = ws_input_data.cell(row=15, column=66).value

if bn15_vinkje in [1, True, "WAAR", "X"]:
    # Loop door de rijen 15 tot en met 30 van de Input-sheet
    for i, input_rij in enumerate(range(15, 31)):
        # Kolom BO is de 67e kolom in Excel
        waarde = ws_input_data.cell(row=input_rij, column=67).value
        
        if waarde:
            # Schrijf de waarde naar kolom A van de pauzeplanning, beginnend bij rij 12
            target_rij = 12 + i
            cel = ws_pauze_sheet.cell(row=target_rij, column=1, value=waarde)
            cel.fill = witte_fill
            cel.border = thin_border
            cel.alignment = Alignment(horizontal="left", vertical="center")
# -------------------------------------



wb_out.save(output)
output.seek(0)  # Zorg dat lezen vanaf begin kan


#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# ============================================================
# DEEL 5: PP optie 2 + Feedback optie 2
# Alleen STAP 1 van de nieuwe logica
# Plaats dit volledig op het einde, net vóór het save-blok
# ============================================================

from collections import defaultdict, Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------
# Veilig bestaande sheets verwijderen indien nodig
# -----------------------------
for sheet_name in ["PP optie 2", "Feedback optie 2"]:
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])

# -----------------------------
# Basis referenties
# -----------------------------
ws_planning = wb_out["Planning"]
ws_pauze_basis = wb_out["Pauzevlinders"]

# -----------------------------
# Maak PP optie 2 als kopie van Pauzevlinders
# Dan blijft de opmaak exact hetzelfde
# -----------------------------
ws_pp2 = wb_out.copy_worksheet(ws_pauze_basis)
ws_pp2.title = "PP optie 2"

# -----------------------------
# Helpers
# -----------------------------

def pp2_is_minderjarig(naam):
    return "-18" in str(naam)


def pp2_parse_kwartier_header(header):
    """
    Zet '12u', '12u15', '12u30', '12u45' om naar minuten sinds 00:00.
    """
    if not header:
        return None
    s = str(header).strip().lower()
    if "u" not in s:
        return None
    parts = s.split("u", 1)
    try:
        uur = int(parts[0])
        mins = int(parts[1]) if parts[1] != "" else 0
        return uur * 60 + mins
    except:
        return None

def pp2_get_pauze_cols(ws_sheet):
    cols = []
    for col in range(2, ws_sheet.max_column + 1):
        header = ws_sheet.cell(1, col).value
        if header and "u" in str(header):
            cols.append(col)
    return cols

def pp2_get_pv_rows(ws_sheet, selected):
    """
    Geeft lijst van tuples: (pv_dict, naam_rij)
    waarbij naam_rij de rij is waar de naam van de pauzevlinder staat.
    """
    rows = []
    for pv in selected:
        found = None
        for r in range(2, ws_sheet.max_row + 1):
            val = ws_sheet.cell(r, 1).value
            if val and str(val).strip() == str(pv["naam"]).strip():
                found = r
                break
        if found is not None:
            rows.append((pv, found))
    return rows

def pp2_get_student_work_hours(naam):
    """
    Leest echte werkuren uit het werkblad Planning.
    """
    uren = set()
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        uur = parse_header_uur(header)
        if uur is None:
            continue
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                uren.add(uur)
                break
    return sorted(uren)

def pp2_is_first_or_last_work_hour(naam, kwartier_col, ws_sheet):
    """
    Checkt of dit kwartier in het eerste of laatste werkuur valt.
    """
    werk_uren = pp2_get_student_work_hours(naam)
    if not werk_uren:
        return True

    header = ws_sheet.cell(1, kwartier_col).value
    pauze_uur = parse_header_uur(header)
    if pauze_uur is None:
        return True

    return pauze_uur == werk_uren[0] or pauze_uur == werk_uren[-1]

def pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols):
    """
    Alle geldige kwartierkolommen voor korte pauze:
    - student werkt dat uur
    - niet in eerste of laatste werkuur
    """
    werk_uren = pp2_get_student_work_hours(naam)
    if len(werk_uren) < 4:
        return []

    first_hour = werk_uren[0]
    last_hour = werk_uren[-1]

    candidates = []
    for col in pauze_cols:
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
        if uur is None:
            continue
        if uur in werk_uren and uur != first_hour and uur != last_hour:
            candidates.append(col)

    return candidates

def pp2_choose_middle_col(naam, ws_sheet, pauze_cols):
    """
    Kies een kwartier zo goed mogelijk in het midden van de shift,
    rekening houdend met de toegelaten kwartieren.
    """
    candidates = pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols)
    if not candidates:
        return None

    werk_uren = pp2_get_student_work_hours(naam)
    shift_start = min(werk_uren) * 60
    shift_end = (max(werk_uren) + 1) * 60
    midpoint = (shift_start + shift_end) / 2

    best_col = None
    best_score = None

    for col in candidates:
        mins = pp2_parse_kwartier_header(ws_sheet.cell(1, col).value)
        if mins is None:
            continue
        score = abs(mins - midpoint)
        if best_score is None or score < best_score:
            best_score = score
            best_col = col

    return best_col

def pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
    """
    Een korte pauze mag alleen als:
    - student werkt in dat kwartier
    - niet in eerste of laatste werkuur
    - student op dat kwartier nog nergens anders in het pauzerooster staat
    """
    header = ws_sheet.cell(1, col).value
    uur = parse_header_uur(header)

    if uur is None:
        return False

    werk_uren = pp2_get_student_work_hours(naam)
    if not werk_uren:
        return False

    if uur not in werk_uren:
        return False

    eerste_uur = werk_uren[0]
    laatste_uur = werk_uren[-1]

    if uur == eerste_uur or uur == laatste_uur:
        return False

    if pp2_student_heeft_al_pauze_op_kolom(
        naam=naam,
        col=col,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows_pp2
    ):
        return False

    return True

def pp2_choose_middle_double_col_for_minor(naam, ws_sheet, pauze_cols):
    """
    Zoek startkolom voor 2 opeenvolgende kwartieren voor minderjarigen:
    - student werkt op beide kwartieren
    - student stopt om of voor 16u (dus laatste werkblok <= 15)
    - student werkt >4u en <=6u
    - start enkel op een half uur (:00 of :30)
    - zo vroeg mogelijk in de shift
    - beide cellen moeten geldig zijn volgens de gewone korte-pauze-regels
    """
    werk_uren = pp2_get_student_work_hours(naam)
    if not werk_uren:
        return None

    if len(werk_uren) <= 4 or len(werk_uren) > 6:
        return None

    if max(werk_uren) > 15:
        return None

    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        # moeten opeenvolgende kwartieren zijn
        if col2 != col1 + 1:
            continue

        header1 = ws_sheet.cell(1, col1).value
        uur1 = parse_header_uur(header1)
        if uur1 is None:
            continue

        # start enkel op heel uur of half uur
        header_text = str(header1).strip().lower()
        if not (header_text.endswith("u") or header_text.endswith("u30")):
            continue

        # beide kwartieren moeten geldig zijn volgens gewone korte-pauze-regels
        if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
            continue
        if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
            continue

        # beide kwartieren moeten effectief tijdens werkuren vallen
        uur2 = parse_header_uur(ws_sheet.cell(1, col2).value)
        if uur2 is None:
            continue

        if uur1 not in werk_uren or uur2 not in werk_uren:
            continue

        # eerste geldige optie meteen nemen
        return col1

    return None



def pp2_same_halfhour(col_a, col_b, ws_sheet):
    mins_a = pp2_parse_kwartier_header(ws_sheet.cell(1, col_a).value)
    mins_b = pp2_parse_kwartier_header(ws_sheet.cell(1, col_b).value)
    if mins_a is None or mins_b is None:
        return False
    return (mins_a // 30) == (mins_b // 30)

def pp2_choose_adjacent_same_halfhour(base_col, student_name, ws_sheet, pauze_cols, pv_name_row):
    """
    Tweede student van het duo moet naast de eerste zitten
    in hetzelfde halfuur, indien dat volgens de regels kan.
    """
    if base_col not in pauze_cols:
        return None

    idx = pauze_cols.index(base_col)
    opties = []

    if idx - 1 >= 0:
        opties.append(pauze_cols[idx - 1])
    if idx + 1 < len(pauze_cols):
        opties.append(pauze_cols[idx + 1])

    # Eerst alleen opties in hetzelfde halfuur
    opties = [c for c in opties if pp2_same_halfhour(base_col, c, ws_sheet)]

    for col in opties:
        # vak moet leeg zijn
        if ws_sheet.cell(pv_name_row, col).value not in [None, ""]:
            continue
        # niet in eerste/laatste werkuur van deze student
        if pp2_is_first_or_last_work_hour(student_name, col, ws_sheet):
            continue
        # student moet effectief dat uur werken
        uur = parse_header_uur(ws_sheet.cell(1, col).value)
        werk_uren = pp2_get_student_work_hours(student_name)
        if uur not in werk_uren:
            continue
        return col

    return None

def pp2_write_name(ws_sheet, row_name, col, naam):
    """
    Schrijf in PP optie 2:
    - bovenste vak: attractie waarop student dat moment staat
    - onderste vak: naam van student
    - korte pauze = paars
    - lange pauze = groen (voor later bruikbaar)
    """
    lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")

    # bepaal uur van deze kolom
    header = ws_sheet.cell(1, col).value
    uur = parse_header_uur(header)

    # attractie erboven invullen
    info_cel = ws_sheet.cell(row_name - 1, col)
    attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
    info_cel.value = attr if attr else ""
    info_cel.alignment = center_align
    info_cel.border = thin_border

    # naam invullen
    cel = ws_sheet.cell(row_name, col)
    cel.value = naam
    cel.alignment = center_align
    cel.border = thin_border

    # check of dit een lange of korte pauze is
    is_lange_pauze = False
    if col - 1 >= 2 and ws_sheet.cell(row_name, col - 1).value == naam:
        is_lange_pauze = True
    if col + 1 <= ws_sheet.max_column and ws_sheet.cell(row_name, col + 1).value == naam:
        is_lange_pauze = True

    cel.fill = lichtgroen_fill if is_lange_pauze else lichtpaars_fill

def pp2_clear_pauze_grid(ws_sheet, pv_rows, pauze_cols):
    """
    Wis enkel de effectieve pauzevakken:
    - rij erboven: attractie/info
    - naamrij: naam
    Kolom A en extra info lager op het blad blijven behouden.
    """
    leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    for pv, naam_rij in pv_rows:
        info_rij = naam_rij - 1
        for col in pauze_cols:
            # bovenste rij leegmaken
            ws_sheet.cell(info_rij, col).value = None
            ws_sheet.cell(info_rij, col).alignment = center_align
            ws_sheet.cell(info_rij, col).border = thin_border

            # naamrij leegmaken
            ws_sheet.cell(naam_rij, col).value = None
            ws_sheet.cell(naam_rij, col).alignment = center_align
            ws_sheet.cell(naam_rij, col).border = thin_border
            ws_sheet.cell(naam_rij, col).fill = leeg_fill


def pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_sheet, pv_rows):
    """
    True als deze student op deze kwartierkolom al ergens in het pauzerooster staat,
    ongeacht op welke pauzevlinder-rij.
    """
    for _pv, pv_row in pv_rows:
        if ws_sheet.cell(pv_row, col).value == naam:
            return True
    return False


def pp2_student_heeft_al_lange_pauze_op_blok(naam, col1, col2, ws_sheet, pv_rows):
    """
    True als deze student deze 2 kwartieren al ergens als lange pauze heeft staan.
    """
    for _pv, pv_row in pv_rows:
        if (
            ws_sheet.cell(pv_row, col1).value == naam and
            ws_sheet.cell(pv_row, col2).value == naam
        ):
            return True
    return False


# -----------------------------
# Vind de pauzevlinder-rijen in PP optie 2
# -----------------------------
pauze_cols_pp2 = pp2_get_pauze_cols(ws_pp2)
pv_rows_pp2 = pp2_get_pv_rows(ws_pp2, selected)

# Maak de grid leeg, maar behoud layout
pp2_clear_pauze_grid(ws_pp2, pv_rows_pp2, pauze_cols_pp2)


# -----------------------------
# STAP 1:
# Vroege stoppers (minstens 4u gewerkt en laatste werkblok <= 15)
# - minderjarige vroege stoppers: eerst halfuur zo vroeg mogelijk,
#   dan kwartier zo laat mogelijk, zo mogelijk op dezelfde PV-rij
# - gewone vroege stoppers: duo-logica zoals voorheen
# Excl. pauzevlinders zelf
# -----------------------------
pauzevlinder_namen_set = {pv["naam"] for pv in selected}

vroege_stoppers_gewoon = []
vroege_stoppers_minderjarig = []

for s in studenten:
    naam = s["naam"]

    if naam in pauzevlinder_namen_set:
        continue

    werk_uren = pp2_get_student_work_hours(naam)
    if len(werk_uren) < 4:
        continue

    laatste_werkblok = max(werk_uren)
    startuur = min(werk_uren)
    aantal_uren = len(werk_uren)

    if laatste_werkblok > 15:
        continue

    item = {
        "naam": naam,
        "werk_uren": werk_uren,
        "einduur": laatste_werkblok,
        "startuur": startuur,
        "aantal_uren": aantal_uren
    }

    if pp2_is_minderjarig(naam):
        vroege_stoppers_minderjarig.append(item)
    else:
        vroege_stoppers_gewoon.append(item)

# Sorteervolgorde: vroegst stoppend, vroegst beginnend, alfabetisch
vroege_stoppers_minderjarig.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))
vroege_stoppers_gewoon.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))

pp2_geplaatste_pauzes = []
pp2_niet_geplaatst = []

# -----------------------------
# STAP 1a: minderjarige vroege stoppers
# Pauze 1: halfuur (2 opeenvolgende kwartieren) zo vroeg mogelijk
# Pauze 2: kwartier zo laat mogelijk
# Beide pauzes: niet in eerste of laatste werkuur
# Pauze 2 zo mogelijk op dezelfde PV-rij als pauze 1
# -----------------------------
pp2_minderjarige_vroege_stopper_rij = {}

if pv_rows_pp2:
    for idx, item in enumerate(vroege_stoppers_minderjarig):
        naam = item["naam"]
        werk_uren = item["werk_uren"]
        eerste_uur = werk_uren[0]
        laatste_uur = werk_uren[-1]

        pv_index = idx % len(pv_rows_pp2)
        pv, pv_name_row = pv_rows_pp2[pv_index]
        pv_label = pv["naam"]

        # -- Pauze 1: halfuur zo vroeg mogelijk --
        col1_gekozen = None
        for i in range(len(pauze_cols_pp2) - 1):
            col1 = pauze_cols_pp2[i]
            col2 = pauze_cols_pp2[i + 1]

            # opeenvolgende kwartieren
            if col2 != col1 + 1:
                continue

            uur1 = parse_header_uur(ws_pp2.cell(1, col1).value)
            uur2 = parse_header_uur(ws_pp2.cell(1, col2).value)

            if uur1 is None or uur2 is None:
                continue

            # niet in eerste of laatste werkuur
            if uur1 == eerste_uur or uur1 == laatste_uur:
                continue
            if uur2 == eerste_uur or uur2 == laatste_uur:
                continue

            # student moet beide uren werken
            if uur1 not in werk_uren or uur2 not in werk_uren:
                continue

            # cellen moeten leeg zijn op deze PV-rij
            if ws_pp2.cell(pv_name_row, col1).value not in [None, ""]:
                continue
            if ws_pp2.cell(pv_name_row, col2).value not in [None, ""]:
                continue

            # eerste geldige optie nemen
            col1_gekozen = col1
            break

        if col1_gekozen is not None:
            col2_gekozen = col1_gekozen + 1
            pp2_write_name(ws_pp2, pv_name_row, col1_gekozen, naam)
            pp2_write_name(ws_pp2, pv_name_row, col2_gekozen, naam)
            pp2_minderjarige_vroege_stopper_rij[naam] = pv_name_row

            pp2_geplaatste_pauzes.append({
                "naam": naam,
                "pauzevlinder": pv_label,
                "tijd": f"{ws_pp2.cell(1, col1_gekozen).value}-{ws_pp2.cell(1, col2_gekozen).value}",
                "type": "minderjarig vroege stopper - halfuur"
            })
        else:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "geen geldig halfuur gevonden voor minderjarige vroege stopper (pauze 1)"
            })

        # -- Pauze 2: kwartier zo laat mogelijk, bij voorkeur zelfde PV-rij --
        vaste_rij = pp2_minderjarige_vroege_stopper_rij.get(naam)

        # Kandidaat-kolommen van achter naar voor
        kandidaten = list(reversed(pauze_cols_pp2))

        kort_geplaatst = False
        for gebruik_rij in ([vaste_rij] if vaste_rij else []) + [r for (_pv2, r) in pv_rows_pp2 if r != vaste_rij]:
            for col in kandidaten:
                uur = parse_header_uur(ws_pp2.cell(1, col).value)
                if uur is None:
                    continue

                # niet in eerste of laatste werkuur
                if uur == eerste_uur or uur == laatste_uur:
                    continue

                # student moet dat uur werken
                if uur not in werk_uren:
                    continue

                # cel moet leeg zijn
                if ws_pp2.cell(gebruik_rij, col).value not in [None, ""]:
                    continue

                # student mag op dit kwartier nog nergens staan
                if pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_pp2, pv_rows_pp2):
                    continue

                pp2_write_name(ws_pp2, gebruik_rij, col, naam)

                pp2_geplaatste_pauzes.append({
                    "naam": naam,
                    "pauzevlinder": ws_pp2.cell(gebruik_rij, 1).value or f"rij {gebruik_rij}",
                    "tijd": ws_pp2.cell(1, col).value,
                    "type": "minderjarig vroege stopper - kort kwartier"
                })

                kort_geplaatst = True
                break

            if kort_geplaatst:
                break

        if not kort_geplaatst:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "geen geldig kwartier gevonden voor minderjarige vroege stopper (pauze 2)"
            })

# -----------------------------
# STAP 1b: gewone vroege stoppers
# Inplannen per duo:
# 1-2 bij PV1, 3-4 bij PV2, 5-6 bij PV3, ...
# als er meer duo's zijn dan pauzevlinders, dan cyclisch verder
# Als de voorkeurs-PV-rij al bezet is op de gekozen kolom,
# worden andere PV-rijen geprobeerd.
# -----------------------------
duo_basis_col = {}
duo_basis_pv_row = {}

if pv_rows_pp2:
    for idx, item in enumerate(vroege_stoppers_gewoon):
        naam = item["naam"]

        duo_nummer = idx // 2
        pv_index_voorkeur = duo_nummer % len(pv_rows_pp2)

        # Eerste van het duo
        if idx % 2 == 0:
            gekozen_col = pp2_choose_middle_col(naam, ws_pp2, pauze_cols_pp2)

            if gekozen_col is None:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldige middenplek gevonden voor eerste van duo"
                })
                continue

            # Probeer eerst voorkeurs-PV-rij, daarna de rest
            pv_volgorde = (
                [pv_rows_pp2[pv_index_voorkeur]]
                + [r for i, r in enumerate(pv_rows_pp2) if i != pv_index_voorkeur]
            )

            geplaatst_eerste = False
            for pv, pv_name_row in pv_volgorde:
                if ws_pp2.cell(pv_name_row, gekozen_col).value not in [None, ""]:
                    continue

                pp2_write_name(ws_pp2, pv_name_row, gekozen_col, naam)
                duo_basis_col[duo_nummer] = gekozen_col
                duo_basis_pv_row[duo_nummer] = pv_name_row

                pp2_geplaatste_pauzes.append({
                    "naam": naam,
                    "pauzevlinder": pv["naam"],
                    "tijd": ws_pp2.cell(1, gekozen_col).value,
                    "type": "eerste van duo"
                })
                geplaatst_eerste = True
                break

            if not geplaatst_eerste:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldige middenplek gevonden voor eerste van duo (alle rijen bezet)"
                })

        # Tweede van het duo
        else:
            basis_col = duo_basis_col.get(duo_nummer)
            pv_name_row = duo_basis_pv_row.get(duo_nummer)

            if basis_col is None or pv_name_row is None:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen basisplek beschikbaar van eerste duo-genoot"
                })
                continue

            pv_label = next(
                (pv["naam"] for pv, r in pv_rows_pp2 if r == pv_name_row),
                f"rij {pv_name_row}"
            )

            buur_cols = []
            if basis_col - 1 in pauze_cols_pp2:
                buur_cols.append(basis_col - 1)
            if basis_col + 1 in pauze_cols_pp2:
                buur_cols.append(basis_col + 1)

            geplaatste_tweede = False

            for buur_col in buur_cols:
                if ws_pp2.cell(pv_name_row, buur_col).value not in [None, ""]:
                    continue

                if not pp2_is_valid_short_break_for_student(naam, buur_col, ws_pp2):
                    continue

                pp2_write_name(ws_pp2, pv_name_row, buur_col, naam)

                pp2_geplaatste_pauzes.append({
                    "naam": naam,
                    "pauzevlinder": pv_label,
                    "tijd": ws_pp2.cell(1, buur_col).value,
                    "type": "tweede van duo"
                })

                geplaatste_tweede = True
                break

            if not geplaatste_tweede:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldige buurplek gevonden voor tweede van duo"
                })


#STAP 2 2222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222

# -----------------------------
# STAP 2 PP optie 2:
# lange pauzes invullen van links naar rechts,
# per halfuurblok en per pauzevlinder
# -----------------------------

lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

def pp2_heeft_al_lange_pauze(naam, ws_sheet, pv_rows, pauze_cols):
    """
    Check of naam al ergens een dubbele blok heeft in PP optie 2.
    """
    for _pv, pv_row in pv_rows:
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
    return False


def pp2_lange_werkers_lijst():
    """
    Studenten die in stap 2 recht hebben op een halfuur pauze:
    - alle minderjarigen met minstens 4 uur werk
    - alle overige studenten met meer dan 6 uur werk
    - inclusief pauzevlinders indien ze eraan voldoen
    """
    result = []
    al_toegevoegd = set()

    for s in studenten:
        naam = s["naam"]
        gewerkte_uren = student_totalen.get(naam, 0)
        is_minderjarig = "-18" in str(naam)

        if is_minderjarig and gewerkte_uren >= 4:
            if naam not in al_toegevoegd:
                result.append(naam)
                al_toegevoegd.add(naam)
        elif gewerkte_uren > 6:
            if naam not in al_toegevoegd:
                result.append(naam)
                al_toegevoegd.add(naam)

    return result



def pp2_aantal_lange_pauzes_nodig_in_stap2(naam):
    """
    Hoeveel halfuren moet deze student in stap 2 krijgen?
    - minderjarige met < 4u werk => 0
    - minderjarige met >= 4u en <= 6u werk => 1
    - minderjarige met > 6u werk => 2
    - niet-minderjarige met > 6u werk => 1
    - anders => 0
    """
    gewerkte_uren = student_totalen.get(naam, 0)
    is_minderjarig = "-18" in str(naam)

    if is_minderjarig:
        if gewerkte_uren < 4:
            return 0
        if gewerkte_uren > 6:
            return 2
        return 1

    if gewerkte_uren > 6:
        return 1

    return 0


def pp2_sort_step2_namen(namenlijst):
    """
    Sorteer voor stap 2:
    - eerst wie vroeger stopt
    - bij gelijke eindtijd: random volgorde
    """
    per_einduur = defaultdict(list)

    for naam in namenlijst:
        werk_uren = pp2_get_student_work_hours(naam)
        if werk_uren:
            einduur = max(werk_uren)
            per_einduur[einduur].append(naam)

    resultaat = []
    for einduur in sorted(per_einduur.keys()):
        groep = per_einduur[einduur][:]
        random.shuffle(groep)
        resultaat.extend(groep)

    return resultaat

def pp2_get_pv_row_for_name(naam, pv_rows):
    """
    Geef de naamrij terug van de pauzevlinder met deze naam.
    """
    for pv, pv_row in pv_rows:
        if pv["naam"] == naam:
            return pv_row
    return None


def pp2_find_first_valid_long_block_any_row(naam, ws_sheet, pv_rows, pauze_cols):
    """
    Zoek het vroegst mogelijke geldige halfuur voor deze student
    over alle pauzevlinder-rijen heen, van links naar rechts.
    Retourneert (pv_row, col1, col2) of None.
    """
    blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)

    for col1, col2 in blokken:
        for _pv, pv_row in pv_rows:
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue

            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue

            return (pv_row, col1, col2)

    return None


def pp2_find_first_valid_long_block_on_fixed_row(naam, ws_sheet, pv_row, pauze_cols):
    """
    Zoek het vroegst mogelijke geldige halfuur voor deze student
    op één vaste pauzevlinder-rij, van links naar rechts.
    Retourneert (col1, col2) of None.
    """
    blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)

    for col1, col2 in blokken:
        if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
            continue
        if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
            continue

        if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
            continue

        return (col1, col2)

    return None
    


def pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
    """
    Een lange pauze mag alleen als:
    - beide kwartieren samen exact 30 min vormen
    - student werkt in beide kwartieren
    - niet in eerste of laatste werkuur
    - student op geen van beide kwartieren al elders in het pauzerooster staat
    """
    header1 = ws_sheet.cell(1, col1).value
    header2 = ws_sheet.cell(1, col2).value

    mins1 = pp2_parse_kwartier_header(header1)
    mins2 = pp2_parse_kwartier_header(header2)

    if mins1 is None or mins2 is None:
        return False

    if mins2 - mins1 != 15:
        return False

    werk_uren = pp2_get_student_work_hours(naam)
    if not werk_uren:
        return False

    uur1 = parse_header_uur(header1)
    uur2 = parse_header_uur(header2)

    if uur1 is None or uur2 is None:
        return False

    if uur1 not in werk_uren or uur2 not in werk_uren:
        return False

    eerste_uur = werk_uren[0]
    laatste_uur = werk_uren[-1]

    if uur1 == eerste_uur or uur1 == laatste_uur:
        return False
    if uur2 == eerste_uur or uur2 == laatste_uur:
        return False

    if pp2_student_heeft_al_pauze_op_kolom(
        naam=naam,
        col=col1,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows_pp2
    ):
        return False

    if pp2_student_heeft_al_pauze_op_kolom(
        naam=naam,
        col=col2,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows_pp2
    ):
        return False

    return True
    

def pp2_write_long_break(ws_sheet, pv_row, col1, col2, naam, leave_top_blank=False):
    """
    Schrijf een lange pauze van 2 kwartieren:
    - normaal: attractie erboven
    - voor pauzevlinder op eigen rij: bovenste cel leeg laten
    - naam in beide vakjes
    - groen kleuren
    """
    for col in [col1, col2]:
        info_cel = ws_sheet.cell(pv_row - 1, col)
        info_cel.alignment = center_align
        info_cel.border = thin_border

        if leave_top_blank:
            info_cel.value = ""
        else:
            header = ws_sheet.cell(1, col).value
            uur = parse_header_uur(header)
            attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
            info_cel.value = attr if attr else ""

        naam_cel = ws_sheet.cell(pv_row, col)
        naam_cel.value = naam
        naam_cel.alignment = center_align
        naam_cel.border = thin_border
        naam_cel.fill = lichtgroen_fill


def pp2_halfuur_blokken(pauze_cols, ws_sheet):
    """
    Geeft alle mogelijke halfuurblokken terug, van links naar rechts.
    Flexibel:
    - mag starten op heel uur
    - mag ook starten op :15
    Dus bv.:
    (12u00, 12u15), (12u15, 12u30), (12u30, 12u45), ...
    zolang de cellen exact 15 minuten uit elkaar liggen.
    """
    blokken = []

    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        mins1 = pp2_parse_kwartier_header(ws_sheet.cell(1, col1).value)
        mins2 = pp2_parse_kwartier_header(ws_sheet.cell(1, col2).value)

        if mins1 is None or mins2 is None:
            continue

        if mins2 - mins1 == 15:
            blokken.append((col1, col2))

    return blokken


def pp2_place_long_break_for_pv_in_own_row(pv, pv_name_row, ws_sheet, pauze_cols, lange_pauze_ontvangers, lange_werkers_random):
    """
    Geef een langwerkende pauzevlinder verplicht een lange pauze in de eigen rij.
    We proberen de blokken strikt van links naar rechts.
    De cellen erboven blijven leeg.
    """
    naam = pv["naam"]

    if naam not in lange_werkers_random:
        return False

    if naam in lange_pauze_ontvangers:
        return False

    blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)

    for col1, col2 in blokken:
        # beide kwartieren moeten leeg zijn op eigen rij
        if ws_sheet.cell(pv_name_row, col1).value not in [None, ""]:
            continue
        if ws_sheet.cell(pv_name_row, col2).value not in [None, ""]:
            continue

        if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
            continue

        pp2_write_long_break(
            ws_sheet=ws_sheet,
            pv_row=pv_name_row,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=True
        )
        lange_pauze_ontvangers.add(naam)
        return True

    return False


# 1) Bouw de kandidatenlijsten voor stap 2
pp2_step2_basis = pp2_lange_werkers_lijst()

pp2_step2_minderjarigen = []
pp2_step2_overige_lange_werkers = []

for naam in pp2_step2_basis:
    if "-18" in str(naam):
        pp2_step2_minderjarigen.append(naam)
    else:
        pp2_step2_overige_lange_werkers.append(naam)

pp2_step2_minderjarigen = pp2_sort_step2_namen(pp2_step2_minderjarigen)
pp2_step2_overige_lange_werkers = pp2_sort_step2_namen(pp2_step2_overige_lange_werkers)

# Deze lijst houden we voor de bestaande latere logica aan
pp2_lange_werkers_random = pp2_step2_minderjarigen + pp2_step2_overige_lange_werkers

# 2) Houd bij wie al minstens één lange pauze kreeg
pp2_lange_pauze_ontvangers = set()
for naam in pp2_lange_werkers_random:
    if pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
        pp2_lange_pauze_ontvangers.add(naam)

# Voor minderjarigen willen we onthouden op welke rij hun EERSTE halfuur kwam
pp2_minderjarige_eerste_halfuur_rij = {}

# 3) Eerst: alle minderjarigen die in stap 2 recht hebben op een halfuur
#    krijgen hun EERSTE halfuur zo vroeg mogelijk
for naam in pp2_step2_minderjarigen:
    nodig = pp2_aantal_lange_pauzes_nodig_in_stap2(naam)
    if nodig <= 0:
        continue

    # Heeft al ergens een lang halfuur? Dan niet nog eens als "eerste" plaatsen
    if naam in pp2_lange_pauze_ontvangers:
        continue

    gevonden = pp2_find_first_valid_long_block_any_row(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2
    )

    if gevonden is None:
        continue

    pv_row, col1, col2 = gevonden

    eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
    leave_top_blank = eigen_pv_row == pv_row

    pp2_write_long_break(
        ws_sheet=ws_pp2,
        pv_row=pv_row,
        col1=col1,
        col2=col2,
        naam=naam,
        leave_top_blank=leave_top_blank
    )

    pp2_lange_pauze_ontvangers.add(naam)
    pp2_minderjarige_eerste_halfuur_rij[naam] = pv_row

# 4) Daarna: bestaande logica voor overige lange pauzevlinders op eigen rij
for pv, pv_name_row in pv_rows_pp2:
    pp2_place_long_break_for_pv_in_own_row(
        pv=pv,
        pv_name_row=pv_name_row,
        ws_sheet=ws_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers,
        lange_werkers_random=pp2_lange_werkers_random
    )

# 5) Daarna: algemene verdeling van andere lange werkers
pp2_blokken = pp2_halfuur_blokken(pauze_cols_pp2, ws_pp2)

for col1, col2 in pp2_blokken:
    for pv, pv_name_row in pv_rows_pp2:
        if ws_pp2.cell(pv_name_row, col1).value not in [None, ""]:
            continue
        if ws_pp2.cell(pv_name_row, col2).value not in [None, ""]:
            continue

        toegewezen_naam = None

        for kandidaat in pp2_step2_overige_lange_werkers:
            if kandidaat in pp2_lange_pauze_ontvangers:
                continue
            if pp2_is_valid_long_break_for_student(kandidaat, col1, col2, ws_pp2):
                toegewezen_naam = kandidaat
                break

        if toegewezen_naam:
            pp2_write_long_break(
                ws_sheet=ws_pp2,
                pv_row=pv_name_row,
                col1=col1,
                col2=col2,
                naam=toegewezen_naam,
                leave_top_blank=False
            )
            pp2_lange_pauze_ontvangers.add(toegewezen_naam)

# 6) Helemaal als laatste:
#    minderjarigen met > 6u krijgen nog een TWEEDE halfuur
#    op exact dezelfde rij als hun eerste halfuur
for naam in pp2_step2_minderjarigen:
    if pp2_aantal_lange_pauzes_nodig_in_stap2(naam) < 2:
        continue

    vaste_rij = pp2_minderjarige_eerste_halfuur_rij.get(naam)
    if vaste_rij is None:
        continue

    gevonden = pp2_find_first_valid_long_block_on_fixed_row(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_row=vaste_rij,
        pauze_cols=pauze_cols_pp2
    )

    if gevonden is None:
        continue

    col1, col2 = gevonden

    eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
    leave_top_blank = eigen_pv_row == vaste_rij

    pp2_write_long_break(
        ws_sheet=ws_pp2,
        pv_row=vaste_rij,
        col1=col1,
        col2=col2,
        naam=naam,
        leave_top_blank=leave_top_blank
    )



#STAP 3 333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333

# -----------------------------
# STAP 3 PP optie 2:
# open spots berekenen en verdelen
# + korte pauzes van pauzevlinders zelf invullen
# -----------------------------

lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
naam_leeg_fill_pp2 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")



def pp2_benodigde_korte_kwartieren(naam):
    """
    Nieuwe regel voor PP optie 2:
    - < 4 uur gewerkt => 0 korte kwartieren
    - >= 4 uur gewerkt => 1 kort kwartier

    Dit geldt nu ook voor minderjarigen:
    - minderjarige 4u t.e.m. 6u => 1 kort kwartier
    - minderjarige > 6u => ook 1 kort kwartier
    """
    gewerkte_uren = student_totalen.get(naam, 0)

    if gewerkte_uren < 4:
        return 0

    return 1


def pp2_count_total_assigned_quarters_for_student(naam, ws_sheet, pv_rows, pauze_cols):
    """
    Telt alle kwartiercellen in PP optie 2 waar deze naam al staat.
    Dit zijn dus ALLE reeds toegekende pauzekwartieren samen.
    """
    count = 0
    for _pv, pv_row in pv_rows:
        for col in pauze_cols:
            if ws_sheet.cell(pv_row, col).value == naam:
                count += 1
    return count


def pp2_count_al_toegekende_lange_kwartieren(naam, ws_sheet, pv_rows, pauze_cols):
    """
    Telt hoeveel reeds toegekende kwartieren deel uitmaken van een LANGE pauze
    voor deze student in PP optie 2.

    Een lange pauze herkennen we als 2 opeenvolgende kwartieren op dezelfde rij
    met exact dezelfde naam.

    Voorbeelden:
    - 1 halfuur lange pauze => 2 kwartieren
    - 2 halve uren lange pauze => 4 kwartieren
    """
    count = 0
    gebruikte_cols_per_row = set()

    for _pv, pv_row in pv_rows:
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]

            if (pv_row, col1) in gebruikte_cols_per_row or (pv_row, col2) in gebruikte_cols_per_row:
                continue

            val1 = ws_sheet.cell(pv_row, col1).value
            val2 = ws_sheet.cell(pv_row, col2).value

            if val1 == naam and val2 == naam:
                count += 2
                gebruikte_cols_per_row.add((pv_row, col1))
                gebruikte_cols_per_row.add((pv_row, col2))

    return count


def pp2_count_al_toegekende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
    """
    Telt hoeveel KORTE kwartieren deze student al heeft.

    Nieuwe logica:
    - tel eerst alle reeds ingevulde kwartieren van deze student
    - trek daar alle kwartieren af die deel uitmaken van een lange pauze
    - wat overblijft, zijn korte kwartieren

    Hierdoor werkt dit ook correct voor minderjarigen die 2 halve uren kregen.
    """
    totaal = pp2_count_total_assigned_quarters_for_student(
        naam=naam,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows,
        pauze_cols=pauze_cols
    )

    lange_kwartieren = pp2_count_al_toegekende_lange_kwartieren(
        naam=naam,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows,
        pauze_cols=pauze_cols
    )

    return max(0, totaal - lange_kwartieren)


def pp2_resterende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
    """
    Hoeveel korte kwartieren heeft deze student nog nodig?
    """
    nodig = pp2_benodigde_korte_kwartieren(naam)
    al_kort = pp2_count_al_toegekende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows,
        pauze_cols=pauze_cols,
        lange_pauze_ontvangers=lange_pauze_ontvangers
    )
    return max(0, nodig - al_kort)


def pp2_heeft_al_voldoende_korte_pauze(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
    """
    True als student al genoeg korte kwartieren heeft gekregen
    volgens de nieuwe PP2-regels.
    """
    return pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows,
        pauze_cols=pauze_cols,
        lange_pauze_ontvangers=lange_pauze_ontvangers
    ) == 0


def pp2_korte_pauze_nodig_namen():
    """
    Iedereen met minstens 4 uur werk heeft recht op 1 kort kwartier,
    BEHALVE minderjarige vroege stoppers.

    Minderjarige vroege stoppers:
    - minderjarig
    - minstens 4u gewerkt
    - laatste werkuur <= 15
    """
    namen = []

    for s in studenten:
        naam = s["naam"]
        werk_uren = pp2_get_student_work_hours(naam)

        is_minor_early_stopper = (
            pp2_is_minderjarig(naam)
            and len(werk_uren) >= 4
            and werk_uren
            and max(werk_uren) <= 15
        )

        if is_minor_early_stopper:
            continue

        if pp2_benodigde_korte_kwartieren(naam) > 0:
            namen.append(naam)

    return namen



def pp2_count_remaining_empty_quarters(ws_sheet, pv_rows, pauze_cols):
    """
    Telt alle nog lege kwartiercellen in de naamrijen van PP optie 2.
    """
    count = 0
    for _pv, pv_row in pv_rows:
        for col in pauze_cols:
            if ws_sheet.cell(pv_row, col).value in [None, ""]:
                count += 1
    return count


def pp2_get_empty_cols_for_pv_row(ws_sheet, pv_row, pauze_cols, open_spots_set):
    """
    Geeft alle lege kwartierkolommen terug voor deze pauzevlinder-rij,
    exclusief reeds gemarkeerde open spots.
    """
    cols = []
    for col in pauze_cols:
        if (pv_row, col) in open_spots_set:
            continue
        if ws_sheet.cell(pv_row, col).value in [None, ""]:
            cols.append(col)
    return cols


def pp2_mark_open_spot(ws_sheet, pv_row, col):
    """
    Open spot blijft gewoon blauw en leeg.
    """
    top_cel = ws_sheet.cell(pv_row - 1, col)
    top_cel.value = ""
    top_cel.alignment = center_align
    top_cel.border = thin_border

    cel = ws_sheet.cell(pv_row, col)
    cel.value = ""
    cel.alignment = center_align
    cel.border = thin_border
    cel.fill = naam_leeg_fill_pp2




def pp2_write_short_break_for_pv(ws_sheet, pv_row, col, naam):
    """
    Schrijf 1 kort kwartier voor een pauzevlinder zelf:
    - bovenliggende cel leeg
    - naam paars
    """
    top_cel = ws_sheet.cell(pv_row - 1, col)
    top_cel.value = ""
    top_cel.alignment = center_align
    top_cel.border = thin_border

    cel = ws_sheet.cell(pv_row, col)
    cel.value = naam
    cel.alignment = center_align
    cel.border = thin_border
    cel.fill = lichtpaars_fill


def pp2_find_short_break_cols_for_pv(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, needed_quarters):
    """
    Zoek geldige kolom/kolommen voor de korte pauze van een pauzevlinder in de eigen rij.

    - needed_quarters == 1:
      neem het eerstvolgende geldige vrije kwartier

    - needed_quarters == 2:
      neem de eerste geldige set van 2 opeenvolgende kwartieren
    """
    if needed_quarters <= 0:
        return []

    if needed_quarters == 1:
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
            return [col]
        return []

    if needed_quarters == 2:
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]

            if col2 != col1 + 1:
                continue

            if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                continue

            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue

            if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                continue
            if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                continue

            return [col1, col2]

        return []

    return []



# 1) Is dit een korte dag?
#    Korte dag = niemand heeft in stap 2 een lange pauze nodig
pp2_is_korte_dag = (len(pp2_lange_werkers_lijst()) == 0)

pp2_open_spots = set()
pp2_pv_short_breaks_placed = []

# -----------------------------
# Hulploop: korte pauzes van pauzevlinders zelf invullen
#
# Nieuwe regel:
# - pauzevlinders die GEEN lange werker zijn: hier al plaatsen
# - pauzevlinders die WEL lange werker zijn: nog NIET hier plaatsen
#   -> die komen later in stap 4, na de korte werkers
# -----------------------------
for pv, pv_row in pv_rows_pp2:
    naam = pv["naam"]

    # Lange pauzevlinders hier nog overslaan:
    # hun korte pauze moet pas later komen
    if naam in pp2_lange_werkers_lijst():
        continue

    resterend_nodig = pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

    if resterend_nodig <= 0:
        continue

    gekozen_cols = pp2_find_short_break_cols_for_pv(
        naam=naam,
        pv_row=pv_row,
        ws_sheet=ws_pp2,
        pauze_cols=pauze_cols_pp2,
        open_spots_set=pp2_open_spots,
        needed_quarters=resterend_nodig
    )

    if not gekozen_cols:
        continue

    for col in gekozen_cols:
        pp2_write_short_break_for_pv(ws_pp2, pv_row, col, naam)

    pp2_pv_short_breaks_placed.append({
        "naam": naam,
        "kolommen": gekozen_cols,
        "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
    })

# -----------------------------
# 2) Tellen hoeveel kwartierblokjes nog leeg zijn
# -----------------------------
pp2_remaining_empty_quarters = pp2_count_remaining_empty_quarters(
    ws_sheet=ws_pp2,
    pv_rows=pv_rows_pp2,
    pauze_cols=pauze_cols_pp2
)

# -----------------------------
# 3) Tellen hoeveel KORTE kwartieren nog gegeven moeten worden
#    Nieuwe telling:
#    - gewone student meestal 1
#    - minderjarige >4u = 2
#    - ook minderjarige >6u met al lange pauze telt hier nog voor 2
# -----------------------------
pp2_korte_pauze_gerechtigden = pp2_korte_pauze_nodig_namen()

pp2_remaining_short_quarters_needed = 0
for naam in pp2_korte_pauze_gerechtigden:
    pp2_remaining_short_quarters_needed += pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

# -----------------------------
# 4) Open spots berekenen
# -----------------------------
pp2_open_spots_count = pp2_remaining_empty_quarters - pp2_remaining_short_quarters_needed
if pp2_open_spots_count < 0:
    pp2_open_spots_count = 0

# -----------------------------
# 5) Open spots verdelen
#
# KORTE DAG:
# - eerst korte pauzes geplaatst
# - dus open spots vallen automatisch NA de korte pauzes
#
# LANGE DAG:
# - script blijft exact hetzelfde gedrag houden als nu
#   => open spots verdelen zoals nu
# -----------------------------

if not pp2_is_korte_dag:
    # ---------------------------------------------------
    # LANGE DAG:
    # - Open spots eerst verdelen (zoals in het originele script)
    # - Daarna enkel de korte pauzes van KORTE pauzevlinders plaatsen
    # - Lange pauzevlinders komen pas in stap 4 aan bod
    # ---------------------------------------------------

    # Reset eerst eventuele eerder geplaatste korte pauzes van pauzevlinders
    for item in pp2_pv_short_breaks_placed:
        naam = item["naam"]

        pv_row = next(
            pv_row for pv, pv_row in pv_rows_pp2
            if pv["naam"] == naam
        )

        for col in item["kolommen"]:
            top_cel = ws_pp2.cell(pv_row - 1, col)
            top_cel.value = ""
            top_cel.alignment = center_align
            top_cel.border = thin_border

            cel = ws_pp2.cell(pv_row, col)
            cel.value = ""
            cel.alignment = center_align
            cel.border = thin_border
            cel.fill = naam_leeg_fill_pp2

    pp2_pv_short_breaks_placed = []

    # ---------------------------------------------------
    # 1) Open spots verdelen
    # ---------------------------------------------------
    ronde_nummer = 0

    while len(pp2_open_spots) < pp2_open_spots_count:
        iets_geplaatst_deze_ronde = False
        vooraan = (ronde_nummer % 2 == 0)

        for _pv, pv_row in pv_rows_pp2:
            if len(pp2_open_spots) >= pp2_open_spots_count:
                break

            lege_cols = pp2_get_empty_cols_for_pv_row(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots
            )

            if not lege_cols:
                continue

            gekozen_col = lege_cols[0] if vooraan else lege_cols[-1]

            pp2_open_spots.add((pv_row, gekozen_col))
            pp2_mark_open_spot(ws_pp2, pv_row, gekozen_col)
            iets_geplaatst_deze_ronde = True

        if not iets_geplaatst_deze_ronde:
            break

        ronde_nummer += 1

    # ---------------------------------------------------
    # 2) Enkel korte pauzes van KORTE pauzevlinders plaatsen
    # Lange pauzevlinders worden hier overgeslagen
    # ---------------------------------------------------
    for pv, pv_row in pv_rows_pp2:
        naam = pv["naam"]

        # Lange pauzevlinders hier overslaan
        if naam in pp2_lange_werkers_lijst():
            continue

        resterend_nodig = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )

        if resterend_nodig <= 0:
            continue

        gekozen_cols = pp2_find_short_break_cols_for_pv(
            naam=naam,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots,
            needed_quarters=resterend_nodig
        )

        if not gekozen_cols:
            continue

        for col in gekozen_cols:
            pp2_write_short_break_for_pv(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                col=col,
                naam=naam
            )

        pp2_pv_short_breaks_placed.append({
            "naam": naam,
            "kolommen": gekozen_cols,
            "tijden": [
                ws_pp2.cell(1, col).value for col in gekozen_cols
            ]
        })


#STAP 4 44444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444


# -----------------------------
# STAP 4 PP optie 2:
# korte pauzes voor:
# 1) studenten die vroeger stoppen dan het einduur
# 2) daarna lange pauzevlinders zelf (in eigen rij)
# met nieuwe minderjarigenlogica
# -----------------------------

lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")


def pp2_get_day_end_hour():
    """
    Einduur van de dag op basis van open_uren.
    """
    if not open_uren:
        return None
    return max(open_uren)


def pp2_get_students_stopping_before_end():
    """
    Studenten die vroeger stoppen dan het einduur van de dag
    en minstens 4 uur werken.
    """
    einduur_dag = pp2_get_day_end_hour()
    result = []

    if einduur_dag is None:
        return result

    for s in studenten:
        naam = s["naam"]
        werk_uren = pp2_get_student_work_hours(naam)

        if len(werk_uren) < 4:
            continue

        if max(werk_uren) < einduur_dag:
            result.append(naam)

    return result


def pp2_write_short_break_regular(ws_sheet, pv_row, col, naam):
    """
    Korte pauze voor gewone student:
    - bovenliggende cel = attractie
    - naamcel:
        * lichtgeel voor minderjarigen die >4u werken
        * lichtpaars voor alle andere korte pauzes
    """
    header = ws_sheet.cell(1, col).value
    uur = parse_header_uur(header)

    attr = vind_attractie_op_uur(naam, uur) if uur is not None else None

    top_cel = ws_sheet.cell(pv_row - 1, col)
    top_cel.value = attr if attr else ""
    top_cel.alignment = center_align
    top_cel.border = thin_border

    cel = ws_sheet.cell(pv_row, col)
    cel.value = naam
    cel.alignment = center_align
    cel.border = thin_border

    if pp2_is_minderjarig(naam) and student_totalen.get(naam, 0) > 4:
        cel.fill = roze_fill
    else:
        cel.fill = lichtpaars_fill


def pp2_get_long_break_owners_on_row(ws_sheet, pv_row, pauze_cols):
    """
    Geeft alle studenten terug die op deze rij een lange pauze hebben,
    gesorteerd op het ankerpunt voor hun korte pauze:
    - minderjarige lange werkers (>6u): gesorteerd op hun LAATSTE halfuur
    - alle anderen: gesorteerd op hun EERSTE halfuur (= volgorde van links naar rechts)
    """
    # Verzamel per student de eerste én laatste halfuur-eindkolom op deze rij
    eerste_col = {}
    laatste_col = {}

    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        val1 = ws_sheet.cell(pv_row, col1).value
        val2 = ws_sheet.cell(pv_row, col2).value

        if val1 and val1 == val2:
            naam = str(val1).strip()
            if naam not in eerste_col:
                eerste_col[naam] = col2
            laatste_col[naam] = col2

    owners = list(eerste_col.keys())

    def sorteersleutel(naam):
        is_minor_long_worker = (
            pp2_is_minderjarig(naam)
            and student_totalen.get(naam, 0) > 6
        )
        if is_minor_long_worker:
            return laatste_col.get(naam, 0)
        else:
            return eerste_col.get(naam, 0)

    owners.sort(key=sorteersleutel)
    return owners

def pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
    """
    Check of deze student een lange pauze heeft op precies deze rij.
    """
    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        if (
            ws_sheet.cell(pv_row, col1).value == naam and
            ws_sheet.cell(pv_row, col2).value == naam
        ):
            return True

    return False


def pp2_student_is_long_worker(naam):
    return naam in pp2_lange_werkers_lijst()


def pp2_find_two_consecutive_valid_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
    """
    Zoek 2 opeenvolgende geldige kwartieren voor deze student op deze specifieke rij.
    """
    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        if col2 != col1 + 1:
            continue

        if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
            continue

        if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
            continue
        if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
            continue

        if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
            continue
        if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
            continue

        return [col1, col2]

    return []


def pp2_find_one_valid_col_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
    """
    Zoek 1 geldig kwartier voor deze student op deze specifieke rij.
    """
    for col in pauze_cols:
        if (pv_row, col) in open_spots_set:
            continue

        if ws_sheet.cell(pv_row, col).value not in [None, ""]:
            continue

        if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
            continue

        return [col]

    return []


def pp2_find_needed_short_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, min_col_exclusive=None, zoek_zo_laat_mogelijk=False):
    """
    Zoek het korte kwartier dat deze student nog nodig heeft op deze specifieke rij.

    - min_col_exclusive: zoek pas NA deze kolom
    - zoek_zo_laat_mogelijk: zoek van rechts naar links (voor minderjarige lange werkers)
    """
    resterend = pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_sheet,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

    if resterend <= 0:
        return []

    kandidaat_cols = list(reversed(pauze_cols)) if zoek_zo_laat_mogelijk else list(pauze_cols)

    for col in kandidaat_cols:
        if min_col_exclusive is not None and col <= min_col_exclusive:
            continue

        if (pv_row, col) in open_spots_set:
            continue

        if ws_sheet.cell(pv_row, col).value not in [None, ""]:
            continue

        if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
            continue

        return [col]

    return []



def pp2_place_short_break_cols_on_row(naam, pv, pv_row, cols):
    """
    Schrijf 1 of 2 korte kwartieren voor gewone student op een bepaalde rij.
    """
    for col in cols:
        pp2_write_short_break_regular(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            col=col,
            naam=naam
        )

    pp2_regular_short_breaks_placed.append({
        "naam": naam,
        "pauzevlinder": pv["naam"],
        "tijden": [ws_pp2.cell(1, col).value for col in cols],
        "zelfde_rij_als_lange_pauze": pp2_student_has_long_break_in_row(
            naam, ws_pp2, pv_row, pauze_cols_pp2
        )
    })





def pp2_student_heeft_nog_lange_pauze_nodig(naam, ws_sheet, pv_rows, pauze_cols):
    """
    Bepaal of deze student volgens de nieuwe regels nog minstens 1 lang halfuur mist.

    Regels minderjarigen:
    - < 4u gewerkt => 0 lange pauzes
    - 4u t.e.m. 6u => 1 lange pauze
    - > 6u => 2 lange pauzes

    Regels meerderjarigen:
    - > 6u => 1 lange pauze
    - anders => 0
    """
    gewerkte_uren = student_totalen.get(naam, 0)
    is_minor = pp2_is_minderjarig(naam)

    if is_minor:
        if gewerkte_uren < 4:
            nodig = 0
        elif gewerkte_uren <= 6:
            nodig = 1
        else:
            nodig = 2
    else:
        nodig = 1 if gewerkte_uren > 6 else 0

    # tel hoeveel lange halve uren al effectief ingepland zijn
    al = 0
    for _pv, pv_row in pv_rows:
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]

            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                al += 1

    return al < nodig


def pp2_find_first_valid_long_block_in_step4(naam, ws_sheet, pv_rows, pauze_cols, open_spots_set):
    """
    Zoek in stap 4 een geldig halfuur voor een student.

    Voor minderjarige vroege stoppers:
    - kies het EERSTE geldige halfuur (dus zo vroeg mogelijk)

    Voor alle anderen:
    - behoud ook het eerste geldige halfuur
    """
    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        for _pv, pv_row in pv_rows:
            if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                continue

            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue

            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue

            return pv_row, col1, col2

    return None

# ---------------------------------------
# 0) Eerst: minderjarigen die nog een LANGE pauze missen alsnog proberen plaatsen
#    Dit vangt het geval op waarin een minderjarige laat start
#    en stap 2 geen geldig halfuur vond.
# ---------------------------------------
pp2_step4_late_long_break_rescue = []
pp2_regular_short_breaks_placed = []

# ---------------------------------------
# 0A) Eerst: minderjarige vroege stoppers
#     die nog een LANGE pauze missen
#     => zo vroeg mogelijk (links naar rechts)
# ---------------------------------------
pp2_minor_early_stoppers = [
    s["naam"] for s in studenten
    if (
        pp2_is_minderjarig(s["naam"])
        and len(pp2_get_student_work_hours(s["naam"])) >= 4
        and pp2_get_student_work_hours(s["naam"])
        and max(pp2_get_student_work_hours(s["naam"])) <= 15
    )
]

for naam in pp2_minor_early_stoppers:
    if not pp2_student_heeft_nog_lange_pauze_nodig(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2
    ):
        continue

    gevonden = None

    for idx in range(len(pauze_cols_pp2) - 1):
        col1 = pauze_cols_pp2[idx]
        col2 = pauze_cols_pp2[idx + 1]

        for _pv, pv_row in pv_rows_pp2:
            if (pv_row, col1) in pp2_open_spots or (pv_row, col2) in pp2_open_spots:
                continue

            if ws_pp2.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_pp2.cell(pv_row, col2).value not in [None, ""]:
                continue

            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_pp2):
                continue

            gevonden = (pv_row, col1, col2)
            break

        if gevonden is not None:
            break

    if gevonden is None:
        continue

    pv_row, col1, col2 = gevonden

    pp2_write_long_break(
        ws_sheet=ws_pp2,
        pv_row=pv_row,
        col1=col1,
        col2=col2,
        naam=naam,
        leave_top_blank=False
    )

    pp2_lange_pauze_ontvangers.add(naam)

    pp2_step4_late_long_break_rescue.append({
        "naam": naam,
        "tijden": [ws_pp2.cell(1, col1).value, ws_pp2.cell(1, col2).value]
    })


# ---------------------------------------
# 0B) Daarna: exact diezelfde minderjarige
#     vroege stoppers hun KORTE pauze
#     => zo laat mogelijk (rechts naar links)
#     => bij voorkeur op dezelfde rij als het laatste halfuur
# ---------------------------------------
for naam in pp2_minor_early_stoppers:
    resterend = pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

    if resterend <= 0:
        continue

    # Zoek de rij én eindkolom van het LAATSTE halfuur van deze student
    laatste_lange_eindcol = None
    laatste_lange_rij = None

    for _pv, pv_row in pv_rows_pp2:
        for idx in range(len(pauze_cols_pp2) - 1):
            col1 = pauze_cols_pp2[idx]
            col2 = pauze_cols_pp2[idx + 1]

            if (
                ws_pp2.cell(pv_row, col1).value == naam
                and ws_pp2.cell(pv_row, col2).value == naam
            ):
                if laatste_lange_eindcol is None or col2 > laatste_lange_eindcol:
                    laatste_lange_eindcol = col2
                    laatste_lange_rij = pv_row

    if laatste_lange_eindcol is None:
        continue

    gekozen = None

    # Bouw PV-rij volgorde: eerst de rij van het laatste halfuur, dan de rest
    pv_volgorde = (
        [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row == laatste_lange_rij]
        + [(pv, pv_row) for pv, pv_row in pv_rows_pp2 if pv_row != laatste_lange_rij]
    )

    for col in reversed(pauze_cols_pp2):
        if col <= laatste_lange_eindcol:
            continue

        for pv, pv_row in pv_volgorde:
            if (pv_row, col) in pp2_open_spots:
                continue

            if ws_pp2.cell(pv_row, col).value not in [None, ""]:
                continue

            if not pp2_is_valid_short_break_for_student(naam, col, ws_pp2):
                continue

            gekozen = (pv, pv_row, col)
            break

        if gekozen is not None:
            break

    if gekozen is None:
        continue

    pv, pv_row, col = gekozen

    pp2_place_short_break_cols_on_row(
        naam=naam,
        pv=pv,
        pv_row=pv_row,
        cols=[col]
    )


# ---------------------------------------
# 1) Daarna pas: gewone korte werkers
#    die vroeger stoppen dan het einde
#    van de dag, maar GEEN minderjarige
#    vroege stoppers zijn
# ---------------------------------------
pp2_students_before_end_all = pp2_get_students_stopping_before_end()

pp2_students_before_end_pending = [
    naam for naam in pp2_students_before_end_all
    if naam not in pp2_minor_early_stoppers
    and pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    ) > 0
]

def pp2_get_last_long_break_end_col_for_sort(naam):
    """
    Geeft de eindkolom van het LAATSTE halfuur van deze student terug,
    over alle PV-rijen heen. Studenten zonder lange pauze krijgen -1,
    zodat ze vooraan komen in de sortering.
    """
    eindcol = -1
    for _pv, pv_row in pv_rows_pp2:
        for idx in range(len(pauze_cols_pp2) - 1):
            col1 = pauze_cols_pp2[idx]
            col2 = pauze_cols_pp2[idx + 1]
            if (
                ws_pp2.cell(pv_row, col1).value == naam
                and ws_pp2.cell(pv_row, col2).value == naam
            ):
                if col2 > eindcol:
                    eindcol = col2
    return eindcol

pp2_students_before_end_pending.sort(
    key=lambda naam: pp2_get_last_long_break_end_col_for_sort(naam)
)

pp2_regular_short_breaks_placed = []



for col in pauze_cols_pp2:
    if not pp2_students_before_end_pending:
        break

    for pv, pv_row in pv_rows_pp2:
        if not pp2_students_before_end_pending:
            break

        if (pv_row, col) in pp2_open_spots:
            continue

        if ws_pp2.cell(pv_row, col).value not in [None, ""]:
            continue

        toegewezen_naam = None
        toegewezen_cols = []

        # ---------------------------------------------------
        # PRIORITEIT 1:
        # studenten die op deze rij al een lange pauze hebben
        # Voor minderjarige lange werkers (>6u): zoek na het
        # LAATSTE halfuur op DEZE rij (niet over alle rijen).
        # ---------------------------------------------------
        rij_lange_pauze_namen = pp2_get_long_break_owners_on_row(
            ws_pp2,
            pv_row,
            pauze_cols_pp2
        )

        for kandidaat in rij_lange_pauze_namen:
            if kandidaat not in pp2_students_before_end_pending:
                continue

            if not pp2_student_has_long_break_in_row(
                kandidaat,
                ws_pp2,
                pv_row,
                pauze_cols_pp2
            ):
                continue

            is_minor_long_worker = (
                pp2_is_minderjarig(kandidaat)
                and student_totalen.get(kandidaat, 0) > 6
            )

            if is_minor_long_worker:
                # Zoek eindkolom van het LAATSTE halfuur op DEZE specifieke rij
                ankercol = None
                for idx in range(len(pauze_cols_pp2) - 1):
                    col1 = pauze_cols_pp2[idx]
                    col2 = pauze_cols_pp2[idx + 1]
                    if (
                        ws_pp2.cell(pv_row, col1).value == kandidaat
                        and ws_pp2.cell(pv_row, col2).value == kandidaat
                    ):
                        ankercol = col2
            else:
                ankercol = None

            cols = pp2_find_needed_short_cols_for_student_on_row(
                naam=kandidaat,
                pv_row=pv_row,
                ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots,
                min_col_exclusive=ankercol,
                zoek_zo_laat_mogelijk=is_minor_long_worker
            )

            if not cols:
                continue

            toegewezen_naam = kandidaat
            toegewezen_cols = cols
            break

        # ---------------------------------------------------
        # PRIORITEIT 2:
        # anders eerste geldige kandidaat uit de vaste lijst
        # ---------------------------------------------------
        if toegewezen_naam is None:
            for kandidaat in pp2_students_before_end_pending:
                cols = pp2_find_needed_short_cols_for_student_on_row(
                    naam=kandidaat,
                    pv_row=pv_row,
                    ws_sheet=ws_pp2,
                    pauze_cols=pauze_cols_pp2,
                    open_spots_set=pp2_open_spots
                )

                if not cols:
                    continue

                toegewezen_naam = kandidaat
                toegewezen_cols = cols
                break

        if toegewezen_naam and toegewezen_cols:
            pp2_place_short_break_cols_on_row(
                naam=toegewezen_naam,
                pv=pv,
                pv_row=pv_row,
                cols=toegewezen_cols
            )

            if toegewezen_naam in pp2_students_before_end_pending:
                pp2_students_before_end_pending.remove(toegewezen_naam)

# ---------------------------------------
# 2) Daarna: lange pauzevlinders zelf
#    - alleen die nog korte kwartieren nodig hebben
#    - alleen in eigen rij
#    - na korte werkers
#    - voor andere lange werkers
# ---------------------------------------
pp2_lange_pv_short_breaks_placed = []

for pv, pv_row in pv_rows_pp2:
    naam = pv["naam"]

    if not pp2_student_is_long_worker(naam):
        continue

    resterend = pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

    if resterend <= 0:
        continue

    gekozen_cols = pp2_find_needed_short_cols_for_student_on_row(
        naam=naam,
        pv_row=pv_row,
        ws_sheet=ws_pp2,
        pauze_cols=pauze_cols_pp2,
        open_spots_set=pp2_open_spots
    )

    if not gekozen_cols:
        continue

    for col in gekozen_cols:
        pp2_write_short_break_for_pv(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            col=col,
            naam=naam
        )

    pp2_lange_pv_short_breaks_placed.append({
        "naam": naam,
        "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
    })


# STAP 5 55555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555


# -----------------------------------
# STAP 5 PP optie 2:
# laatste resterende korte kwartieren invullen
# - werkt met resterende kwartieren i.p.v. ja/nee
# - minderjarigen >4u krijgen hier ook 2 opeenvolgende kwartieren
# - eerst overige pending korte kwartieren
# - pas daarna eindwerkers zonder lange pauze
# -----------------------------------

def pp2_get_long_break_students_on_row_in_order(ws_sheet, pv_row, pauze_cols):
    """
    Geef de studenten terug die op deze rij een lange pauze hebben,
    in dezelfde volgorde als de lange pauzes op de rij zelf:
    dus van links naar rechts.

    Dit zorgt ervoor dat korte pauzes later ook in een logische
    volgorde kunnen volgen, gelijklopend met de lange pauzes.
    """
    found = {}

    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        val1 = ws_sheet.cell(pv_row, col1).value
        val2 = ws_sheet.cell(pv_row, col2).value

        if val1 and val1 == val2:
            naam = str(val1).strip()
            # bewaar de startkolom van de eerste lange pauze op deze rij
            if naam not in found:
                found[naam] = col1

    return [naam for naam, _col in sorted(found.items(), key=lambda x: x[1])]


def pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
    """
    True als deze student op deze specifieke rij ergens een lange pauze heeft.
    """
    for idx in range(len(pauze_cols) - 1):
        col1 = pauze_cols[idx]
        col2 = pauze_cols[idx + 1]

        if (
            ws_sheet.cell(pv_row, col1).value == naam and
            ws_sheet.cell(pv_row, col2).value == naam
        ):
            return True

    return False


def pp2_student_works_until_day_end(naam):
    """
    True als student werkt tot het einduur van de dag.
    """
    werk_uren = pp2_get_student_work_hours(naam)
    if not werk_uren or not open_uren:
        return False
    return max(werk_uren) == max(open_uren)


def pp2_build_step5_pending_groups():
    """
    Splits alle NIET-pauzevlinders die nog korte kwartieren nodig hebben in:
    A) overige pending korte kwartieren
    B) eindwerkers zonder lange pauze

    Minderjarige vroege stoppers horen hier ook NIET meer in:
    die werden al eerder apart behandeld en mogen in stap 5
    niet opnieuw een kort kwartier krijgen.

    Pauzevlinders zelf horen hier ook niet meer in:
    - korte pauzevlinders werden al eerder verwerkt
    - lange pauzevlinders kregen in stap 4 hun eigen aparte fase,
      enkel in hun eigen rij
    """
    pauzevlinder_namen_set = {pv["naam"] for pv in selected}
    all_pending = []

    for s in studenten:
        naam = s["naam"]

        # Pauzevlinders hier NIET meer meenemen
        if naam in pauzevlinder_namen_set:
            continue

        # Minderjarige vroege stoppers hier ook NIET meer meenemen
        if (
            pp2_is_minderjarig(naam)
            and len(pp2_get_student_work_hours(naam)) >= 4
            and pp2_get_student_work_hours(naam)
            and max(pp2_get_student_work_hours(naam)) <= 15
        ):
            continue

        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )

        if resterend > 0:
            all_pending.append(naam)

    endworkers_without_long_break = []
    other_pending_short_breaks = []

    for naam in all_pending:
        heeft_lange = (naam in pp2_lange_pauze_ontvangers)
        werkt_tot_einduur = pp2_student_works_until_day_end(naam)

        if werkt_tot_einduur and not heeft_lange:
            endworkers_without_long_break.append(naam)
        else:
            other_pending_short_breaks.append(naam)

    random.shuffle(other_pending_short_breaks)
    random.shuffle(endworkers_without_long_break)

    return other_pending_short_breaks, endworkers_without_long_break


def pp2_try_assign_from_candidate_list_on_row(candidate_list, pv, pv_row, shuffle_candidates=False):
    """
    Probeer op deze rij een kandidaat te plaatsen uit de opgegeven lijst.
    Werkt met 1 of 2 kwartieren, afhankelijk van wat nog nodig is.

    Belangrijk:
    - de volgorde van candidate_list blijft behouden als dat een prioriteitslijst is
      (bv. dezelfde volgorde als de lange pauzes op de rij)
    - voor gewone fallback-lijsten kan shuffle_candidates=True gebruikt worden
      zodat niet-minderjarigen daar opnieuw randomer verdeeld worden
    """
    kandidaten = candidate_list[:]

    if shuffle_candidates and len(kandidaten) > 1:
        random.shuffle(kandidaten)

    for kandidaat in kandidaten:
        cols = pp2_find_needed_short_cols_for_student_on_row(
            naam=kandidaat,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots
        )

        if not cols:
            continue

        pp2_place_short_break_cols_on_row(
            naam=kandidaat,
            pv=pv,
            pv_row=pv_row,
            cols=cols
        )

        return kandidaat, cols

    return None, []


pp2_other_pending_short_breaks, pp2_endworkers_without_long_break = pp2_build_step5_pending_groups()

pp2_step5_short_breaks_placed = []

# -----------------------------------
# 2) Eerst alle "gewone" resterende korte kwartieren invullen
#    inclusief:
#    - studenten met lange pauze die nog korte kwartieren missen
#    - minderjarigen met dubbele korte kwartieren
# -----------------------------------
for col in pauze_cols_pp2:
    if not pp2_other_pending_short_breaks:
        break

    for pv, pv_row in pv_rows_pp2:
        if not pp2_other_pending_short_breaks:
            break

        # open spots overslaan
        if (pv_row, col) in pp2_open_spots:
            continue

        # vak moet leeg zijn
        if ws_pp2.cell(pv_row, col).value not in [None, ""]:
            continue

        toegewezen_naam = None
        toegewezen_cols = []

        # -----------------------------------
        # PRIORITEIT 1:
        # studenten die in deze rij al een lange pauze kregen,
        # in dezelfde volgorde als hun lange pauzes op die rij
        # -----------------------------------
        rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            pauze_cols=pauze_cols_pp2
        )

        prioriteitslijst = [
            naam for naam in rij_lange_pauze_namen
            if naam in pp2_other_pending_short_breaks
        ]

        toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
            candidate_list=prioriteitslijst,
            pv=pv,
            pv_row=pv_row,
            shuffle_candidates=False
        )

        # -----------------------------------
        # PRIORITEIT 2:
        # fallback naar overige nog open korte kwartieren
        # hier mag het randomer blijven voor niet-minderjarigen
        # -----------------------------------
        if toegewezen_naam is None:
            fallback_lijst = [
                naam for naam in pp2_other_pending_short_breaks
                if naam not in prioriteitslijst
            ]

            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=fallback_lijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=True
            )

        # schrijven/loggen indien kandidaat gevonden
        if toegewezen_naam:
            pp2_step5_short_breaks_placed.append({
                "naam": toegewezen_naam,
                "pauzevlinder": pv["naam"],
                "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
            })

            if pp2_resterende_korte_kwartieren(
                naam=toegewezen_naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            ) <= 0:
                if toegewezen_naam in pp2_other_pending_short_breaks:
                    pp2_other_pending_short_breaks.remove(toegewezen_naam)


# -----------------------------------
# 3) Pas daarna:
#    studenten die tot het einduur werken én geen lange pauze kregen
# -----------------------------------
for col in pauze_cols_pp2:
    if not pp2_endworkers_without_long_break:
        break

    for pv, pv_row in pv_rows_pp2:
        if not pp2_endworkers_without_long_break:
            break

        # open spots overslaan
        if (pv_row, col) in pp2_open_spots:
            continue

        # vak moet leeg zijn
        if ws_pp2.cell(pv_row, col).value not in [None, ""]:
            continue

        toegewezen_naam = None
        toegewezen_cols = []

        rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            pauze_cols=pauze_cols_pp2
        )

        prioriteitslijst = [
            naam for naam in rij_lange_pauze_namen
            if naam in pp2_endworkers_without_long_break
        ]

        toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
            candidate_list=prioriteitslijst,
            pv=pv,
            pv_row=pv_row,
            shuffle_candidates=False
        )

        if toegewezen_naam is None:
            fallback_lijst = [
                naam for naam in pp2_endworkers_without_long_break
                if naam not in prioriteitslijst
            ]

            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=fallback_lijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=True
            )

        if toegewezen_naam:
            pp2_step5_short_breaks_placed.append({
                "naam": toegewezen_naam,
                "pauzevlinder": pv["naam"],
                "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
            })

            if pp2_resterende_korte_kwartieren(
                naam=toegewezen_naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            ) <= 0:
                if toegewezen_naam in pp2_endworkers_without_long_break:
                    pp2_endworkers_without_long_break.remove(toegewezen_naam)

# -----------------------------------
# 3) Pas daarna:
#    studenten die tot het einduur werken én geen lange pauze kregen
# -----------------------------------
for col in pauze_cols_pp2:
    if not pp2_endworkers_without_long_break:
        break

    for pv, pv_row in pv_rows_pp2:
        if not pp2_endworkers_without_long_break:
            break

        # open spots overslaan
        if (pv_row, col) in pp2_open_spots:
            continue

        # vak moet leeg zijn
        if ws_pp2.cell(pv_row, col).value not in [None, ""]:
            continue

        toegewezen_naam = None
        toegewezen_cols = []

        rij_lange_pauze_namen = pp2_get_long_break_students_on_row_in_order(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            pauze_cols=pauze_cols_pp2
        )

        prioriteitslijst = [
            naam for naam in rij_lange_pauze_namen
            if naam in pp2_endworkers_without_long_break
        ]

        toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
            candidate_list=prioriteitslijst,
            pv=pv,
            pv_row=pv_row,
            shuffle_candidates=False
        )

        if toegewezen_naam is None:
            fallback_lijst = [
                naam for naam in pp2_endworkers_without_long_break
                if naam not in prioriteitslijst
            ]

            toegewezen_naam, toegewezen_cols = pp2_try_assign_from_candidate_list_on_row(
                candidate_list=fallback_lijst,
                pv=pv,
                pv_row=pv_row,
                shuffle_candidates=True
            )

        if toegewezen_naam:
            pp2_step5_short_breaks_placed.append({
                "naam": toegewezen_naam,
                "pauzevlinder": pv["naam"],
                "tijden": [ws_pp2.cell(1, c).value for c in toegewezen_cols],
                "via_lange_pauze_prioriteit": toegewezen_naam in rij_lange_pauze_namen
            })

            if pp2_resterende_korte_kwartieren(
                naam=toegewezen_naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            ) <= 0:
                if toegewezen_naam in pp2_endworkers_without_long_break:
                    pp2_endworkers_without_long_break.remove(toegewezen_naam)

# -----------------------------
# BN15 vinkje: tekst uit BO15:BO30 tonen op PP optie 2
# Drie rijen onder de laatste rij van de pauzeplanning
# -----------------------------
bn15_vinkje_pp2 = ws.cell(15, 66).value  # BN15 in Input

if bn15_vinkje_pp2 in [1, True, "WAAR", "X"]:
    start_rij_pp2 = ws_pp2.max_row + 3

    for i, input_rij in enumerate(range(15, 31)):
        waarde = ws.cell(input_rij, 67).value  # BO kolom = 67

        if waarde:
            cel = ws_pp2.cell(row=start_rij_pp2 + i, column=1, value=waarde)
            cel.alignment = Alignment(horizontal="left", vertical="center")


#FEEDBACKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK
# =============================
# FEEDBACK SHEET - OPTIE 2
# =============================
ws_feedback2 = wb_out.create_sheet("Feedback optie 2")

groen_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
rood_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

row_fb2 = 1

ws_feedback2.cell(row_fb2, 1, "Feedback PP optie 2").font = Font(bold=True)
row_fb2 += 2

# -----------------------------------
# 1) Lange pauzes controleren
# Nieuwe logica PP optie 2:
# - alleen studenten met >6 uur werk moeten een lange pauze hebben
# -----------------------------------
pp2_lange_pauze_ontbreekt = []

for s in studenten:
    naam = s["naam"]
    gewerkte_uren = student_totalen.get(naam, 0)

    if gewerkte_uren > 6:
        if not pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
            pp2_lange_pauze_ontbreekt.append(naam)

if not pp2_lange_pauze_ontbreekt:
    cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle lange pauzes toegekend")
    cel.fill = groen_fill
    cel.font = Font(bold=True, color="006100")
    row_fb2 += 2
else:
    cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende lange pauzes:")
    cel.fill = rood_fill
    cel.font = Font(bold=True)
    row_fb2 += 1

    for naam in sorted(pp2_lange_pauze_ontbreekt):
        ws_feedback2.cell(row_fb2, 1, naam)
        row_fb2 += 1

    row_fb2 += 1

# -----------------------------------
# 2) Korte kwartieren controleren
# Gebruik exact dezelfde logica als de planner zelf:
# - pp2_benodigde_korte_kwartieren(...)
# - pp2_resterende_korte_kwartieren(...)
# Dus geen aparte feedbacktelling meer
# -----------------------------------
pp2_korte_kwartieren_ontbreekt = []

for s in studenten:
    naam = s["naam"]

    nodig = pp2_benodigde_korte_kwartieren(naam)
    if nodig <= 0:
        continue

    resterend = pp2_resterende_korte_kwartieren(
        naam=naam,
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2,
        lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
    )

    if resterend > 0:
        pp2_korte_kwartieren_ontbreekt.append((naam, resterend))

if not pp2_korte_kwartieren_ontbreekt:
    cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle korte kwartieren toegekend")
    cel.fill = groen_fill
    cel.font = Font(bold=True, color="006100")
    row_fb2 += 2
else:
    cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende korte kwartieren:")
    cel.fill = rood_fill
    cel.font = Font(bold=True)
    row_fb2 += 1

    for naam, resterend in sorted(pp2_korte_kwartieren_ontbreekt, key=lambda x: x[0].lower()):
        if resterend == 1:
            ws_feedback2.cell(row_fb2, 1, f"{naam} - nog 1 kwartier tekort")
        else:
            ws_feedback2.cell(row_fb2, 1, f"{naam} - nog {resterend} kwartieren tekort")
        row_fb2 += 1

    row_fb2 += 1

# -----------------------------------
# kolombreedte en opmaak
# -----------------------------------
ws_feedback2.column_dimensions["A"].width = 45

for row in ws_feedback2.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

# PART 6 6666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666
# PART 6 666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666

# -----------------------------
# DEEL 6: Wissels detecteren, classificeren en exporteren
# -----------------------------

from collections import defaultdict, deque
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------
# Helpers
# -----------------------------
def build_student_per_hour_map(assigned_map):
    student_per_uur = defaultdict(dict)
    for (uur, attr), namen in assigned_map.items():
        for naam in namen:
            student_per_uur[naam][uur] = attr
    return student_per_uur


def extract_hourly_changes(student_per_uur, open_uren):
    """
    Bouw per uur alle veranderingen op:
    - newcomers: studenten die op dit uur starten
    - movers: studenten die op dit uur van attractie wisselen
    - leavers: studenten die vorig uur wel werkten en nu niet meer
    - disappearing_sources: attractieplekken die verdwijnen tussen vorig uur en dit uur
    """
    changes_per_hour = {}

    def capaciteit_op_uur(uur, attr):
        if uur not in open_uren:
            return 0
        return max(0, aantallen.get(uur, {}).get(attr, 0))

    all_attrs = set()
    for uur2 in open_uren:
        all_attrs.update(aantallen.get(uur2, {}).keys())

    for uur in sorted(open_uren):
        prev_uur = uur - 1

        prev_students = {}
        curr_students = {}

        for naam, uren_dict in student_per_uur.items():
            if prev_uur in uren_dict:
                prev_students[naam] = uren_dict[prev_uur]
            if uur in uren_dict:
                curr_students[naam] = uren_dict[uur]

        newcomers = []
        movers = []
        leavers = []

        for naam, curr_attr in curr_students.items():
            if naam not in prev_students:
                newcomers.append({
                    "naam": naam,
                    "naar": curr_attr
                })
            else:
                prev_attr = prev_students[naam]
                if prev_attr != curr_attr:
                    movers.append({
                        "naam": naam,
                        "van": prev_attr,
                        "naar": curr_attr,
                        "uur": uur,
                        "type": "normaal"
                    })

        for naam, prev_attr in prev_students.items():
            if naam not in curr_students:
                leavers.append({
                    "naam": naam,
                    "van": prev_attr
                })

        disappearing_sources = []
        if prev_uur in open_uren:
            for attr in sorted(all_attrs):
                prev_cap = capaciteit_op_uur(prev_uur, attr)
                curr_cap = capaciteit_op_uur(uur, attr)

                if curr_cap < prev_cap:
                    for pos in range(curr_cap + 1, prev_cap + 1):
                        disappearing_sources.append({
                            "attr": attr,
                            "pos": pos,
                            "reason": "capacity_drop"
                        })

        changes_per_hour[uur] = {
            "newcomers": newcomers,
            "movers": movers,
            "leavers": leavers,
            "disappearing_sources": disappearing_sources
        }

    return changes_per_hour




def classify_hourly_switches(uur, newcomers, movers, leavers=None, disappearing_sources=None):
    """
    Types:
    - volledig automatisch:
        een nieuwkomer komt toe op attractie A,
        daardoor kan een student van A weg,
        waardoor ketting verder loopt
    - half-automatisch:
        een ketting die start vanuit een verdwijnende plek
        of een logisch vervolg daarop is
    - normaal:
        losse wissels of resterende lussen zonder duidelijk startpunt

    Belangrijk:
    - de eerste edge van een ketting krijgt 'half-start'
    - de rest krijgt 'half-automatisch'
    - losse enkele wissels blijven 'normaal'

    Extra regels:
    - bij echte ronde lussen kiezen we het startpunt liefst op een attractie
      met 2 plekken op dit uur
    - niet-ronde kettingen komen vóór ronde lussen in de output
    - groene wissels starten altijd bij de attractie waar de nieuwkomer toekomt
    - enkel kettingen met lengte > 1 komen in de half-automatische output
      zodat er geen dubbels ontstaan
    """
    if not movers:
        return []

    if leavers is None:
        leavers = []

    if disappearing_sources is None:
        disappearing_sources = []

    # -----------------------------
    # Helpers
    # -----------------------------
    def stable_edge_key(edge):
        return (edge["van"], edge["naar"], edge["naam"])

    def next_edge_key(edge):
        return (edge["naar"], edge["naam"])

    def has_two_spots(attr):
        try:
            return aantallen[uur].get(attr, 1) >= 2
        except Exception:
            return False

    def roll_chain_from_start_edge(start_edge, edge_pool, used_ids):
        chain = []
        current = start_edge

        while current and current["id"] not in used_ids:
            chain.append(current)
            used_ids.add(current["id"])

            next_candidates = [
                e for e in edge_pool
                if e["id"] not in used_ids and e["van"] == current["naar"]
            ]
            next_candidates.sort(key=next_edge_key)
            current = next_candidates[0] if next_candidates else None

        return chain

    def classify_chain_shape(chain):
        """
        Geeft terug:
        - 'open' als begin en einde verschillen
        - 'cycle' als begin en einde terug sluiten
        """
        if len(chain) <= 1:
            return "single"

        eerste_van = chain[0]["van"]
        laatste_naar = chain[-1]["naar"]

        if eerste_van == laatste_naar:
            return "cycle"
        return "open"

    def add_chain_record_if_needed(chain_records, chain):
        """
        Enkel echte kettingen (lengte > 1) komen in chain_records.
        Singles blijven 'normaal' en worden later via normal_edges getoond.
        """
        if not chain:
            return

        shape = classify_chain_shape(chain)

        if len(chain) == 1:
            chain[0]["type"] = "normaal"
            return

        chain[0]["type"] = "half-start"
        for e in chain[1:]:
            e["type"] = "half-automatisch"

        chain_records.append({
            "shape": shape,
            "start_has_two_spots": has_two_spots(chain[0]["van"]),
            "edges": chain
        })

    # -----------------------------
    # Edges opbouwen
    # -----------------------------
    edges = []
    for idx, m in enumerate(movers):
        edges.append({
            "id": idx,
            "naam": m["naam"],
            "van": m["van"],
            "naar": m["naar"],
            "uur": uur,
            "type": "normaal"
        })

    # -----------------------------
    # Maps
    # -----------------------------
    outgoing = defaultdict(list)
    incoming = defaultdict(list)

    for e in edges:
        outgoing[e["van"]].append(e)
        incoming[e["naar"]].append(e)

    for attr in outgoing:
        outgoing[attr].sort(key=next_edge_key)
    for attr in incoming:
        incoming[attr].sort(key=lambda x: (x["van"], x["naam"]))

    # -----------------------------
    # 1. Volledig automatische kettingen
    # -----------------------------
    newcomers_by_attr = defaultdict(list)
    for n in newcomers:
        newcomers_by_attr[n["naar"]].append(n["naam"])

    auto_edge_ids = set()
    queue = deque()

    # Groen start ALTIJD bij de attractie waar de nieuwkomer toekomt
    # De nieuwkomer zet daar de ketting in gang.
    for attr in newcomers_by_attr.keys():
        for e in outgoing.get(attr, []):
            if e["id"] not in auto_edge_ids:
                auto_edge_ids.add(e["id"])
                queue.append(e)

    while queue:
        current = queue.popleft()
        next_attr = current["naar"]

        for next_edge in outgoing.get(next_attr, []):
            if next_edge["id"] not in auto_edge_ids:
                auto_edge_ids.add(next_edge["id"])
                queue.append(next_edge)

    for e in edges:
        if e["id"] in auto_edge_ids:
            e["type"] = "volledig automatisch"

    # -----------------------------
    # 2. Resterende edges
    # -----------------------------
    remaining_edges = [e for e in edges if e["id"] not in auto_edge_ids]

    if not remaining_edges:
        auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]

        ordered_auto = []
        used_auto = set()

        # Volg exact de volgorde van newcomers, niet alfabetisch op attractie
        for newcomer in newcomers:
            start_attr = newcomer["naar"]

            start_candidates = [
                e for e in auto_edges
                if e["id"] not in used_auto and e["van"] == start_attr
            ]
            start_candidates.sort(key=next_edge_key)

            for start in start_candidates:
                current = start
                while current and current["id"] not in used_auto:
                    ordered_auto.append(current)
                    used_auto.add(current["id"])

                    next_candidates = [
                        e for e in auto_edges
                        if e["id"] not in used_auto and e["van"] == current["naar"]
                    ]
                    next_candidates.sort(key=next_edge_key)
                    current = next_candidates[0] if next_candidates else None

        leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
        leftovers_auto.sort(key=stable_edge_key)
        ordered_auto.extend(leftovers_auto)

        return ordered_auto

    source_attrs = [x["attr"] for x in disappearing_sources]

    chain_records = []
    used_ids = set()

    # -----------------------------
    # 3. Eerst kettingen vanuit verdwijnende plekken
    # -----------------------------
    for start_attr in source_attrs:
        start_candidates = [
            e for e in remaining_edges
            if e["id"] not in used_ids and e["van"] == start_attr
        ]
        start_candidates.sort(key=stable_edge_key)

        for start_edge in start_candidates:
            if start_edge["id"] in used_ids:
                continue

            chain = roll_chain_from_start_edge(start_edge, remaining_edges, used_ids)
            add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 4. Restjes groeperen in componenten
    # -----------------------------
    leftovers = [e for e in remaining_edges if e["id"] not in used_ids]

    if leftovers:
        remaining_by_id = {e["id"]: e for e in leftovers}
        adjacency = defaultdict(set)

        for e1 in leftovers:
            for e2 in leftovers:
                if e1["id"] == e2["id"]:
                    continue
                if e1["naar"] == e2["van"] or e2["naar"] == e1["van"]:
                    adjacency[e1["id"]].add(e2["id"])
                    adjacency[e2["id"]].add(e1["id"])

        visited = set()
        components = []

        for e in leftovers:
            if e["id"] in visited:
                continue

            stack = [e["id"]]
            comp_ids = []

            while stack:
                curr = stack.pop()
                if curr in visited:
                    continue
                visited.add(curr)
                comp_ids.append(curr)

                for nb in adjacency[curr]:
                    if nb not in visited:
                        stack.append(nb)

            components.append([remaining_by_id[i] for i in comp_ids])

        for comp_edges in components:
            if not comp_edges:
                continue

            comp_used = set()

            start_candidates = []
            for e in comp_edges:
                has_prev = any(
                    other["id"] != e["id"] and other["naar"] == e["van"]
                    for other in comp_edges
                )
                if not has_prev:
                    start_candidates.append(e)

            # ---------------------------------
            # NIET-RONDE KETTINGEN
            # ---------------------------------
            if start_candidates:
                start_candidates.sort(key=stable_edge_key)

                for start_edge in start_candidates:
                    if start_edge["id"] in comp_used:
                        continue

                    chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)

                for edge in rest:
                    if edge["id"] in comp_used:
                        continue
                    chain = roll_chain_from_start_edge(edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

            # ---------------------------------
            # ECHTE RONDE LUS
            # ---------------------------------
            else:
                two_spot_candidates = [e for e in comp_edges if has_two_spots(e["van"])]

                if two_spot_candidates:
                    two_spot_candidates.sort(key=stable_edge_key)
                    start_edge = two_spot_candidates[0]
                else:
                    comp_edges.sort(key=stable_edge_key)
                    start_edge = comp_edges[0]

                chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)
                chain.extend(rest)

                add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 5. Definitieve volgorde
    # -----------------------------
    auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]
    normal_edges = [e for e in edges if e["type"] == "normaal"]

    ordered_auto = []
    used_auto = set()

    # Groen start ALTIJD vanuit de attractie van de nieuwkomer
    # en volgt dan pas de ketting verder.
    for newcomer in newcomers:
        start_attr = newcomer["naar"]

        start_candidates = [
            e for e in auto_edges
            if e["id"] not in used_auto and e["van"] == start_attr
        ]
        start_candidates.sort(key=next_edge_key)

        for start in start_candidates:
            current = start
            while current and current["id"] not in used_auto:
                ordered_auto.append(current)
                used_auto.add(current["id"])

                next_candidates = [
                    e for e in auto_edges
                    if e["id"] not in used_auto and e["van"] == current["naar"]
                ]
                next_candidates.sort(key=next_edge_key)
                current = next_candidates[0] if next_candidates else None

    leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
    leftovers_auto.sort(key=stable_edge_key)
    ordered_auto.extend(leftovers_auto)

    # niet-ronde kettingen eerst, dan ronde lussen
    chain_records.sort(
        key=lambda rec: (
            0 if rec["shape"] == "open" else 1,
            0 if rec["shape"] == "cycle" and rec["start_has_two_spots"] else 1,
            stable_edge_key(rec["edges"][0]) if rec["edges"] else ("", "", "")
        )
    )

    ordered_half = []
    for rec in chain_records:
        ordered_half.extend(rec["edges"])

    normal_edges.sort(key=stable_edge_key)

    # Extra veiligheid tegen dubbels
    seen_ids = set()
    final_order = []

    for e in ordered_auto + ordered_half + normal_edges:
        if e["id"] in seen_ids:
            continue
        seen_ids.add(e["id"])
        final_order.append(e)

    return final_order



# -----------------------------
# Stap 1: student → uur → attractie
# -----------------------------
student_per_uur = build_student_per_hour_map(assigned_map)

# -----------------------------
# Stap 2: veranderingen per uur opbouwen
# -----------------------------
changes_per_hour = extract_hourly_changes(student_per_uur, open_uren)

# -----------------------------
# Stap 3: per uur classificeren en ordenen
# -----------------------------
wissels_per_uur = {}

for uur in sorted(open_uren):
    newcomers = changes_per_hour[uur]["newcomers"]
    movers = changes_per_hour[uur]["movers"]
    leavers = changes_per_hour[uur]["leavers"]
    disappearing_sources = changes_per_hour[uur]["disappearing_sources"]

    ordered_switches = classify_hourly_switches(
        uur,
        newcomers,
        movers,
        leavers,
        disappearing_sources
    )

    if ordered_switches:
        wissels_per_uur[uur] = ordered_switches


# KPI berekenen
totaal_wissels = 0
aantal_auto = 0

for uur in wissels_per_uur:
    for w in wissels_per_uur[uur]:
        totaal_wissels += 1
        if w["type"] == "volledig automatisch":
            aantal_auto += 1

niet_groen = totaal_wissels - aantal_auto

# -----------------------------
# Stap 4: werkblad "Wissels" maken
# -----------------------------
ws_wissels = wb_out.create_sheet(title="Wissels")

# -----------------------------
# KPI rechts van de tabel (kolom G)
# -----------------------------
ws_wissels.cell(1, 7, "KPI Wissels").font = Font(bold=True)

ws_wissels.cell(2, 7, "Totaal wissels:")
ws_wissels.cell(2, 8, totaal_wissels)

ws_wissels.cell(3, 7, "Volledig automatisch:")
ws_wissels.cell(3, 8, aantal_auto)

ws_wissels.cell(4, 7, "Niet-groen (KPI):")
ws_wissels.cell(4, 8, niet_groen)
ws_wissels.cell(4, 8).font = Font(bold=True)

center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

current_row = 1


for uur in sorted(wissels_per_uur.keys()):
    # Titelrij per uur
    title_cell = ws_wissels.cell(current_row, 1, f"Wissels om {uur}:00")
    title_cell.font = Font(bold=True)
    title_cell.alignment = center_align
    current_row += 1

    # Headers
    headers = ["Student", "Van", "Naar"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_wissels.cell(current_row, col_idx, header)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1

    # Wissels
    for w in wissels_per_uur[uur]:
        ws_wissels.cell(current_row, 1, w["naam"])
        ws_wissels.cell(current_row, 2, w["van"])
        ws_wissels.cell(current_row, 3, w["naar"])

        # Basis layout
        for col_idx in range(1, 4):
            cell = ws_wissels.cell(current_row, col_idx)
            cell.alignment = center_align
            cell.border = thin_border

        # Kleuren enkel op kolom B en C
        if w["type"] == "volledig automatisch":
            ws_wissels.cell(current_row, 2).fill = green_fill
            ws_wissels.cell(current_row, 3).fill = green_fill

        elif w["type"] == "half-automatisch":
            ws_wissels.cell(current_row, 2).fill = yellow_fill
            ws_wissels.cell(current_row, 3).fill = yellow_fill

        elif w["type"] == "half-start":
            ws_wissels.cell(current_row, 2).fill = orange_fill
            ws_wissels.cell(current_row, 3).fill = orange_fill

        current_row += 1

    # Lege rij tussen uren
    current_row += 1

# -----------------------------
# Stap 5: kolombreedtes
# -----------------------------
breedtes = {
    1: 22,
    2: 25,
    3: 25,
    7: 24,
    8: 18
}

for col_idx, breedte in breedtes.items():
    ws_wissels.column_dimensions[get_column_letter(col_idx)].width = breedte

#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# -----------------------------
# Werkbladen altijd verbergen
# -----------------------------
for bladnaam in ["Pauzevlinders", "Feedback"]:
    if bladnaam in wb_out.sheetnames:
        ws_hide = wb_out[bladnaam]
        ws_hide.sheet_state = "veryHidden" 

#ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


# -----------------------------
# Opslaan in hetzelfde unieke bestand als DEEL 3
# -----------------------------
output = BytesIO()
wb_out.save(output)
output.seek(0)
# st.success("Planning gegenereerd!")
st.download_button(
    "Download planning",
    data=output.getvalue(),
    file_name=f"Planning_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)
